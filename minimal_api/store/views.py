# Standard library
import io
from collections import defaultdict
from datetime import datetime, timedelta
from decimal import Decimal

# Third-party libraries
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle

# Django imports
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum, Count, F, Q, DecimalField, ExpressionWrapper
from django.db.models.functions import Coalesce, TruncMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone

# Local imports - only models
from .models import (
    Product, Sale, Receipt, Payment, PaymentMethod,
    Customer, TaxConfiguration
)

# =====================================================
# CONSTANTS - Defined locally to avoid import errors
# =====================================================

SHOP_TYPES = [
    ('STORE', 'Store (Shop Floor)'),
    ('WAREHOUSE', 'Warehouse'),
]

PAYMENT_METHODS = [
    ('cash', 'Cash'),
    ('transfer', 'Bank Transfer'),
    ('pos', 'POS'),
    ('credit', 'Credit'),
]

# =====================================================
# HELPER FUNCTIONS
# =====================================================

def is_md(user):
    """Check if user is MD (Managing Director)"""
    if user.is_superuser:
        return True
    if hasattr(user, 'profile'):
        return user.profile.access_level == 'md'
    return False


# =====================================================
# REPORT VIEWS
# =====================================================

def sales_report(request):
    """Sales report view with filtering and export options"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    payment_method_filter = request.GET.get('payment_method')
    shop_type = request.GET.get('shop_type')
    export_format = request.GET.get('export')

    # Base queryset
    sales = Sale.objects.select_related(
        'product', 'receipt', 'payment', 'delivery', 'customer'
    ).prefetch_related(
        'payment__payment_methods'
    ).order_by('-sale_date')

    # Apply date filters
    start_date_obj = None
    end_date_obj = None
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            sales = sales.filter(sale_date__range=[start_date_obj, end_date_obj])
        except ValueError:
            start_date_obj = None
            end_date_obj = None

    # Apply shop type filter
    if shop_type:
        sales = sales.filter(product__shop=shop_type)

    # Apply payment method filter
    if payment_method_filter:
        sales = sales.filter(payment__payment_methods__payment_method=payment_method_filter).distinct()

    # Group sales by receipt
    grouped_sales = defaultdict(list)
    for sale in sales:
        grouped_sales[sale.receipt].append(sale)

    # Convert to sorted list
    grouped_sales = sorted(
        grouped_sales.items(),
        key=lambda x: x[1][0].sale_date if x[1] else timezone.now(),
        reverse=True
    )

    # Calculate total sales
    unique_receipts = Receipt.objects.filter(sales__in=sales).distinct()
    total_sales = sum(receipt.total_with_delivery for receipt in unique_receipts if receipt.total_with_delivery)

    # Handle exports
    if export_format == 'excel':
        return export_sales_to_excel(grouped_sales, start_date_obj, end_date_obj, total_sales)
    elif export_format == 'pdf':
        return export_sales_to_pdf(grouped_sales, start_date_obj, end_date_obj, total_sales)

    # Return JSON response for API
    return JsonResponse({
        'status': 'success',
        'data': {
            'sales': [
                {
                    'receipt_number': receipt.receipt_number,
                    'receipt_total': float(receipt.total_with_delivery) if receipt.total_with_delivery else 0,
                    'date': sale_list[0].sale_date.strftime('%Y-%m-%d %H:%M:%S'),
                    'items': [
                        {
                            'product': sale.product.brand,
                            'quantity': sale.quantity,
                            'total_price': float(sale.total_price),
                            'payment_method': ', '.join([pm.get_payment_method_display() for pm in sale.payment.payment_methods.all()]) if sale.payment else ''
                        }
                        for sale in sale_list
                    ]
                }
                for receipt, sale_list in grouped_sales
            ],
            'total_sales': float(total_sales),
            'filters': {
                'start_date': start_date,
                'end_date': end_date,
                'payment_method': payment_method_filter,
                'shop_type': shop_type
            }
        }
    })


def financial_report(request):
    """Financial report view"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    export_format = request.GET.get('export')

    # Base queryset
    sales = Sale.objects.select_related('product', 'payment', 'receipt')

    # Apply date filters
    start_date_obj = None
    end_date_obj = None
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            sales = sales.filter(sale_date__range=[start_date_obj, end_date_obj])
        except ValueError:
            pass

    # Calculate metrics
    total_revenue = sales.aggregate(
        total=Coalesce(Sum('total_price'), Decimal('0'))
    )['total']

    total_cost = sales.aggregate(
        total=Coalesce(Sum(
            ExpressionWrapper(
                F('product__price') * F('quantity'),
                output_field=DecimalField()
            )
        ), Decimal('0'))
    )['total']

    profit = total_revenue - total_cost
    profit_margin = (profit / total_revenue * 100) if total_revenue > 0 else 0

    # Payment method breakdown
    payment_breakdown = {}
    for pm_code, pm_display in PAYMENT_METHODS:
        pm_total = sales.filter(
            payment__payment_methods__payment_method=pm_code
        ).distinct().aggregate(
            total=Coalesce(Sum('total_price'), Decimal('0'))
        )['total']
        payment_breakdown[pm_display] = float(pm_total)

    # Monthly trends
    monthly_sales = sales.annotate(
        month=TruncMonth('sale_date')
    ).values('month').annotate(
        revenue=Coalesce(Sum('total_price'), Decimal('0')),
        count=Count('id')
    ).order_by('month')

    # Return JSON response for API
    return JsonResponse({
        'status': 'success',
        'data': {
            'summary': {
                'total_revenue': float(total_revenue),
                'total_cost': float(total_cost),
                'profit': float(profit),
                'profit_margin': float(profit_margin),
            },
            'payment_breakdown': payment_breakdown,
            'monthly_trends': [
                {
                    'month': item['month'].strftime('%Y-%m'),
                    'revenue': float(item['revenue']),
                    'count': item['count']
                }
                for item in monthly_sales
            ],
            'filters': {
                'start_date': start_date,
                'end_date': end_date
            }
        }
    })


def inventory_report(request):
    """Inventory report view"""

    # Get all products with stock info
    products = Product.objects.all().order_by('brand')

    # Calculate metrics
    total_products = products.count()
    total_stock_value = sum(
        (product.price or 0) * product.quantity
        for product in products
    )
    low_stock_count = products.filter(quantity__lt=10).count()
    out_of_stock_count = products.filter(quantity=0).count()

    # Return JSON response for API
    return JsonResponse({
        'status': 'success',
        'data': {
            'summary': {
                'total_products': total_products,
                'total_stock_value': float(total_stock_value),
                'low_stock_count': low_stock_count,
                'out_of_stock_count': out_of_stock_count,
            },
            'products': [
                {
                    'brand': product.brand,
                    'category': product.category,
                    'quantity': product.quantity,
                    'price': float(product.price or 0),
                    'selling_price': float(product.selling_price or 0),
                    'shop': product.shop,
                    'location': product.location,
                }
                for product in products
            ]
        }
    })


def tax_report(request):
    """Tax report view"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Get active tax configurations
    active_taxes = TaxConfiguration.objects.filter(is_active=True)

    # Base queryset for sales
    sales = Sale.objects.select_related('product', 'receipt')

    # Apply date filters
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            sales = sales.filter(sale_date__range=[start_date_obj, end_date_obj])
        except ValueError:
            pass

    # Calculate tax totals
    total_sales = sales.aggregate(
        total=Coalesce(Sum('total_price'), Decimal('0'))
    )['total']

    tax_breakdown = []
    total_tax = Decimal('0')

    for tax in active_taxes:
        tax_amount = total_sales * (tax.rate / 100)
        total_tax += tax_amount
        tax_breakdown.append({
            'name': tax.name,
            'rate': float(tax.rate),
            'amount': float(tax_amount),
            'is_inclusive': tax.is_inclusive,
        })

    # Return JSON response for API
    return JsonResponse({
        'status': 'success',
        'data': {
            'summary': {
                'total_sales': float(total_sales),
                'total_tax': float(total_tax),
                'net_sales': float(total_sales - total_tax),
            },
            'tax_breakdown': tax_breakdown,
            'filters': {
                'start_date': start_date,
                'end_date': end_date
            }
        }
    })


def reports_dashboard(request):
    """Reports dashboard with key metrics"""

    # Get date range (default to last 30 days)
    end_date = timezone.now()
    start_date = end_date - timedelta(days=30)

    # Sales metrics
    recent_sales = Sale.objects.filter(
        sale_date__range=[start_date, end_date]
    ).select_related('product', 'receipt')

    total_revenue = recent_sales.aggregate(
        total=Coalesce(Sum('total_price'), Decimal('0'))
    )['total']

    sales_count = recent_sales.count()

    # Product metrics
    total_products = Product.objects.count()
    low_stock = Product.objects.filter(quantity__lt=10).count()

    # Customer metrics
    total_customers = Customer.objects.count()
    active_customers = recent_sales.values('customer').distinct().count()

    # Return JSON response for API
    return JsonResponse({
        'status': 'success',
        'data': {
            'sales': {
                'total_revenue': float(total_revenue),
                'sales_count': sales_count,
                'average_sale': float(total_revenue / sales_count) if sales_count > 0 else 0,
            },
            'inventory': {
                'total_products': total_products,
                'low_stock_count': low_stock,
            },
            'customers': {
                'total_customers': total_customers,
                'active_customers': active_customers,
            },
            'date_range': {
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
            }
        }
    })


# =====================================================
# EXPORT HELPER FUNCTIONS
# =====================================================

def export_sales_to_pdf(grouped_sales, start_date, end_date, total_sales):
    """Export sales report to PDF"""
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(1 * inch, height - 1 * inch, "Sales Report")

    # Date range
    p.setFont("Helvetica", 12)
    date_text = "All Dates"
    if start_date and end_date:
        date_text = f"From {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    p.drawString(1 * inch, height - 1.2 * inch, date_text)

    # Table data
    data = [['Receipt', 'Item', 'Qty', 'Item Total', 'Receipt Total', 'Payment', 'Date']]
    receipt_header_rows = []
    row_index = 1

    for receipt, sale_list in grouped_sales:
        sale = sale_list[0]
        receipt_total = receipt.total_with_delivery if receipt.total_with_delivery else 0

        # Receipt header row
        data.append([
            f"Receipt #{receipt.receipt_number}",
            "", "", "",
            f"₦{receipt_total:.2f}",
            "",
            sale.sale_date.strftime("%m/%d %H:%M")
        ])
        receipt_header_rows.append(row_index)
        row_index += 1

        # Items
        for item in sale_list:
            payment_method = ""
            if item.payment:
                payment_methods = item.payment.payment_methods.all()
                payment_method = ", ".join([pm.get_payment_method_display() for pm in payment_methods])

            data.append([
                "",
                item.product.brand,
                str(item.quantity),
                f"₦{item.total_price:.2f}",
                "",
                payment_method,
                ""
            ])
            row_index += 1

        # Spacer
        data.append(["", "", "", "", "", "", ""])
        row_index += 1

    # Total row
    data.append([
        "TOTAL SALES", "", "", "",
        f"₦{total_sales:.2f}",
        "", ""
    ])

    # Create table
    col_widths = [90, 120, 35, 70, 70, 90, 65]
    table = Table(data, colWidths=col_widths, repeatRows=1)

    # Style
    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ]

    for row in receipt_header_rows:
        style_commands.append(('FONTNAME', (0, row), (6, row), 'Helvetica-Bold'))
        style_commands.append(('BACKGROUND', (0, row), (6, row), colors.beige))
        style_commands.append(('BOTTOMPADDING', (0, row), (6, row), 6))

    table.setStyle(TableStyle(style_commands))

    # Draw table
    table.wrapOn(p, width, height)
    y = height - 2 * inch
    table.drawOn(p, 0.5 * inch, y - table._height)

    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


def export_sales_to_excel(grouped_sales, start_date, end_date, total_sales):
    """Export sales report to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Title
    ws['A1'] = "Sales Report"
    ws['A1'].font = Font(bold=True, size=14)

    # Date range
    date_text = "All Dates"
    if start_date and end_date:
        date_text = f"From {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    ws['A2'] = date_text

    # Headers
    headers = ['Receipt', 'Item', 'Qty', 'Item Total', 'Receipt Total', 'Payment', 'Date']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Data
    row = 5
    for receipt, sale_list in grouped_sales:
        sale = sale_list[0]
        receipt_total = receipt.total_with_delivery if receipt.total_with_delivery else 0

        # Receipt header
        ws.cell(row=row, column=1, value=f"Receipt #{receipt.receipt_number}")
        ws.cell(row=row, column=5, value=receipt_total)
        ws.cell(row=row, column=7, value=sale.sale_date.strftime("%Y-%m-%d %H:%M"))

        for col in range(1, 8):
            ws.cell(row=row, column=col).font = Font(bold=True)
        row += 1

        # Items
        for item in sale_list:
            payment_method = ""
            if item.payment:
                payment_methods = item.payment.payment_methods.all()
                payment_method = ", ".join([pm.get_payment_method_display() for pm in payment_methods])

            ws.cell(row=row, column=2, value=item.product.brand)
            ws.cell(row=row, column=3, value=item.quantity)
            ws.cell(row=row, column=4, value=float(item.total_price))
            ws.cell(row=row, column=6, value=payment_method)
            row += 1

        row += 1  # Spacer

    # Total
    ws.cell(row=row, column=1, value="TOTAL SALES")
    ws.cell(row=row, column=5, value=float(total_sales))
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=5).font = Font(bold=True)

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 18

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


# =====================================================
# DUMMY VIEWS FOR REQUIRED URL PATTERNS
# =====================================================

def access_denied(request):
    """Access denied page"""
    return JsonResponse({
        'status': 'error',
        'message': 'Access denied. You do not have permission to access this resource.'
    }, status=403)


def login_view(request):
    """Login view - stub for minimal API"""
    return JsonResponse({
        'status': 'info',
        'message': 'This is a minimal API endpoint. Please authenticate using your API client.'
    })


def logout_view(request):
    """Logout view - stub for minimal API"""
    return JsonResponse({
        'status': 'success',
        'message': 'Logged out successfully.'
    })
