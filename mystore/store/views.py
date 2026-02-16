2# Standard library
import hashlib
import io
import json
import logging
import threading
import time
import urllib.request  # ✅ Added urllib
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal

# Django settings
from django.conf import settings

# Third-party libraries
import barcode
import openpyxl
import pandas as pd
from barcode.writer import ImageWriter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
)
from reportlab.lib.utils import ImageReader
from weasyprint import HTML, CSS
import win32print  # Windows-specific — consider wrapping in try/except if cross-platform needed

# Django imports
from django import template
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.sites.shortcuts import get_current_site
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.core.mail import EmailMessage
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db import models, transaction
from django.db.models import (
    Q, F, Sum, Avg, Count, FloatField, DecimalField, ExpressionWrapper
)
from django.db.models.functions import (
    Coalesce, TruncMonth, TruncWeek, TruncDay
)
from django.forms import formset_factory
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string, get_template
from django.templatetags.static import static
from django.urls import reverse
from django.utils import timezone
from django.utils.timezone import make_aware
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_http_methods
from django_redis import get_redis_connection

# Local app imports
from .choices import ProductChoices
from .forms import (
    ProductForm, PreOrderForm, InvoiceForm, GoodsReceivedForm, DeliveryForm,
    CustomerForm, SaleForm, PreOrderStatusForm, PaymentForm,
    PaymentValidationForm, PaymentMethodFormSet, LocationTransferForm,
    TransferItemForm, ExcelUploadForm, CustomUserCreationForm,
    UserEditForm, UserProfileForm
)
from .models import (
    Product, PreOrder, Invoice, GoodsReceived, Delivery, Customer, Sale,
    InvoiceProduct, Receipt, Payment, PaymentMethod, LocationTransfer,
    ProductHistory, TransferItem, UserProfile, ActivityLog, StoreConfiguration,
    LoyaltyConfiguration, LoyaltyTransaction, TaxConfiguration
)
from .utils import (
    get_cached_choices,
    get_product_stats,
    get_location_cached_choices,
)
import csv

# Re-export or alias commonly used types (optional)
from io import BytesIO  # ✅ Added BytesIO

# Logger
logger = logging.getLogger(__name__)




register = template.Library()

@register.filter
def add_class(field, css_class):
    return field.as_widget(attrs={"class": css_class})


def is_md(user):
    if not user.is_authenticated:
        return False
    return user.is_staff  # Adjust this based on your admin check logic


def access_denied(request):
    return render(request, 'access_denied.html')


def is_cashier(user):
    return user.groups.filter(name='Cashier').exists()


def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                # Log successful login
                ActivityLog.log_activity(
                    user=user,
                    action='login',
                    description=f'User {username} logged in successfully',
                    request=request
                )
                return redirect('homepage')  # Replace with your success URL
            else:
                # Log failed login attempt
                ActivityLog.log_activity(
                    user=None,
                    action='failed_login',
                    description=f'Failed login attempt for username: {username}',
                    success=False,
                    request=request
                )
                messages.error(request, 'Invalid username or password. Please try again.')
        else:
            # Form validation errors (username/password format errors)
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, error)
    else:
        form = AuthenticationForm()

    return render(request, 'loginout/login.html', {'form': form})


def logout_view(request):
    # Log logout before actually logging out
    if request.user.is_authenticated:
        ActivityLog.log_activity(
            user=request.user,
            action='logout',
            description=f'User {request.user.username} logged out',
            request=request
        )
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect('login')  # Redirect to login page


@login_required(login_url='login')
def homepage(request):
    return render(request, 'homepage.html')


def user_required_access(access_levels):
    """Decorator to check if user has required access level"""

    def decorator(view_func):
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                messages.error(request, "You need to be logged in.")
                return redirect('login')

            try:
                user_profile = request.user.profile
                if user_profile.access_level not in access_levels:
                    messages.error(request, "You don't have permission to access this page.")
                    return redirect('dashboard')  # Redirect to your dashboard
            except UserProfile.DoesNotExist:
                messages.error(request, "User profile not found.")
                return redirect('login')

            return view_func(request, *args, **kwargs)

        return _wrapped_view

    return decorator


@login_required
@user_required_access(['md'])
def user_management_dashboard(request):
    """Main user management dashboard - Only MD can access"""
    users = User.objects.select_related('profile').all().order_by('-date_joined')

    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        users = users.filter(
            username__icontains=search_query
        ) | users.filter(
            first_name__icontains=search_query
        ) | users.filter(
            last_name__icontains=search_query
        )

    # Filter by access level
    access_filter = request.GET.get('access_level')
    if access_filter:
        users = users.filter(profile__access_level=access_filter)

    # Pagination
    paginator = Paginator(users, 10)  # Show 10 users per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'users': page_obj,
        'search_query': search_query,
        'access_filter': access_filter,
        'access_levels': UserProfile.ACCESS_LEVEL_CHOICES,
        'total_users': users.count(),
    }

    return render(request, 'user_management/dashboard.html', context)


@login_required(login_url='login')
def create_customer(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save()
            # Log customer creation
            ActivityLog.log_activity(
                user=request.user,
                action='customer_create',
                description=f'Created customer: {customer.name} - {customer.phone_number}',
                model_name='Customer',
                object_id=customer.id,
                object_repr=str(customer),
                request=request
            )
            return redirect('customer_list')
    else:
        form = CustomerForm()
    return render(request, 'customer/create_customer.html', {'form': form})


@login_required
@user_required_access(['md'])
def create_user(request):
    """Create new user - Only MD can access"""
    if request.method == 'POST':
        user_form = CustomUserCreationForm(request.POST)

        if user_form.is_valid():
            try:
                with transaction.atomic():
                    # Create user
                    user = user_form.save()
                    access_level = user_form.cleaned_data['access_level']

                    # Create user profile
                    UserProfile.objects.create(
                        user=user,
                        access_level=access_level,
                        phone_number=user_form.cleaned_data.get('phone_number', ''),
                        created_by=request.user
                    )

                    # Assign Django groups based on access level
                    from django.contrib.auth.models import Group

                    # Clear existing groups
                    user.groups.clear()

                    # Assign appropriate group
                    if access_level == 'md':
                        user.is_staff = True
                        user.save()
                        # MD gets both MD access (via is_staff) and can be in a group
                        group, _ = Group.objects.get_or_create(name='MD')
                        user.groups.add(group)
                    elif access_level == 'cashier':
                        group, _ = Group.objects.get_or_create(name='Cashier')
                        user.groups.add(group)
                    elif access_level == 'accountant':
                        group, _ = Group.objects.get_or_create(name='Accountant')
                        user.groups.add(group)

                    # Log user creation
                    ActivityLog.log_activity(
                        user=request.user,
                        action='user_create',
                        description=f'Created user: {user.username} ({user.get_full_name()}) - Access level: {access_level}',
                        model_name='User',
                        object_id=user.id,
                        object_repr=user.username,
                        extra_data={'access_level': access_level},
                        request=request
                    )

                    messages.success(request, f'User "{user.username}" created successfully with {access_level} access!')
                    return redirect('user_management_dashboard')
            except Exception as e:
                messages.error(request, f'Error creating user: {str(e)}')
    else:
        user_form = CustomUserCreationForm()

    return render(request, 'user_management/create_user.html', {
        'user_form': user_form
    })


@login_required
@user_required_access(['md'])
def edit_user(request, user_id):
    """Edit user - Only MD can access"""
    user = get_object_or_404(User, id=user_id)

    # Don't allow editing superuser accounts unless current user is superuser
    if user.is_superuser and not request.user.is_superuser:
        messages.error(request, "You cannot edit superuser accounts.")
        return redirect('user_management_dashboard')

    try:
        profile = user.profile
    except UserProfile.DoesNotExist:
        # Create profile if it doesn't exist
        profile = UserProfile.objects.create(user=user)

    if request.method == 'POST':
        user_form = UserEditForm(request.POST, instance=user)
        profile_form = UserProfileForm(request.POST, instance=profile)

        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            saved_profile = profile_form.save()

            # Sync Django groups based on access level
            from django.contrib.auth.models import Group
            access_level = saved_profile.access_level

            # Clear existing groups
            user.groups.clear()

            # Assign appropriate group
            if access_level == 'md':
                user.is_staff = True
                user.save()
                group, _ = Group.objects.get_or_create(name='MD')
                user.groups.add(group)
            elif access_level == 'cashier':
                user.is_staff = False
                user.save()
                group, _ = Group.objects.get_or_create(name='Cashier')
                user.groups.add(group)
            elif access_level == 'accountant':
                user.is_staff = False
                user.save()
                group, _ = Group.objects.get_or_create(name='Accountant')
                user.groups.add(group)

            # Log user update
            ActivityLog.log_activity(
                user=request.user,
                action='user_update',
                description=f'Updated user: {user.username} - Access level: {saved_profile.access_level}',
                model_name='User',
                object_id=user.id,
                object_repr=user.username,
                extra_data={'access_level': saved_profile.access_level, 'is_active': user.is_active},
                request=request
            )

            messages.success(request, f'User "{user.username}" updated successfully!')
            return redirect('user_management_dashboard')
    else:
        user_form = UserEditForm(instance=user)
        profile_form = UserProfileForm(instance=profile)

    return render(request, 'user_management/edit_user.html', {
        'user_form': user_form,
        'profile_form': profile_form,
        'user': user
    })


@login_required
@user_required_access(['md'])
def toggle_user_status(request, user_id):
    """Toggle user active status via AJAX"""
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id)

        # Don't allow deactivating superuser accounts
        if user.is_superuser:
            return JsonResponse({'success': False, 'message': 'Cannot deactivate superuser accounts.'})

        user.is_active = not user.is_active
        user.save()

        status = "activated" if user.is_active else "deactivated"

        # Log user status change
        ActivityLog.log_activity(
            user=request.user,
            action='user_update',
            description=f'User {user.username} {status}',
            model_name='User',
            object_id=user.id,
            object_repr=user.username,
            extra_data={'is_active': user.is_active, 'action': status},
            request=request
        )

        return JsonResponse({
            'success': True,
            'message': f'User "{user.username}" has been {status}.',
            'is_active': user.is_active
        })

    return JsonResponse({'success': False, 'message': 'Invalid request method.'})


@login_required
@user_required_access(['md'])
def delete_user(request, user_id):
    """Delete user - Only MD can access"""
    user = get_object_or_404(User, id=user_id)

    # Don't allow deleting superuser accounts or self
    if user.is_superuser:
        messages.error(request, "Cannot delete superuser accounts.")
        return redirect('user_management_dashboard')

    if user == request.user:
        messages.error(request, "You cannot delete your own account.")
        return redirect('user_management_dashboard')

    if request.method == 'POST':
        username = user.username
        user.delete()
        messages.success(request, f'User "{username}" deleted successfully!')
        return redirect('user_management_dashboard')

    return render(request, 'user_management/delete_user.html', {'user': user})


@login_required
def user_profile_view(request):
    """View current user's profile"""
    try:
        profile = request.user.profile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user)

    # Get loyalty information if user is also a customer
    loyalty_info = None
    try:
        # Check if this user has an associated customer account (via email matching)
        customer = Customer.objects.filter(email=request.user.email).first()
        if customer:
            from .loyalty_utils import get_customer_loyalty_summary
            loyalty_info = get_customer_loyalty_summary(customer)
    except Exception as e:
        logger.error(f"Error fetching loyalty info for user {request.user.username}: {e}")

    return render(request, 'user_management/profile.html', {
        'user': request.user,
        'profile': profile,
        'loyalty_info': loyalty_info
    })


def is_superuser(user):
    """Check if user is a superuser"""
    return user.is_superuser


@login_required(login_url='login')
@user_passes_test(is_superuser, login_url='login')
@require_POST
@csrf_exempt
def update_product_quantity(request):
    """Update product quantity via AJAX for superusers only"""
    try:
        data = json.loads(request.body)
        product_id = data.get('product_id')
        action = data.get('action')  # 'increase' or 'decrease'

        if not product_id or not action:
            return JsonResponse({'success': False, 'error': 'Missing product_id or action'})

        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Product not found'})

        # Update quantity based on action
        if action == 'increase':
            product.quantity += 1
        elif action == 'decrease':
            if product.quantity > 0:
                product.quantity -= 1
            else:
                return JsonResponse({'success': False, 'error': 'Quantity cannot be negative'})
        else:
            return JsonResponse({'success': False, 'error': 'Invalid action'})

        # Save the product
        product.save()

        # Return updated data
        return JsonResponse({
            'success': True,
            'new_quantity': product.quantity,
            'product_brand': product.brand
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def filter_products(request, queryset):
    """Apply filters to product queryset based on request parameters"""
    query = request.GET.get('search', '')
    category = request.GET.get('category', '')
    shop = request.GET.get('shop', '')
    size = request.GET.get('size', '')
    color = request.GET.get('color', '')
    design = request.GET.get('design', '')
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')
    min_quantity = request.GET.get('min_quantity', '')
    max_quantity = request.GET.get('max_quantity', '')

    if query:
        queryset = queryset.filter(
            Q(brand__icontains=query) |
            Q(color__icontains=query) |
            Q(category__icontains=query) |
            Q(design__icontains=query) |
            Q(size__icontains=query)
        )
    if category:
        queryset = queryset.filter(category=category)
    if shop:
        queryset = queryset.filter(shop=shop)
    if size:
        queryset = queryset.filter(size__icontains=size)
    if color:
        queryset = queryset.filter(color=color)
    if design:
        queryset = queryset.filter(design=design)
    if min_price:
        queryset = queryset.filter(price__gte=min_price)
    if max_price:
        queryset = queryset.filter(price__lte=max_price)
    if min_quantity:
        queryset = queryset.filter(quantity__gte=min_quantity)
    if max_quantity:
        queryset = queryset.filter(quantity__lte=max_quantity)

    return queryset


@login_required(login_url='login')
def product_list(request):
    # Get all filter parameters
    query = request.GET.get('search', '')
    barcode = request.GET.get('barcode', '')
    category = request.GET.get('category', '')
    shop = request.GET.get('shop', '')
    size = request.GET.get('size', '')
    color = request.GET.get('color', '')
    design = request.GET.get('design', '')
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')
    min_quantity = request.GET.get('min_quantity', '')
    max_quantity = request.GET.get('max_quantity', '')
    sort_by_name = request.GET.get('sort_by_name', '')
    sort_by_price = request.GET.get('sort_by_price', '')
    sort_by_quantity = request.GET.get('sort_by_quantity', '')

    # Build a single Q object for all filters
    filters = Q()

    # Always exclude products with zero or negative quantities
    filters &= Q(quantity__gt=0)

    # Default to showing only STORE (shop floor) items unless user explicitly selects warehouse
    if shop:
        filters &= Q(shop=shop)
    else:
        filters &= Q(shop='STORE')  # Default to shop floor items

    if query:
        filters &= (
                Q(brand__icontains=query) |
                Q(color__icontains=query) |
                Q(category__icontains=query) |
                Q(design__icontains=query) |
                Q(size__icontains=query)
        )

    if barcode:
        filters &= Q(barcode_number__icontains=barcode)

    if category:
        filters &= Q(category=category)
    if size:
        filters &= Q(size__icontains=size)
    if color:
        filters &= Q(color=color)
    if design:
        filters &= Q(design=design)
    if min_price:
        filters &= Q(price__gte=min_price)
    if max_price:
        filters &= Q(price__lte=max_price)
    if min_quantity:
        filters &= Q(quantity__gte=min_quantity)
    if max_quantity:
        filters &= Q(quantity__lte=max_quantity)

    # Apply all filters at once
    products = Product.objects.filter(filters)

    # Apply sorting
    if sort_by_name:
        products = products.order_by(sort_by_name)
    elif sort_by_price:
        products = products.order_by(sort_by_price)
    elif sort_by_quantity:
        products = products.order_by(sort_by_quantity)
    else:
        products = products.order_by('brand')  # Default sorting

    # ✅ GET CACHED STATS (FAST!)
    stats = get_product_stats()
    total_items = stats['total_items']
    total_quantity = stats['total_quantity']
    total_inventory_value = stats['total_inventory_value']

    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(products, 25)
    try:
        products = paginator.page(page)
    except PageNotAnInteger:
        products = paginator.page(1)
    except EmptyPage:
        products = paginator.page(paginator.num_pages)

    # Check if any filters are applied
    has_filters = any([query, barcode, category, shop, size, color, design, min_price, max_price, min_quantity, max_quantity])

    # ✅ GET CACHED CHOICES (NO MORE FLATTENING IN VIEW!)
    color_choices = get_cached_choices('color')
    design_choices = get_cached_choices('design')
    category_choices = get_cached_choices('category')

    # Process shop choices - ensure they're in tuple format
    shop_choices = []
    for shop in ProductChoices.SHOP_TYPE:
        if isinstance(shop, (tuple, list)) and len(shop) >= 2:
            shop_choices.append((shop[0], shop[1]))
        else:
            shop_choices.append((shop, shop))

    return render(request, 'product/product_list.html', {
        'products': products,
        'query': query,
        'category_choices': category_choices,
        'total_items': total_items,
        'total_quantity': total_quantity,
        'total_inventory_value': total_inventory_value,
        'shop_choices': shop_choices,
        'has_filters': has_filters,
        'COLOR_CHOICES': color_choices,
        'DESIGN_CHOICES': design_choices,
        'is_superuser': request.user.is_superuser,
        'current_filters': {
            'barcode': barcode,
            'category': category,
            'shop': shop,
            'size': size,
            'color': color,
            'design': design,
            'min_price': min_price,
            'max_price': max_price,
            'min_quantity': min_quantity,
            'max_quantity': max_quantity,
            'sort_by_name': sort_by_name,
            'sort_by_price': sort_by_price,
            'sort_by_quantity': sort_by_quantity,
        }
    })


@login_required(login_url='login')
def barcode_print_manager(request):
    """Barcode print manager with selection and quantity controls and optional sorting"""
    # Get filter parameters
    search = request.GET.get('search', '')
    category = request.GET.get('category', '')
    shop = request.GET.get('shop', '')
    size = request.GET.get('size', '')
    color = request.GET.get('color', '')
    design = request.GET.get('design', '')
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')
    min_quantity = request.GET.get('min_quantity', '')
    max_quantity = request.GET.get('max_quantity', '')
    sort_by_name = request.GET.get('sort_by_name', '')
    sort_by_price = request.GET.get('sort_by_price', '')
    sort_by_quantity = request.GET.get('sort_by_quantity', '')
    sort_by_id = request.GET.get('sort_by_id', '')  # New: sort by ID (creation order)

    # Start with all products
    products = Product.objects.all()

    # Apply filters
    if search:
        products = products.filter(
            Q(brand__icontains=search) |
            Q(barcode_number__icontains=search) |
            Q(color__icontains=search) |
            Q(category__icontains=search) |
            Q(design__icontains=search) |
            Q(size__icontains=search)
        )

    if category:
        products = products.filter(category=category)

    if shop:
        products = products.filter(shop=shop)

    if size:
        products = products.filter(size__icontains=size)

    if color:
        products = products.filter(color=color)

    if design:
        products = products.filter(design=design)

    if min_price:
        products = products.filter(price__gte=min_price)

    if max_price:
        products = products.filter(price__lte=max_price)

    if min_quantity:
        products = products.filter(quantity__gte=min_quantity)

    if max_quantity:
        products = products.filter(quantity__lte=max_quantity)

    # Apply sorting
    if sort_by_name:
        products = products.order_by(sort_by_name)
    elif sort_by_price:
        products = products.order_by(sort_by_price)
    elif sort_by_quantity:
        products = products.order_by(sort_by_quantity)
    elif sort_by_id:
        products = products.order_by(sort_by_id)  # Sort by ID (creation order)
    else:
        products = products.order_by('brand')  # Default sorting

    # ✅ GET CACHED STATS
    stats = get_product_stats()
    total_quantity = stats['total_quantity']
    total_inventory_value = stats['total_inventory_value']

    # Calculate barcode-specific stats
    total_products = products.count()
    products_with_barcode = products.filter(barcode_image__isnull=False).exclude(barcode_image='').count()
    products_without_barcode = total_products - products_with_barcode

    # Pagination
    paginator = Paginator(products, 25)
    page_number = request.GET.get('page')
    products_page = paginator.get_page(page_number)

    # Check if filters are active
    has_filters = any([search, category, shop, size, color, design, min_price, max_price, min_quantity, max_quantity])

    # ✅ GET CACHED CHOICES — NO MORE FLATTENING IN VIEW!
    color_choices = get_cached_choices('color')
    design_choices = get_cached_choices('design')
    category_choices = get_cached_choices('category')

    # Process shop choices - ensure they're in tuple format
    shop_choices = []
    for shop in ProductChoices.SHOP_TYPE:
        if isinstance(shop, (tuple, list)) and len(shop) >= 2:
            shop_choices.append((shop[0], shop[1]))
        else:
            shop_choices.append((shop, shop))

    context = {
        'products': products_page,
        'query': search,
        'current_filters': {
            'search': search,
            'category': category,
            'shop': shop,
            'size': size,
            'color': color,
            'design': design,
            'min_price': min_price,
            'max_price': max_price,
            'min_quantity': min_quantity,
            'max_quantity': max_quantity,
            'sort_by_name': sort_by_name,
            'sort_by_price': sort_by_price,
            'sort_by_quantity': sort_by_quantity,
            'sort_by_id': sort_by_id,
        },
        'has_filters': has_filters,
        'category_choices': category_choices,
        'shop_choices': shop_choices,
        'COLOR_CHOICES': color_choices,
        'DESIGN_CHOICES': design_choices,
        'total_products': total_products,
        'total_quantity': total_quantity,
        'total_inventory_value': total_inventory_value,
        'products_with_barcode': products_with_barcode,
        'products_without_barcode': products_without_barcode,
        'is_superuser': request.user.is_superuser,
    }

    return render(request, 'barcode/barcode_print_manager.html', context)


@login_required(login_url='login')
def transfer_menu(request):
    return render(request, 'transfers/transfer_menu.html')


@login_required(login_url='login')
def internal_transfer_create_view(request):
    """View for internal transfers (Warehouse ↔ Shop Floor)"""
    from .forms import InternalTransferForm

    try:
        current_location = request.user.profile.location
    except:
        current_location = 'ABUJA'  # Fallback

    # Get filters
    search = request.GET.get('search', '')
    category = request.GET.get('category', '')
    size = request.GET.get('size', '')
    color = request.GET.get('color', '')
    design = request.GET.get('design', '')
    min_quantity = request.GET.get('min_quantity', '')
    max_quantity = request.GET.get('max_quantity', '')
    shop_filter = request.GET.get('shop', '')  # Filter by shop type

    # Filter warehouse inventory at current location
    from .models import WarehouseInventory
    warehouse_items = WarehouseInventory.objects.filter(location=current_location, quantity__gt=0)

    if search:
        warehouse_items = warehouse_items.filter(
            Q(brand__icontains=search) |
            Q(category__icontains=search) |
            Q(size__icontains=search) |
            Q(color__icontains=search) |
            Q(design__icontains=search)
        )

    if category:
        warehouse_items = warehouse_items.filter(category=category)
    if size:
        warehouse_items = warehouse_items.filter(size__icontains=size)
    if color:
        warehouse_items = warehouse_items.filter(color=color)
    if design:
        warehouse_items = warehouse_items.filter(design=design)
    if min_quantity:
        warehouse_items = warehouse_items.filter(quantity__gte=min_quantity)
    if max_quantity:
        warehouse_items = warehouse_items.filter(quantity__lte=max_quantity)

    warehouse_items = warehouse_items.order_by('brand', 'category', 'size')

    # Pagination
    paginator = Paginator(warehouse_items, 100)
    page_number = request.GET.get('page')
    products_page = paginator.get_page(page_number)

    # Unique values for dropdowns from warehouse inventory
    categories = WarehouseInventory.objects.filter(location=current_location).values_list('category', flat=True).distinct()
    sizes = WarehouseInventory.objects.filter(location=current_location).values_list('size', flat=True).distinct()
    colors = WarehouseInventory.objects.filter(location=current_location).values_list('color', flat=True).distinct()
    designs = WarehouseInventory.objects.filter(location=current_location).values_list('design', flat=True).distinct()

    has_filters = any([search, category, size, color, design, min_quantity, max_quantity, shop_filter])

    # Handle POST
    if request.method == 'POST':
        transfer_form = InternalTransferForm(request.POST)
        selected_products = []
        has_errors = False

        # Get selected products data
        selected_products_data = request.POST.get('selected_products_data', '')

        if selected_products_data:
            try:
                import json
                frontend_selections = json.loads(selected_products_data)

                # Validate and collect warehouse items
                for item_id, item_data in frontend_selections.items():
                    try:
                        warehouse_item = WarehouseInventory.objects.get(id=item_id, location=current_location)
                        qty = int(item_data.get('quantity', 1))

                        if qty <= 0:
                            messages.error(request, f"Invalid quantity for {warehouse_item.brand}.")
                            has_errors = True
                        elif qty > warehouse_item.quantity:
                            messages.error(request,
                                           f"Not enough stock for {warehouse_item.brand}. Available: {warehouse_item.quantity}, Requested: {qty}")
                            has_errors = True
                        else:
                            selected_products.append({'warehouse_item': warehouse_item, 'quantity': qty})
                    except WarehouseInventory.DoesNotExist:
                        messages.error(request, f"Warehouse item with ID {item_id} not found.")
                        has_errors = True
                    except (ValueError, TypeError):
                        messages.error(request, f"Invalid quantity data for warehouse item ID {item_id}.")
                        has_errors = True

            except json.JSONDecodeError:
                messages.error(request, "Invalid selection data format.")
                has_errors = True

        if not selected_products and not has_errors:
            messages.error(request, "Please select at least one product for internal transfer.")
            has_errors = True

        if not has_errors and transfer_form.is_valid():
            destination = transfer_form.cleaned_data['destination']
            notes = transfer_form.cleaned_data.get('notes', '')

            try:
                with transaction.atomic():
                    # Determine if this is location transfer or shop floor transfer
                    if destination == 'STORE':
                        # Transfer to shop floor (same location)
                        transfer_type = 'internal'
                        to_location = None
                        to_shop = 'STORE'
                        destination_desc = f'Shop Floor ({current_location})'
                    else:
                        # Transfer to different location (Abuja or Lagos)
                        transfer_type = 'location'
                        to_location = destination
                        to_shop = 'STORE'  # Goes to shop floor at destination
                        destination_desc = destination

                    # Create transfer record
                    transfer = LocationTransfer.objects.create(
                        transfer_type=transfer_type,
                        from_shop='WAREHOUSE',
                        to_shop=to_shop,
                        from_location=current_location if transfer_type == 'location' else None,
                        to_location=to_location,
                        internal_location=current_location if transfer_type == 'internal' else None,
                        transfer_reference=LocationTransfer.generate_transfer_reference(
                            transfer_type=transfer_type,
                            from_location=current_location,
                            to_location=to_location,
                            from_shop='WAREHOUSE',
                            to_shop=to_shop
                        ),
                        created_by=request.user,
                        notes=notes,
                        status='COMPLETED' if transfer_type == 'internal' else 'PENDING'
                    )

                    total_items = 0
                    total_value = 0

                    for item in selected_products:
                        warehouse_item = item['warehouse_item']
                        quantity = item['quantity']

                        # Create transfer item record (using a dummy Product reference for now)
                        # Note: We'll need to adjust TransferItem to handle warehouse items properly
                        # For now, create a temporary product reference
                        temp_product = Product.objects.filter(
                            brand=warehouse_item.brand,
                            category=warehouse_item.category,
                            size=warehouse_item.size,
                            color=warehouse_item.color,
                            design=warehouse_item.design,
                            location=current_location
                        ).first()

                        if not temp_product:
                            # Create a placeholder if needed
                            temp_product = Product.objects.create(
                                brand=warehouse_item.brand,
                                category=warehouse_item.category,
                                size=warehouse_item.size,
                                color=warehouse_item.color,
                                design=warehouse_item.design,
                                price=warehouse_item.price,
                                markup_type=warehouse_item.markup_type,
                                markup=warehouse_item.markup,
                                selling_price=warehouse_item.selling_price,
                                shop='STORE',
                                location=current_location,
                                quantity=0,
                                barcode_number=warehouse_item.original_barcode
                            )

                        TransferItem.objects.create(
                            transfer=transfer,
                            product=temp_product,
                            quantity=quantity,
                            unit_price=warehouse_item.price
                        )

                        # Update product based on destination
                        if destination == 'STORE':
                            # Move to shop floor at same location (internal transfer FROM warehouse TO shop floor)
                            # Find or create shop floor product
                            shop_product = temp_product  # Use the temp_product we created above

                            # Add quantity to shop floor
                            shop_product.quantity += quantity

                            # Restore barcode if it was saved in warehouse
                            if warehouse_item.original_barcode and not shop_product.barcode_number:
                                shop_product.barcode_number = warehouse_item.original_barcode

                            shop_product.save()

                            # Subtract from warehouse inventory
                            warehouse_item.quantity -= quantity
                            if warehouse_item.quantity <= 0:
                                # Warehouse depleted, delete it
                                warehouse_item.delete()
                            else:
                                warehouse_item.save()
                        else:
                            # Move to different location shop floor (location transfer from warehouse)
                            # Find or create product at destination location
                            dest_product = Product.objects.filter(
                                location=destination,
                                shop='STORE',
                                brand=warehouse_item.brand,
                                category=warehouse_item.category,
                                size=warehouse_item.size,
                                color=warehouse_item.color,
                                design=warehouse_item.design
                            ).first()

                            if not dest_product:
                                # Create product at destination location
                                dest_product = Product.objects.create(
                                    brand=warehouse_item.brand,
                                    category=warehouse_item.category,
                                    size=warehouse_item.size,
                                    color=warehouse_item.color,
                                    design=warehouse_item.design,
                                    price=warehouse_item.price,
                                    markup_type=warehouse_item.markup_type,
                                    markup=warehouse_item.markup,
                                    selling_price=warehouse_item.selling_price,
                                    shop='STORE',
                                    location=destination,
                                    quantity=0,
                                    barcode_number=warehouse_item.original_barcode
                                )

                            # Add quantity to destination
                            dest_product.quantity += quantity
                            dest_product.save()

                            # Subtract from source warehouse
                            warehouse_item.quantity -= quantity
                            if warehouse_item.quantity <= 0:
                                warehouse_item.delete()
                            else:
                                warehouse_item.save()

                        total_items += quantity
                        total_value += Decimal(str(warehouse_item.price)) * quantity

                    # Update transfer totals
                    transfer.total_items = total_items
                    transfer.total_value = total_value
                    transfer.save()

                    if not has_errors:
                        # Log the transfer
                        ActivityLog.log_activity(
                            user=request.user,
                            action='warehouse_transfer',
                            description=f'Warehouse transfer {transfer.transfer_reference} - Warehouse → {destination_desc} - {total_items} items - ₦{total_value:,.2f}',
                            model_name='LocationTransfer',
                            object_id=transfer.id,
                            object_repr=transfer.transfer_reference,
                            extra_data={
                                'from': 'WAREHOUSE',
                                'to': destination_desc,
                                'total_items': total_items,
                                'total_value': float(total_value),
                                'transfer_type': transfer_type
                            },
                            request=request
                        )

                        success_message = f"Transfer {transfer.transfer_reference} completed! {total_items} items moved from Warehouse to {destination_desc} (₦{total_value:,.2f})"
                        messages.success(request, success_message)

                        # Clear the session storage
                        request.session['transfer_created'] = True

                        return redirect('transfer_detail', transfer_id=transfer.id)

            except Exception as e:
                import traceback
                error_msg = f"Error completing internal transfer: {str(e)}"
                print(error_msg)
                print(traceback.format_exc())
                messages.error(request, f"An error occurred: {str(e)}")
        else:
            # Form validation errors
            if transfer_form.errors:
                for field, errors in transfer_form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field.title()}: {error}")
    else:
        transfer_form = InternalTransferForm()

    # Get store config for currency symbol
    from .models import StoreConfiguration
    store_config = StoreConfiguration.get_active_config()

    context = {
        'transfer_form': transfer_form,
        'products': products_page,
        'categories': categories,
        'sizes': sizes,
        'colors': colors,
        'designs': designs,
        'current_location': current_location,
        'has_filters': has_filters,
        'current_filters': {
            'search': search,
            'category': category,
            'size': size,
            'color': color,
            'design': design,
            'min_quantity': min_quantity,
            'max_quantity': max_quantity,
            'shop': shop_filter,
        },
        'transfer_created': request.session.pop('transfer_created', False),
        'currency_symbol': store_config.currency_symbol if store_config else '₦',
        'is_internal_transfer': True,  # Flag to indicate this is internal transfer
    }

    return render(request, 'transfers/internal_transfer.html', context)


import logging
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.shortcuts import render, redirect
from django.urls import reverse
from .models import Product, LocationTransfer, TransferItem
from .forms import LocationTransferForm

logger = logging.getLogger(__name__)

@login_required(login_url='login')
def transfer_create_view(request):
    try:
        from_location = request.user.profile.location
    except:
        from_location = 'ABUJA'  # Fallback

    # Get filters
    search = request.GET.get('search', '')
    category = request.GET.get('category', '')
    size = request.GET.get('size', '')
    color = request.GET.get('color', '')
    design = request.GET.get('design', '')
    min_quantity = request.GET.get('min_quantity', '')
    max_quantity = request.GET.get('max_quantity', '')

    # Filter products at source location
    products = Product.objects.filter(location=from_location, quantity__gt=0)

    if search:
        products = products.filter(
            Q(brand__icontains=search) |
            Q(category__icontains=search) |
            Q(size__icontains=search) |
            Q(color__icontains=search) |
            Q(design__icontains=search)
        )

    if category:
        products = products.filter(category=category)
    if size:
        products = products.filter(size__icontains=size)
    if color:
        products = products.filter(color=color)
    if design:
        products = products.filter(design=design)
    if min_quantity:
        products = products.filter(quantity__gte=min_quantity)
    if max_quantity:
        products = products.filter(quantity__lte=max_quantity)

    products = products.order_by('brand', 'category', 'size')

    # Pagination
    paginator = Paginator(products, 100)
    page_number = request.GET.get('page')
    products_page = paginator.get_page(page_number)

    # Unique values for dropdowns
    categories = Product.objects.filter(location=from_location).values_list('category', flat=True).distinct()
    sizes = Product.objects.filter(location=from_location).values_list('size', flat=True).distinct()
    colors = Product.objects.filter(location=from_location).values_list('color', flat=True).distinct()
    designs = Product.objects.filter(location=from_location).values_list('design', flat=True).distinct()

    has_filters = any([search, category, size, color, design, min_quantity, max_quantity])

    # Handle POST
    if request.method == 'POST':
        transfer_form = LocationTransferForm(request.POST)
        selected_products = []
        has_errors = False

        # Debug: Print POST data
        print("POST data:", request.POST)

        # Check if we have selected_products_data from frontend
        selected_products_data = request.POST.get('selected_products_data', '')

        if selected_products_data:
            try:
                # Parse the JSON data from frontend
                import json
                frontend_selections = json.loads(selected_products_data)

                # Process frontend selections
                for product_id, item_data in frontend_selections.items():
                    try:
                        product = Product.objects.get(id=product_id, location=from_location)
                        qty = int(item_data.get('quantity', 1))

                        if qty <= 0:
                            messages.error(request, f"Invalid quantity for {product.brand}.")
                            has_errors = True
                        elif qty > product.quantity:
                            messages.error(request,
                                           f"Not enough stock for {product.brand}. Available: {product.quantity}, Requested: {qty}")
                            has_errors = True
                        else:
                            selected_products.append({'product': product, 'quantity': qty})
                            print(f"Added from frontend data: {product.brand} - Qty: {qty}")
                    except Product.DoesNotExist:
                        messages.error(request,
                                       f"Product with ID {product_id} not found or not available at {from_location}.")
                        has_errors = True
                    except (ValueError, TypeError) as e:
                        messages.error(request, f"Invalid quantity data for product ID {product_id}.")
                        has_errors = True

            except json.JSONDecodeError:
                messages.error(request, "Invalid selection data format.")
                has_errors = True
        else:
            # Fallback to traditional form processing (for products visible on current page)
            # Get all products (not just paginated) to capture selections from all pages
            all_products = Product.objects.filter(location=from_location, quantity__gt=0)

            for product in all_products:
                selected_key = f'product_{product.id}_selected'
                quantity_key = f'product_{product.id}_quantity'

                if request.POST.get(selected_key):
                    try:
                        qty = int(request.POST.get(quantity_key, 1))
                        if qty <= 0:
                            messages.error(request, f"Invalid quantity for {product.brand}.")
                            has_errors = True
                        elif qty > product.quantity:
                            messages.error(request,
                                           f"Not enough stock for {product.brand}. Available: {product.quantity}, Requested: {qty}")
                            has_errors = True
                        else:
                            selected_products.append({'product': product, 'quantity': qty})
                            print(f"Added from form data: {product.brand} - Qty: {qty}")
                    except (ValueError, TypeError):
                        messages.error(request, f"Invalid quantity for {product.brand}.")
                        has_errors = True

        if not selected_products and not has_errors:
            messages.error(request, "Please select at least one product for transfer.")
            has_errors = True

        if not has_errors and transfer_form.is_valid():
            try:
                with transaction.atomic():
                    transfer = transfer_form.save(commit=False)
                    transfer.from_location = from_location
                    transfer.transfer_reference = LocationTransfer.generate_transfer_reference(
                        from_location, transfer.to_location
                    )
                    transfer.created_by = request.user
                    transfer.save()

                    # Debug: Print transfer details
                    print(f"Creating transfer: {transfer.transfer_reference}")

                    total_items = 0
                    total_value = 0

                    for item in selected_products:
                        # Debug: Print each transfer item
                        print(f"Adding product {item['product'].id}, quantity: {item['quantity']}")

                        # Create transfer item - let the model handle quantity deduction
                        transfer_item = TransferItem.objects.create(
                            transfer=transfer,
                            product=item['product'],
                            quantity=item['quantity'],
                            unit_price=item['product'].price
                        )

                        total_items += item['quantity']
                        total_value += item['product'].price * item['quantity']

                        print(
                            f"Transfer item created - model will handle quantity deduction for {item['product'].brand}")

                    # Update transfer totals
                    transfer.total_items = total_items
                    transfer.total_value = total_value
                    transfer.save()

                    # Log transfer creation
                    ActivityLog.log_activity(
                        user=request.user,
                        action='transfer_create',
                        description=f'Created transfer {transfer.transfer_reference} - {from_location} → {transfer.to_location} - {total_items} items - ₦{total_value:,.2f}',
                        model_name='LocationTransfer',
                        object_id=transfer.id,
                        object_repr=transfer.transfer_reference,
                        extra_data={
                            'from_location': from_location,
                            'to_location': transfer.to_location,
                            'total_items': total_items,
                            'total_value': float(total_value)
                        },
                        request=request
                    )

                    success_message = f"Transfer {transfer.transfer_reference} created successfully! {total_items} items transferred (₦{total_value:,.2f})"
                    messages.success(request, success_message)

                    # Clear the session storage by adding a success flag
                    request.session['transfer_created'] = True

                    return redirect('transfer_detail', transfer_id=transfer.id)

            except Exception as e:
                # More detailed error logging
                import traceback
                error_msg = f"Error creating transfer: {str(e)}"
                print(error_msg)
                print(traceback.format_exc())
                messages.error(request, f"An error occurred while creating the transfer: {str(e)}")
        else:
            # Form validation errors
            if transfer_form.errors:
                for field, errors in transfer_form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field.title()}: {error}")

            # If we have selected products but form errors, preserve the selection
            if selected_products:
                messages.info(request,
                              f"Please fix the form errors. You have {len(selected_products)} products selected.")
    else:
        transfer_form = LocationTransferForm()

    # Check if transfer was just created (to clear frontend storage)
    transfer_created = request.session.pop('transfer_created', False)

    context = {
        'transfer_form': transfer_form,
        'products': products_page,
        'from_location': from_location,
        'categories': categories,
        'sizes': sizes,
        'colors': colors,
        'designs': designs,
        'current_filters': {
            'search': search,
            'category': category,
            'size': size,
            'color': color,
            'design': design,
            'min_quantity': min_quantity,
            'max_quantity': max_quantity,
        },
        'has_filters': has_filters,
        'transfer_created': transfer_created,
    }
    return render(request, 'transfers/create_transfer.html', context)



@login_required(login_url='login')
def download_transfer_document(request, transfer_id, format_type):
    transfer = get_object_or_404(LocationTransfer, id=transfer_id)
    transfer_items = transfer.transfer_items.all()

    if format_type == 'pdf':
        return generate_transfer_pdf(transfer, transfer_items)
    elif format_type == 'excel':
        return generate_transfer_excel(transfer, transfer_items)
    else:
        messages.error(request, "Invalid format requested.")
        return redirect('transfer_detail', transfer_id=transfer_id)


@login_required(login_url='login')
def transfer_list_view(request):
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    type_filter = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Start with all transfers
    transfers = LocationTransfer.objects.all().order_by('-transfer_date')

    # Apply filters
    if search_query:
        transfers = transfers.filter(transfer_reference__icontains=search_query)

    if status_filter:
        transfers = transfers.filter(status=status_filter)

    if type_filter:
        transfers = transfers.filter(transfer_type=type_filter)

    # Date filters
    if date_from:
        transfers = transfers.filter(transfer_date__gte=date_from)
    if date_to:
        transfers = transfers.filter(transfer_date__lte=date_to)

    # No need to annotate - total_value and total_items are now fields on the model

    context = {
        'transfers': transfers,
        'search_query': search_query,
        'status_filter': status_filter,
        'type_filter': type_filter,
        'date_from': date_from,
        'date_to': date_to,
        'status_choices': LocationTransfer.TRANSFER_STATUS_CHOICES,
        'type_choices': LocationTransfer.TRANSFER_TYPE_CHOICES,
    }
    return render(request, 'transfers/transfer_list.html', context)


@login_required(login_url='login')
def transfer_detail_view(request, transfer_id):
    transfer = get_object_or_404(LocationTransfer, id=transfer_id)
    transfer_items = transfer.transfer_items.all()

    # Calculate totals (still safe since item.total_price should be Decimal)
    total_items = sum(item.quantity for item in transfer_items)
    total_value = sum(item.total_price for item in transfer_items)

    context = {
        'transfer': transfer,
        'transfer_items': transfer_items,
        'total_items': total_items,
        'total_value': total_value,
    }
    return render(request, 'transfers/transfer_detail.html', context)


@login_required(login_url='login')
def download_transfer_pdf(request, transfer_id):
    # Get transfer and related data
    transfer = get_object_or_404(LocationTransfer, id=transfer_id)
    transfer_items = transfer.transfer_items.select_related('product').all()

    # Calculate total value
    total_value = sum(item.total_price for item in transfer_items)

    # Build logo URL
    domain = get_current_site(request).domain
    protocol = 'https' if request.is_secure() else 'http'
    logo_url = f'{protocol}://{domain}{static("img/Wlogo.png")}'

    # Context for Template
    context = {
        'transfer': transfer,
        'transfer_items': transfer_items,
        'total_value': total_value,
        'logo_url': logo_url,
        'current_date': timezone.now().strftime('%Y-%m-%d %H:%M'),
    }

    # Render HTML & Generate PDF
    html_string = render_to_string('transfers/transfer_pdf.html', context)
    pdf = HTML(string=html_string).write_pdf()

    # HTTP Response
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Transfer_{transfer.transfer_reference}.pdf"'
    return response


def generate_transfer_pdf(transfer, transfer_items, logo_url=None):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=40, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    # Add custom styles
    styles.add(ParagraphStyle(
        name='Header',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center
    ))

    styles.add(ParagraphStyle(
        name='Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey
    ))

    # Title with logo (if available)
    if logo_url:
        try:
            # Download and add logo
            with urllib.request.urlopen(logo_url) as response:
                img_data = response.read()
            logo = Image(ImageReader(io.BytesIO(img_data)), width=1.5 * inch, height=0.75 * inch)
            elements.append(logo)
        except:
            # Fallback if logo can't be loaded
            title = Paragraph("TRANSFER DOCUMENT", styles['Header'])
            elements.append(title)
    else:
        title = Paragraph("TRANSFER DOCUMENT", styles['Header'])
        elements.append(title)

    elements.append(Spacer(1, 12))

    # Transfer Info
    from_to = [
        ['Transfer Reference:', transfer.transfer_reference],
        ['Transfer Type:', transfer.get_transfer_type_display()],
    ]

    # Add location or shop info based on transfer type
    if transfer.transfer_type == 'location':
        from_to.extend([
            ['From Location:', transfer.get_from_location_display()],
            ['To Location:', transfer.get_to_location_display()],
        ])
    else:  # internal transfer
        from_to.extend([
            ['Location:', transfer.get_internal_location_display()],
            ['From:', transfer.get_from_shop_display()],
            ['To:', transfer.get_to_shop_display()],
        ])

    from_to.extend([
        ['Date:', transfer.transfer_date.strftime('%Y-%m-%d %H:%M')],
        ['Prepared By:', transfer.created_by.get_full_name() or transfer.created_by.username],
        ['Status:', transfer.get_status_display()],
    ])

    info_table = Table(from_to, colWidths=[120, 400])
    info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 15))

    # Items Table
    headers = ['Brand', 'Category', 'Size', 'Color', 'Design', 'Qty', 'Unit Price(₦)', 'Total(₦)']
    rows = [headers]

    total_value = 0
    for item in transfer_items:
        p = item.product
        total_price = item.total_price
        rows.append([
            p.brand,
            p.get_display_category() or p.category,
            p.size or '-',
            p.get_display_color() or p.color or '-',
            p.get_display_design() or p.design or '-',
            str(item.quantity),
            f"{item.unit_price:,.2f}",
            f"{total_price:,.2f}"
        ])
        total_value += total_price

    # Add total row
    rows.append(['', '', '', '', '', 'Total:', '', f"{total_value:,.2f}"])

    # Create table
    col_widths = [80, 90, 50, 60, 60, 40, 70, 80]
    item_table = Table(rows, colWidths=col_widths, repeatRows=1)
    item_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONT', (0, 1), (-1, -2), 'Helvetica', 9),
        ('FONT', (-1, -1), (-1, -1), 'Helvetica-Bold', 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),  # Blue header
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (5, 1), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('PADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#D9E1F2')]),  # Light blue alternating
    ]))
    elements.append(item_table)
    elements.append(Spacer(1, 20))

    # Signature section
    signature_data = [
        ['SENDER (ABUJA)', 'RECEIVER (LAGOS)'],
        ['', ''],
        ['_________________________', '_________________________'],
        ['Name:', 'Name:'],
        ['Date:', 'Date:'],
        ['Signature:', 'Signature:'],
    ]

    signature_table = Table(signature_data, colWidths=[250, 250])
    signature_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 12),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEABOVE', (0, 2), (-1, 2), 1, colors.black),
        ('SPAN', (0, 1), (0, 1)),
        ('SPAN', (1, 1), (1, 1)),
    ]))
    elements.append(signature_table)
    elements.append(Spacer(1, 15))

    # Footer
    footer = Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')} • Transfer ID: {transfer.id}",
                       styles['Footer'])
    elements.append(footer)

    # Notes section if available
    if transfer.notes:
        elements.append(Spacer(1, 10))
        notes_style = ParagraphStyle(
            name='Notes',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.darkgrey,
            borderPadding=5,
            borderColor=colors.lightgrey,
            backgroundColor=colors.whitesmoke,
            borderWidth=1
        )
        notes = Paragraph(f"<b>Notes:</b> {transfer.notes}", notes_style)
        elements.append(notes)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="transfer_{transfer.transfer_reference}.pdf"'
    return response


def generate_transfer_excel(transfer, transfer_items):
    output = io.BytesIO()

    # Essential data only
    data = []
    for item in transfer_items:
        p = item.product

        # Determine From and To based on transfer type
        if transfer.transfer_type == 'location':
            from_location = transfer.get_from_location_display()
            to_location = transfer.get_to_location_display()
        else:  # internal transfer
            from_location = f"{transfer.get_internal_location_display()} - {transfer.get_from_shop_display()}"
            to_location = f"{transfer.get_internal_location_display()} - {transfer.get_to_shop_display()}"

        data.append({
            'Brand': p.brand,
            'Category': p.get_display_category() or p.category,
            'Size': p.size,
            'Color': p.get_display_color() or p.color or '',
            'Design': p.get_display_design() or p.design or '',
            'From': from_location,
            'To': to_location,
            'Quantity': item.quantity,
            'Unit Price (₦)': round(float(item.unit_price), 2),
            'Total (₦)': round(float(item.total_price), 2),
        })

    df = pd.DataFrame(data)

    # Write to Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Items', index=False)
        worksheet = writer.sheets['Items']

        # Basic styling: header bold + auto-width
        for col_idx, column in enumerate(df.columns, 1):
            max_length = max(df[column].astype(str).map(len).max(), len(column)) + 2
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length, 20)

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="transfer_{transfer.transfer_reference}.xlsx"'
    return response


@login_required(login_url='login')
def update_transfer_status(request, transfer_id):
    transfer = get_object_or_404(LocationTransfer, id=transfer_id)

    if request.method == 'POST':
        new_status = request.POST.get('status')

        # Validate status choice
        if new_status not in dict(LocationTransfer.TRANSFER_STATUS_CHOICES):
            messages.error(request, "Invalid status.")
        else:
            old_status = transfer.status
            transfer.status = new_status
            transfer.save()

            # Optional: Notify only
            if new_status == 'RECEIVED':
                messages.success(request, f"Transfer {transfer.transfer_reference} marked as received. No inventory update was made at destination.")
            else:
                messages.success(request, f"Transfer status updated to {transfer.get_status_display()}.")

    return redirect('transfer_detail', transfer_id=transfer_id)



def send_receipt_email_background(receipt_id, domain, protocol='https', max_retries=2, retry_delay=600):
    """
    Send receipt email in background thread with retries
    """

    def email_task():
        for attempt in range(max_retries + 1):
            try:
                # Import models inside function to avoid import issues in thread
                from .models import Receipt, Delivery  # Adjust import path as needed

                # Get receipt and related data
                receipt = Receipt.objects.select_related('customer', 'user').get(pk=receipt_id)
                sales = receipt.sales.select_related('product').all()

                if not receipt.customer or not receipt.customer.email:
                    logger.warning(f"Receipt {receipt_id}: No customer email")
                    return

                # Get payment info
                payment = None
                if sales.exists():
                    first_sale = sales.first()
                    if hasattr(first_sale, 'payment') and first_sale.payment:
                        payment = first_sale.payment

                # Calculate totals (same as your existing logic)
                total_item_discount = sum(
                    (sale.discount_amount or Decimal('0.00')) * sale.quantity
                    for sale in sales
                )
                total_price_before_discount = sum(
                    sale.product.selling_price * sale.quantity
                    for sale in sales
                )
                total_bill_discount = payment.discount_amount if payment else Decimal('0.00')
                final_subtotal = total_price_before_discount - total_item_discount - total_bill_discount

                # Get delivery info
                delivery_cost = Decimal('0.00')
                delivery = None
                if receipt.customer:
                    try:
                        delivery = Delivery.objects.filter(customer=receipt.customer).latest('delivery_date')
                        if delivery.delivery_option == 'delivery':
                            delivery_cost = Decimal(str(delivery.delivery_cost))
                    except Delivery.DoesNotExist:
                        pass

                final_total_with_delivery = final_subtotal + delivery_cost
                logo_url = f'{protocol}://{domain}{static("img/Wlogo.png")}'

                # Get loyalty points information if customer has loyalty account
                loyalty_info = None
                try:
                    from .loyalty_utils import get_customer_loyalty_summary
                    config = LoyaltyConfiguration.get_active_config()
                    if config.is_active and receipt.customer:
                        loyalty_summary = get_customer_loyalty_summary(receipt.customer)
                        if loyalty_summary['has_account']:
                            # Get the loyalty transaction for this receipt
                            loyalty_transaction = LoyaltyTransaction.objects.filter(
                                receipt=receipt,
                                transaction_type='earned'
                            ).first()

                            if loyalty_transaction:
                                loyalty_info = {
                                    'program_name': config.program_name,
                                    'points_earned': loyalty_transaction.points,
                                    'previous_balance': loyalty_transaction.balance_after - loyalty_transaction.points,
                                    'new_balance': loyalty_transaction.balance_after,
                                    'redeemable_value': receipt.customer.loyalty_account.get_redeemable_value(),
                                }
                except Exception as e:
                    logger.error(f"Error fetching loyalty info for receipt email: {e}")

                # Get store configuration
                store_config = StoreConfiguration.get_active_config()

                # Context for templates
                context = {
                    'receipt': receipt,
                    'sales': sales,
                    'payment': payment,
                    'customer_name': receipt.customer.name,
                    'user': receipt.user,
                    'total_item_discount': total_item_discount,
                    'total_bill_discount': total_bill_discount,
                    'total_price_before_discount': total_price_before_discount,
                    'final_total': final_subtotal,
                    'final_total_with_delivery': final_total_with_delivery,
                    'delivery': delivery,
                    'logo_url': logo_url,
                    'loyalty_info': loyalty_info,
                    'store_config': store_config,
                    'store_name': store_config.store_name,
                    'store_phone': store_config.phone,
                    'store_email': store_config.email,
                    'currency_symbol': store_config.currency_symbol,
                }

                # Generate email and PDF
                html_message = render_to_string('receipt/receipt_email_template.html', context)
                pdf_html = render_to_string('receipt/receipt_pdf.html', context)

                pdf_file = BytesIO()
                HTML(string=pdf_html).write_pdf(pdf_file)
                pdf_content = pdf_file.getvalue()

                if not pdf_content:
                    raise Exception("Generated PDF is empty")

                # Validate PDF before sending
                from .pdf_validator import validate_receipt_pdf
                is_valid, error_msg = validate_receipt_pdf(pdf_content, receipt, sales, store_config)

                if not is_valid:
                    raise Exception(f"PDF validation failed: {error_msg}")

                logger.info(f"✅ PDF validation passed for receipt {receipt_id} - all required data present")

                # Send email
                logger.info(f"📧 Preparing to send email for receipt {receipt_id} to {receipt.customer.email}")
                logger.info(f"   Loyalty info included: {bool(loyalty_info)}")
                if loyalty_info:
                    logger.info(f"   Points earned: {loyalty_info['points_earned']}")

                subject = f"Your Receipt #{receipt.receipt_number}"
                email = EmailMessage(
                    subject=subject,
                    body=html_message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[receipt.customer.email],
                    attachments=[
                        (f'Receipt_{receipt.receipt_number}.pdf', pdf_content, 'application/pdf')
                    ]
                )
                email.content_subtype = "html"
                email.send()

                logger.info(f"✅ Receipt email sent successfully for receipt {receipt_id} to {receipt.customer.email}")
                logger.info(f"   Email included loyalty points: {bool(loyalty_info)}")
                return  # Success - exit the retry loop

            except Exception as e:
                logger.error(f"❌ Attempt {attempt + 1} failed for receipt {receipt_id}: {str(e)}")
                if attempt < max_retries:
                    logger.info(f"🔄 Retrying in {retry_delay} seconds...")
                    time.sleep(retry_delay)  # Wait before retry
                else:
                    logger.error(f"💥 Failed to send email for receipt {receipt_id} after {max_retries + 1} attempts")

    # Start the background thread
    thread = threading.Thread(target=email_task)
    thread.daemon = True  # Dies when main program exits
    thread.start()





@login_required(login_url='login')
def sell_product(request):
    from .models import PaymentMethod, TaxConfiguration

    # Only show products on shop floor (exclude warehouse)
    products = Product.objects.filter(quantity__gt=0, shop='STORE')
    customers = Customer.objects.all()
    SaleFormSet = formset_factory(SaleForm, extra=1)

    # Get dynamic payment method choices
    payment_method_choices = PaymentMethod.get_payment_method_choices()

    # Get active taxes for display
    active_taxes = TaxConfiguration.get_active_taxes()

    if request.method == 'POST':
        formset = SaleFormSet(request.POST, prefix='form')
        payment_form = PaymentForm(request.POST)
        delivery_form = DeliveryForm(request.POST)
        payment_methods_formset = PaymentMethodFormSet(request.POST, prefix='payment_method')
        customer_id = request.POST.get('customer')

        # Extract payment totals for validation
        total_sale_amount = Decimal(request.POST.get('total_price', '0'))
        payment_methods_total = Decimal('0')

        # Calculate total from payment methods
        total_forms = int(request.POST.get('payment_method-TOTAL_FORMS', 0))
        for i in range(total_forms):
            amount_field = f'payment_method-{i}-amount'
            if amount_field in request.POST:
                try:
                    amount_value = request.POST[amount_field]
                    if amount_value:
                        payment_methods_total += Decimal(amount_value)
                except (ValueError, TypeError):
                    pass

        # Validation form to ensure payment amounts match
        validation_form = PaymentValidationForm({
            'total_sale_amount': total_sale_amount,
            'payment_methods_total': payment_methods_total
        })

        if (formset.is_valid() and payment_form.is_valid() and
                delivery_form.is_valid() and payment_methods_formset.is_valid() and
                validation_form.is_valid()):

            try:
                customer = get_object_or_404(Customer, id=customer_id) if customer_id else None

                # Validate stock availability before processing
                stock_errors = []
                for form in formset:
                    if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                        product = form.cleaned_data['product']
                        quantity = form.cleaned_data['quantity']

                        if quantity > product.quantity:
                            stock_errors.append(
                                f"{product.brand} - Size: {product.size} - Color: {product.color}: "
                                f"Requested {quantity}, but only {product.quantity} available"
                            )

                if stock_errors:
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': False,
                            'message': 'Please adjust quantities to match available stock.',
                            'errors': stock_errors
                        })
                    else:
                        for error in stock_errors:
                            messages.error(request, error)
                        messages.error(request, "Please adjust quantities to match available stock.")
                        return render(request, 'sales/sell_product_multi_payment.html', {
                            'formset': formset,
                            'payment_form': payment_form,
                            'payment_methods_formset': payment_methods_formset,
                            'delivery_form': delivery_form,
                            'products': products,
                            'customers': customers,
                            'payment_method_choices': payment_method_choices
                        })

                # Validate payment methods
                valid_payment_methods = []
                total_payment_amount = Decimal('0')

                for form in payment_methods_formset:
                    if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                        if form.cleaned_data.get('payment_method') and form.cleaned_data.get('amount'):
                            valid_payment_methods.append(form.cleaned_data)
                            total_payment_amount += Decimal(str(form.cleaned_data['amount']))

                if not valid_payment_methods:
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': False,
                            'message': 'At least one payment method is required.',
                            'errors': ['At least one payment method is required.']
                        })
                    else:
                        messages.error(request, "At least one payment method is required.")
                        return render(request, 'sales/sell_product_multi_payment.html', {
                            'formset': formset,
                            'payment_form': payment_form,
                            'payment_methods_formset': payment_methods_formset,
                            'delivery_form': delivery_form,
                            'products': products,
                            'customers': customers,
                            'payment_method_choices': payment_method_choices
                        })

                # Use database transaction to ensure consistency
                with transaction.atomic():
                    # Create core objects
                    receipt = Receipt.objects.create(
                        date=timezone.now(),
                        user=request.user,
                        customer=customer
                    )

                    # Process delivery
                    delivery = delivery_form.save(commit=False)
                    if delivery.delivery_option == 'delivery':
                        delivery.customer = customer
                        delivery.delivery_date = timezone.now()
                        delivery.save()
                        # Add delivery cost to receipt
                        receipt.delivery_cost = Decimal(str(delivery.delivery_cost))
                    else:
                        # For pickup, set cost to 0
                        delivery.delivery_cost = Decimal('0')
                        delivery.customer = customer
                        delivery.delivery_date = timezone.now()
                        delivery.save()
                        # Ensure receipt delivery cost is also 0
                        receipt.delivery_cost = Decimal('0')

                    # Create main payment record
                    payment = Payment.objects.create(
                        payment_status='pending',
                        discount_percentage=Decimal(str(payment_form.cleaned_data.get('discount_percentage', 0)))
                    )

                    # Process sales (with additional stock validation)
                    sale_items = []
                    subtotal = Decimal('0')

                    for idx, form in enumerate(formset):
                        if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                            product = form.cleaned_data['product']
                            quantity = form.cleaned_data['quantity']

                            # Refresh product from database to get latest stock
                            product.refresh_from_db()

                            if quantity > product.quantity:
                                raise ValidationError(
                                    f"Insufficient stock for {product.brand} - Size: {product.size} - Color: {product.color}. "
                                    f"Available: {product.quantity}, Requested: {quantity}"
                                )

                            # Create and save sale
                            sale = form.save(commit=False)
                            sale.product = product
                            sale.quantity = quantity
                            sale.customer = customer
                            sale.payment = payment
                            sale.delivery = delivery if delivery.delivery_option == 'delivery' else None
                            sale.receipt = receipt
                            sale.save()  # Triggers total_price calculation

                            # Check if this item is marked as gift (admin only)
                            is_gift = request.POST.get(f'is_gift_{idx}') == 'true'
                            if is_gift and request.user.is_superuser:
                                sale.is_gift = True
                                sale.gift_reason = request.POST.get(f'gift_reason_{idx}', '').strip()
                                sale.original_value = sale.total_price  # Store original price before making it ₦0
                                sale.total_price = Decimal('0')  # Gift items are ₦0
                                sale.save()
                                logger.info(f"Item marked as GIFT: {product.brand} - Original value: ₦{sale.original_value}")

                                # Update stock for gift items too
                                product.quantity -= quantity
                                product.save()

                                # Add sale item to list but don't add to subtotal (gifts are ₦0)
                                sale_items.append(sale)
                                subtotal += Decimal('0')
                            else:
                                # Normal sale - update stock and add to subtotal
                                product.quantity -= quantity
                                product.save()

                                # Add sale item to list and update subtotal
                                sale_items.append(sale)
                                subtotal += sale.total_price

                    # ============================================
                    # PRICING CALCULATION WITH TAX SUPPORT
                    # ============================================

                    # Step 1: Calculate items subtotal (products only, no delivery yet)
                    items_subtotal = subtotal  # This is sum of all sale items

                    # Step 2: Add delivery cost
                    delivery_cost = Decimal('0')
                    if delivery.delivery_option == 'delivery':
                        delivery_cost = Decimal(str(delivery.delivery_cost))

                    # Subtotal including delivery
                    subtotal_with_delivery = items_subtotal + delivery_cost

                    # Step 3: Apply discount (on subtotal including delivery)
                    discount_percentage = payment_form.cleaned_data.get('discount_percentage', 0)
                    discount_amount = subtotal_with_delivery * (Decimal(str(discount_percentage)) / 100) if discount_percentage else Decimal('0')
                    amount_after_discount = subtotal_with_delivery - discount_amount

                    # Step 4: Get loyalty redemption data (will apply after total is calculated)
                    loyalty_points_redeemed = int(request.POST.get('loyalty_points_redeemed', 0))
                    loyalty_discount_amount = Decimal(request.POST.get('loyalty_discount_amount', '0'))
                    loyalty_discount_applied = Decimal('0')

                    # Temporarily apply loyalty discount to calculate correct total
                    if loyalty_points_redeemed > 0 and loyalty_discount_amount > 0:
                        loyalty_discount_applied = loyalty_discount_amount
                        amount_after_discount -= loyalty_discount_applied

                    # Step 5: Calculate taxes
                    import json
                    from .models import TaxConfiguration

                    active_taxes = TaxConfiguration.get_active_taxes()
                    total_tax_amount = Decimal('0')
                    total_exclusive_tax = Decimal('0')
                    total_inclusive_tax = Decimal('0')
                    tax_details = {}

                    # Taxable amount = items after all discounts (excluding delivery)
                    # We tax items only, not delivery cost
                    items_after_discount = amount_after_discount - delivery_cost

                    # Calculate each tax
                    for tax in active_taxes:
                        tax_amount = tax.calculate_tax_amount(items_after_discount)
                        total_tax_amount += tax_amount

                        # Track inclusive vs exclusive separately
                        if tax.calculation_method == 'inclusive':
                            total_inclusive_tax += tax_amount
                        else:
                            total_exclusive_tax += tax_amount

                        # Store tax details for receipt
                        tax_details[tax.code] = {
                            'name': tax.name,
                            'rate': float(tax.rate),
                            'amount': float(tax_amount),
                            'method': tax.calculation_method,
                            'type': tax.tax_type,
                            'taxable_amount': float(items_after_discount)
                        }

                    # Step 6: Calculate final total
                    # IMPORTANT:
                    # - Inclusive tax: Already in the price, so we DON'T add it
                    # - Exclusive tax: Added on top of price, so we DO add it
                    final_total = amount_after_discount + total_exclusive_tax

                    # Step 7: Update receipt with complete pricing breakdown
                    receipt.subtotal = items_subtotal  # Items only, before delivery and tax
                    receipt.tax_amount = total_tax_amount  # Total tax (both inclusive and exclusive)
                    receipt.tax_details = json.dumps(tax_details)  # Detailed tax breakdown
                    receipt.delivery_cost = delivery_cost
                    receipt.loyalty_discount_amount = loyalty_discount_applied  # Track loyalty discount
                    receipt.loyalty_points_redeemed = loyalty_points_redeemed  # Track points redeemed
                    receipt.total_with_delivery = final_total  # Grand total including exclusive tax

                    # Step 8: NOW actually redeem the loyalty points (after total is set)
                    if loyalty_points_redeemed > 0 and customer and loyalty_discount_applied > 0:
                        try:
                            from .loyalty_utils import apply_loyalty_discount as apply_loyalty_util
                            loyalty_result = apply_loyalty_util(receipt, loyalty_points_redeemed, request.user)

                            if loyalty_result['success']:
                                logger.info(f"Successfully redeemed {loyalty_points_redeemed} loyalty points "
                                          f"(₦{loyalty_discount_applied}) for receipt {receipt.receipt_number}")
                            else:
                                logger.error(f"Failed to redeem loyalty points: {loyalty_result.get('error')}")
                                # Rollback the discount if redemption failed
                                receipt.loyalty_discount_amount = Decimal('0')
                                receipt.loyalty_points_redeemed = 0
                        except Exception as e:
                            logger.error(f"Error redeeming loyalty points for receipt {receipt.receipt_number}: {e}")
                            # Rollback the discount if redemption failed
                            receipt.loyalty_discount_amount = Decimal('0')
                            receipt.loyalty_points_redeemed = 0

                    receipt.save()

                    # ============================================
                    # PARTIAL PAYMENT HANDLING
                    # ============================================
                    enable_partial_payment = request.POST.get('enable_partial_payment') == 'true'
                    if enable_partial_payment:
                        partial_amount_str = request.POST.get('partial_amount_paying', '0')
                        try:
                            amount_paying = Decimal(partial_amount_str)
                        except (ValueError, TypeError):
                            amount_paying = Decimal('0')

                        # Validate partial payment amount
                        if amount_paying >= final_total:
                            # Paying full amount or more - treat as full payment
                            receipt.payment_status = 'paid'
                            receipt.amount_paid = final_total
                            receipt.balance_remaining = Decimal('0')
                            logger.info(f"Partial payment enabled but full amount paid: ₦{amount_paying}")
                        elif amount_paying <= 0:
                            # No payment made - mark as pending
                            receipt.payment_status = 'pending'
                            receipt.amount_paid = Decimal('0')
                            receipt.balance_remaining = final_total
                            logger.info(f"No initial payment - receipt marked as pending")
                        else:
                            # Actual partial payment
                            receipt.payment_status = 'partial'
                            receipt.amount_paid = amount_paying
                            receipt.balance_remaining = final_total - amount_paying

                            # Create initial partial payment record
                            from .models import PartialPayment
                            first_payment_method = valid_payment_methods[0]['payment_method'] if valid_payment_methods else 'Cash'
                            PartialPayment.objects.create(
                                receipt=receipt,
                                amount=amount_paying,
                                payment_method=first_payment_method,
                                notes=f"Initial partial payment - Balance: ₦{receipt.balance_remaining}",
                                received_by=request.user
                            )
                            logger.info(f"Partial payment created: Paid ₦{amount_paying}, Remaining ₦{receipt.balance_remaining}")

                        receipt.save()
                    else:
                        # Full payment
                        receipt.payment_status = 'paid'
                        receipt.amount_paid = final_total
                        receipt.balance_remaining = Decimal('0')
                        receipt.save()

                    # Log pricing breakdown for debugging
                    logger.info(f"Receipt {receipt.receipt_number} - Pricing breakdown:")
                    logger.info(f"  Items subtotal: ₦{items_subtotal}")
                    logger.info(f"  Delivery: ₦{delivery_cost}")
                    logger.info(f"  Discount: -₦{discount_amount}")
                    logger.info(f"  Loyalty discount: -₦{loyalty_discount_applied}")
                    logger.info(f"  Inclusive tax: ₦{total_inclusive_tax} (in price)")
                    logger.info(f"  Exclusive tax: ₦{total_exclusive_tax} (added)")
                    logger.info(f"  Grand total: ₦{final_total}")
                    logger.info(f"  Payment status: {receipt.payment_status}")

                    # Update payment total
                    payment.total_amount = final_total
                    payment.discount_amount = discount_amount
                    payment.loyalty_discount_amount = loyalty_discount_applied
                    payment.save()

                    # Create individual payment method records
                    payment_method_summaries = []
                    for method_data in valid_payment_methods:
                        payment_method = PaymentMethod.objects.create(
                            payment=payment,
                            payment_method=method_data['payment_method'],
                            amount=Decimal(str(method_data['amount'])),
                            reference_number=method_data.get('reference_number', ''),
                            notes=method_data.get('notes', ''),
                            status='completed',
                            # Assuming immediate completion, can be 'pending' if verification needed
                            confirmed_date=timezone.now(),
                            processed_by=request.user
                        )
                        payment_method_summaries.append({
                            'method': payment_method.get_payment_method_display(),
                            'amount': payment_method.amount,
                            'reference': payment_method.reference_number or 'N/A'
                        })

                    # Finalize payment status
                    payment.refresh_from_db()
                    payment.update_payment_status()
                    payment.save()

                    # Log sale creation
                    sale_description = f'Sale created - Receipt #{receipt.receipt_number} - Total: ₦{final_total:.2f}'
                    if customer:
                        sale_description += f' - Customer: {customer.name}'
                    ActivityLog.log_activity(
                        user=request.user,
                        action='sale_create',
                        description=sale_description,
                        model_name='Receipt',
                        object_id=receipt.id,
                        object_repr=f'Receipt #{receipt.receipt_number}',
                        extra_data={
                            'total_amount': float(final_total),
                            'items_count': len(sale_items),
                            'tax_amount': float(total_tax_amount),
                            'discount_amount': float(discount_amount)
                        },
                        request=request
                    )

                    # Prepare success message with tax breakdown
                    payment_details = "; ".join([
                        f"{method['method']}: ₦{method['amount']:.2f} ({method['reference']})"
                        for method in payment_method_summaries
                    ])

                    # Build tax summary for message
                    tax_message = ""
                    if total_tax_amount > 0:
                        tax_parts = []
                        if total_inclusive_tax > 0:
                            tax_parts.append(f"₦{total_inclusive_tax:.2f} incl.")
                        if total_exclusive_tax > 0:
                            tax_parts.append(f"₦{total_exclusive_tax:.2f} excl.")
                        tax_message = f", Tax: {' + '.join(tax_parts)} = ₦{total_tax_amount:.2f}"

                    success_message = (
                        f"Sale completed successfully! "
                        f"Total: ₦{final_total:.2f}"
                        f"{f', Discount: ₦{discount_amount:.2f}' if discount_amount > 0 else ''}"
                        f"{tax_message}. "
                        f"Payment methods: {payment_details}"
                    )

                    # Handle successful transaction based on request type
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        # AJAX request - return JSON response

                        # Trigger background email
                        if receipt.customer and receipt.customer.email:
                            domain = get_current_site(request).domain
                            protocol = 'https' if request.is_secure() else 'http'
                            send_receipt_email_background(receipt.id, domain, protocol)

                        return JsonResponse({
                            'success': True,
                            'message': success_message,
                            'redirect_url': reverse('sale_success', kwargs={'receipt_id': receipt.id}),
                            'receipt_id': receipt.id
                        })
                    else:
                        # Regular form submission - redirect normally

                        # Trigger background email
                        if receipt.customer and receipt.customer.email:
                            domain = get_current_site(request).domain
                            protocol = 'https' if request.is_secure() else 'http'
                            send_receipt_email_background(receipt.id, domain, protocol)

                        messages.success(request, success_message)
                        return redirect('sale_success', receipt_id=receipt.id)

            except ValidationError as ve:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': str(ve),
                        'errors': [str(ve)]
                    })
                else:
                    messages.error(request, str(ve))
            except Exception as e:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': f"An error occurred: {str(e)}",
                        'errors': [f"An error occurred: {str(e)}"]
                    })
                else:
                    messages.error(request, f"An error occurred: {str(e)}")
        else:
            # Form validation failed
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Collect all error messages for AJAX response
                error_messages = []

                if formset.errors:
                    error_messages.append("Product form errors found")
                    for i, error_dict in enumerate(formset.errors):
                        if error_dict:
                            field_errors = []
                            for field, errors in error_dict.items():
                                if field != 'DELETE':
                                    error_strings = [str(error) for error in errors]
                                    field_errors.append(f"{field}: {', '.join(error_strings)}")

                            if field_errors:
                                error_messages.append(f"Product row {i + 1}: {', '.join(field_errors)}")

                if payment_form.errors:
                    form_errors = []
                    for field, errors in payment_form.errors.items():
                        error_strings = [str(error) for error in errors]
                        form_errors.append(f"{field}: {', '.join(error_strings)}")
                    error_messages.append(f"Payment errors: {', '.join(form_errors)}")

                if payment_methods_formset.errors:
                    error_messages.append("Payment method errors found")
                    for i, error_dict in enumerate(payment_methods_formset.errors):
                        if error_dict:
                            field_errors = []
                            for field, errors in error_dict.items():
                                if field != 'DELETE':
                                    error_strings = [str(error) for error in errors]
                                    field_errors.append(f"{field}: {', '.join(error_strings)}")

                            if field_errors:
                                error_messages.append(f"Payment method {i + 1}: {', '.join(field_errors)}")

                if delivery_form.errors:
                    form_errors = []
                    for field, errors in delivery_form.errors.items():
                        error_strings = [str(error) for error in errors]
                        form_errors.append(f"{field}: {', '.join(error_strings)}")
                    error_messages.append(f"Delivery errors: {', '.join(form_errors)}")

                if not validation_form.is_valid():
                    validation_errors = []
                    for field, errors in validation_form.errors.items():
                        error_strings = [str(error) for error in errors]
                        validation_errors.append(f"{field}: {', '.join(error_strings)}")
                    error_messages.extend(validation_errors)

                return JsonResponse({
                    'success': False,
                    'message': 'Please fix the following errors and try again:',
                    'errors': error_messages
                })
            else:
                # Regular form submission - show messages and re-render form
                error_messages = []

                if formset.errors:
                    error_messages.append("Product form errors found")
                    for i, error_dict in enumerate(formset.errors):
                        if error_dict:
                            field_errors = []
                            for field, errors in error_dict.items():
                                if field != 'DELETE':
                                    error_strings = [str(error) for error in errors]
                                    field_errors.append(f"{field}: {', '.join(error_strings)}")

                            if field_errors:
                                error_messages.append(f"Product row {i + 1}: {', '.join(field_errors)}")

                if payment_form.errors:
                    form_errors = []
                    for field, errors in payment_form.errors.items():
                        error_strings = [str(error) for error in errors]
                        form_errors.append(f"{field}: {', '.join(error_strings)}")
                    error_messages.append(f"Payment errors: {', '.join(form_errors)}")

                if payment_methods_formset.errors:
                    error_messages.append("Payment method errors found")
                    for i, error_dict in enumerate(payment_methods_formset.errors):
                        if error_dict:
                            field_errors = []
                            for field, errors in error_dict.items():
                                if field != 'DELETE':
                                    error_strings = [str(error) for error in errors]
                                    field_errors.append(f"{field}: {', '.join(error_strings)}")

                            if field_errors:
                                error_messages.append(f"Payment method {i + 1}: {', '.join(field_errors)}")

                if delivery_form.errors:
                    form_errors = []
                    for field, errors in delivery_form.errors.items():
                        error_strings = [str(error) for error in errors]
                        form_errors.append(f"{field}: {', '.join(error_strings)}")
                    error_messages.append(f"Delivery errors: {', '.join(form_errors)}")

                if not validation_form.is_valid():
                    validation_errors = []
                    for field, errors in validation_form.errors.items():
                        error_strings = [str(error) for error in errors]
                        validation_errors.append(f"{field}: {', '.join(error_strings)}")
                    error_messages.extend(validation_errors)

                for error in error_messages:
                    messages.error(request, error)

    else:
        formset = SaleFormSet(prefix='form')
        payment_form = PaymentForm()
        payment_methods_formset = PaymentMethodFormSet(prefix='payment_method')
        delivery_form = DeliveryForm()

    return render(request, 'sales/sell_product.html', {
        'formset': formset,
        'payment_form': payment_form,
        'payment_methods_formset': payment_methods_formset,
        'delivery_form': delivery_form,
        'products': products,
        'customers': customers,
        'payment_method_choices': payment_method_choices,
        'active_taxes': active_taxes
    })



@login_required(login_url='login')
def sale_success(request, receipt_id):
    receipt = get_object_or_404(Receipt, id=receipt_id)

    return render(request, 'sales/sale_success.html', {
        'receipt': receipt,
    })




@login_required(login_url='login')
def payment_details(request, payment_id):
    """View to show detailed payment breakdown"""
    payment = get_object_or_404(Payment, id=payment_id)
    payment_methods = payment.payment_methods.all()

    context = {
        'payment': payment,
        'payment_methods': payment_methods,
        'sales': payment.sale_set.all(),
    }

    return render(request, 'sales/payment_details.html', context)


@login_required(login_url='login')
def update_payment_status(request, payment_method_id):
    """Update individual payment method status (for pending payments)"""
    payment_method = get_object_or_404(PaymentMethod, id=payment_method_id)

    if request.method == 'POST':
        new_status = request.POST.get('status')
        reference = request.POST.get('reference_number', '')
        notes = request.POST.get('notes', '')

        if new_status in dict(PaymentMethod.PAYMENT_STATUS):
            old_status = payment_method.status
            payment_method.status = new_status
            payment_method.reference_number = reference
            payment_method.notes = notes

            if new_status == 'completed':
                payment_method.confirmed_date = timezone.now()

            payment_method.save()

            # Log the status change
            from .models import PaymentLog
            PaymentLog.objects.create(
                payment_method=payment_method,
                action='status_update',
                previous_status=old_status,
                new_status=new_status,
                notes=f"Updated by {request.user.username}. {notes}",
                user=request.user
            )

            messages.success(request, f"Payment method status updated to {payment_method.get_status_display()}")
        else:
            messages.error(request, "Invalid status provided")

    return redirect('payment_details', payment_id=payment_method.payment.id)


def customer_display(request):
    return render(request, 'sales/customer_display.html')

@login_required(login_url='login')
def delivered_items_view(request):
    status_filter = request.GET.get('status', 'all')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Filter deliveries based on status
    deliveries = Delivery.objects.select_related('customer').all()
    if status_filter == 'pending':
        deliveries = deliveries.filter(delivery_status='pending')
    elif status_filter == 'delivered':
        deliveries = deliveries.filter(delivery_status='delivered')

    # Filter by date range if provided
    if start_date and end_date:
        try:
            deliveries = deliveries.filter(delivery_date__range=[start_date, end_date])
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")

    # Handle status update
    if request.method == 'POST':
        delivery_id = request.POST.get('delivery_id')
        new_status = request.POST.get('delivery_status')
        delivery = Delivery.objects.get(id=delivery_id)
        delivery.delivery_status = new_status
        delivery.save()
        messages.success(request, "Delivery status updated successfully!")
        return redirect('delivered_items')

    return render(request, 'delivery/delivered_items.html', {
        'deliveries': deliveries,
        'status_filter': status_filter,
        'start_date': start_date,
        'end_date': end_date,
    })



def receipt_list(request):
    # Get all filter parameters from GET request
    search_query = request.GET.get('search', '')
    customer_filter = request.GET.get('customer', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    amount_min = request.GET.get('amount_min', '')
    amount_max = request.GET.get('amount_max', '')
    sort_by = request.GET.get('sort_by', '-date')  # Default sort by date descending

    # Start with all receipts, prefetch related data for efficiency
    receipts = Receipt.objects.prefetch_related('sales', 'customer').order_by('-date')

    # Apply filters
    if search_query:
        # Search in receipt number, customer name, or customer phone
        receipts = receipts.filter(
            Q(receipt_number__icontains=search_query) |
            Q(customer__name__icontains=search_query) |
            Q(customer__phone_number__icontains=search_query)
        )

    if customer_filter:
        receipts = receipts.filter(customer__name__icontains=customer_filter)

    # Date filtering
    if date_from:
        try:
            date_from_parsed = datetime.strptime(date_from, '%Y-%m-%d')
            receipts = receipts.filter(date__date__gte=date_from_parsed.date())
        except ValueError:
            pass  # Invalid date format, ignore filter

    if date_to:
        try:
            date_to_parsed = datetime.strptime(date_to, '%Y-%m-%d')
            receipts = receipts.filter(date__date__lte=date_to_parsed.date())
        except ValueError:
            pass  # Invalid date format, ignore filter

    receipt_data = []

    for receipt in receipts:
        total_amount = Decimal('0.00')
        customer_name = "N/A"

        # Safely get customer name
        if hasattr(receipt, 'customer') and receipt.customer:
            customer_name = receipt.customer.name
        elif receipt.customer_id:
            try:
                customer_name = receipt.customer.name
            except AttributeError:
                customer_name = "Unknown Customer"

        # Calculate total amount including delivery cost
        if hasattr(receipt, 'sales'):
            for sale in receipt.sales.all():
                if sale and hasattr(sale, 'total_price') and sale.total_price:
                    total_amount += Decimal(str(sale.total_price))

        # Add delivery cost if exists
        if receipt.delivery_cost:
            total_amount += Decimal(str(receipt.delivery_cost))

        receipt_info = {
            'receipt': receipt,
            'total_amount': total_amount.quantize(Decimal('0.00')),
            'customer_name': customer_name
        }

        receipt_data.append(receipt_info)

    # Apply amount filtering after calculating totals
    if amount_min:
        try:
            min_amount = Decimal(amount_min)
            receipt_data = [r for r in receipt_data if r['total_amount'] >= min_amount]
        except (ValueError, TypeError):
            pass

    if amount_max:
        try:
            max_amount = Decimal(amount_max)
            receipt_data = [r for r in receipt_data if r['total_amount'] <= max_amount]
        except (ValueError, TypeError):
            pass

    # Apply sorting
    if sort_by == 'receipt_number':
        receipt_data.sort(key=lambda x: x['receipt']['receipt_number'])
    elif sort_by == '-receipt_number':
        receipt_data.sort(key=lambda x: x['receipt']['receipt_number'], reverse=True)
    elif sort_by == 'customer':
        receipt_data.sort(key=lambda x: x['customer_name'])
    elif sort_by == '-customer':
        receipt_data.sort(key=lambda x: x['customer_name'], reverse=True)
    elif sort_by == 'amount':
        receipt_data.sort(key=lambda x: x['total_amount'])
    elif sort_by == '-amount':
        receipt_data.sort(key=lambda x: x['total_amount'], reverse=True)
    elif sort_by == 'date':
        receipt_data.sort(key=lambda x: x['receipt'].date)
    else:  # Default: -date
        receipt_data.sort(key=lambda x: x['receipt'].date, reverse=True)

    # Get unique customers for dropdown filter
    customers = Customer.objects.filter(receipt__isnull=False).distinct().order_by('name')

    # Calculate summary statistics
    total_receipts = len(receipt_data)
    total_revenue = sum(r['total_amount'] for r in receipt_data)

    # Quick filter date calculations
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    context = {
        'receipt_data': receipt_data,
        'search_query': search_query,
        'customer_filter': customer_filter,
        'date_from': date_from,
        'date_to': date_to,
        'amount_min': amount_min,
        'amount_max': amount_max,
        'sort_by': sort_by,
        'customers': customers,
        'total_receipts': total_receipts,
        'total_revenue': total_revenue,
        'today': today.strftime('%Y-%m-%d'),
        'week_ago': week_ago.strftime('%Y-%m-%d'),
        'month_ago': month_ago.strftime('%Y-%m-%d'),
    }

    return render(request, 'receipt/receipt_list.html', context)


@login_required(login_url='login')
def receipt_detail(request, pk):
    receipt = get_object_or_404(Receipt, pk=pk)

    # Log receipt view
    ActivityLog.log_activity(
        user=request.user,
        action='receipt_view',
        description=f'Viewed receipt #{receipt.receipt_number}',
        model_name='Receipt',
        object_id=receipt.id,
        object_repr=f'Receipt #{receipt.receipt_number}',
        request=request
    )

    # Fixed: Remove 'payment__delivery' from select_related since delivery is related to Sale, not Payment
    sales = receipt.sales.select_related('product', 'payment', 'delivery').prefetch_related(
        'payment__payment_methods'
    ).all()

    payment = sales.first().payment if sales.exists() else None
    customer_name = receipt.customer.name if receipt.customer else "No customer"
    user = receipt.user

    # Total item-level discounts
    total_item_discount = sum(
        (sale.discount_amount or Decimal('0.00')) for sale in sales
    )

    # Total before any discount
    total_price_before_discount = sum(
        sale.product.selling_price * sale.quantity for sale in sales
    )

    # Bill-level discount
    total_bill_discount = payment.discount_amount if payment else Decimal('0.00')

    # Get delivery cost from the receipt or from sales delivery
    delivery_cost = receipt.delivery_cost or Decimal('0.00')

    # If receipt doesn't have delivery_cost, get it from the first sale's delivery
    if not delivery_cost and sales.exists():
        first_sale_delivery = sales.first().delivery
        if first_sale_delivery:
            delivery_cost = first_sale_delivery.delivery_cost or Decimal('0.00')

    # Final total (after discounts + delivery)
    final_total = payment.total_amount if payment else Decimal('0.00')

    # Total paid via all completed methods
    if payment:
        total_paid = sum(
            pm.amount for pm in payment.payment_methods.filter(status='completed')
        )
        payment_methods = payment.payment_methods.all()
    else:
        total_paid = Decimal('0.00')
        payment_methods = []

    change_amount = max(total_paid - final_total, Decimal('0.00'))

    # Get store configuration
    store_config = StoreConfiguration.get_active_config()

    # Get delivery details
    delivery = None
    if sales.exists():
        first_sale_delivery = sales.first().delivery
        if first_sale_delivery:
            delivery = first_sale_delivery

    # Get loyalty info
    loyalty_info = None
    if receipt.customer and hasattr(receipt.customer, 'loyalty_account'):
        try:
            config = LoyaltyConfiguration.get_active_config()
            if config and config.is_active:
                loyalty_transaction = LoyaltyTransaction.objects.filter(
                    customer=receipt.customer,
                    receipt=receipt
                ).order_by('-created_at').first()

                if loyalty_transaction:
                    loyalty_info = {
                        'program_name': config.program_name,
                        'points_earned': loyalty_transaction.points,
                        'previous_balance': loyalty_transaction.balance_after - loyalty_transaction.points,
                        'new_balance': loyalty_transaction.balance_after,
                        'redeemable_value': receipt.customer.loyalty_account.get_redeemable_value(),
                    }
        except Exception as e:
            logger.error(f"Error fetching loyalty info: {e}")

    return render(request, 'receipt/receipt_detail.html', {
        'receipt': receipt,
        'sales': sales,
        'payment': payment,
        'customer_name': customer_name,
        'user': user,
        'total_item_discount': total_item_discount,
        'total_bill_discount': total_bill_discount,
        'total_price_before_discount': total_price_before_discount,
        'delivery_cost': delivery_cost,
        'final_total': final_total,
        'total_paid': total_paid,
        'change_amount': change_amount,
        'payment_methods': payment_methods,
        'store_config': store_config,
        'store_name': store_config.store_name,
        'store_phone': store_config.phone,
        'store_email': store_config.email,
        'currency_symbol': store_config.currency_symbol,
        'delivery': delivery,
        'loyalty_info': loyalty_info,
    })


@login_required(login_url='login')
def send_receipt_email(request, pk):
    logger.info(f"📧 Starting email send process for receipt {pk}")

    receipt = get_object_or_404(Receipt, pk=pk)
    sales = receipt.sales.select_related('product').all()

    if not receipt.customer or not receipt.customer.email:
        logger.warning(f"⚠️ Receipt {pk} has no customer email")
        messages.error(request, "❌ Customer does not have an email address.")
        return redirect('receipt_list')

    logger.info(f"📧 Sending email to: {receipt.customer.email} for receipt {receipt.receipt_number}")

    payment = sales.first().payment if sales.exists() and hasattr(sales.first(), 'payment') else None

    total_item_discount = sum(
        (sale.discount_amount or Decimal('0.00')) * sale.quantity
        for sale in sales
    )
    total_price_before_discount = sum(
        sale.product.selling_price * sale.quantity
        for sale in sales
    )
    total_bill_discount = payment.discount_amount if payment else Decimal('0.00')
    final_total = payment.total_amount if payment else Decimal('0.00')

    # ✅ Generate logo URL using `request` here — inside the view
    domain = get_current_site(request).domain
    protocol = 'https' if request.is_secure() else 'http'
    logo_url = f'{protocol}://{domain}{static("img/Wlogo.png")}'

    # Get loyalty points information if customer has loyalty account
    loyalty_info = None
    try:
        from .loyalty_utils import get_customer_loyalty_summary
        config = LoyaltyConfiguration.get_active_config()
        if config.is_active and receipt.customer:
            loyalty_summary = get_customer_loyalty_summary(receipt.customer)
            if loyalty_summary['has_account']:
                # Get the loyalty transaction for this receipt
                loyalty_transaction = LoyaltyTransaction.objects.filter(
                    receipt=receipt,
                    transaction_type='earned'
                ).first()

                if loyalty_transaction:
                    loyalty_info = {
                        'program_name': config.program_name,
                        'points_earned': loyalty_transaction.points,
                        'previous_balance': loyalty_transaction.balance_after - loyalty_transaction.points,
                        'new_balance': loyalty_transaction.balance_after,
                        'redeemable_value': receipt.customer.loyalty_account.get_redeemable_value(),
                    }
    except Exception as e:
        logger.error(f"Error fetching loyalty info for receipt email: {e}")

    # Get store configuration
    store_config = StoreConfiguration.get_active_config()

    # === Generate Location QR Code ===
    location_qr_code_url = None
    try:
        import qrcode
        from io import BytesIO
        import base64

        # Full address for Google Maps search
        full_address = "Wrighteous Wearhouse, Suit 10/11, Amma Centre, near AP Filling Station, opposite Old CBN, Garki, Abuja 900103, Federal Capital Territory"

        # Create Google Maps search URL
        import urllib.parse
        google_maps_url = f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote(full_address)}"

        # Generate QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(google_maps_url)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        # Convert to base64
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        qr_base64 = base64.b64encode(buffer.getvalue()).decode()
        location_qr_code_url = f'data:image/png;base64,{qr_base64}'
    except Exception as e:
        logger.error(f"Error generating location QR code for email: {e}")

    # === Get Payment Methods ===
    # Get unique payment IDs from sales in this receipt
    payment_ids = sales.values_list('payment_id', flat=True).distinct()
    payments = PaymentMethod.objects.filter(payment_id__in=payment_ids)

    context = {
        'receipt': receipt,
        'sales': sales,
        'payment': payment,
        'payments': payments,
        'customer_name': receipt.customer.name,
        'user': receipt.user,
        'total_item_discount': total_item_discount,
        'total_bill_discount': total_bill_discount,
        'total_price_before_discount': total_price_before_discount,
        'final_total': final_total,
        'final_total_with_delivery': receipt.total_with_delivery or final_total,
        'delivery': None,  # Can be added if needed
        'logo_url': logo_url,
        'location_qr_code_url': location_qr_code_url,
        'loyalty_info': loyalty_info,
        'store_config': store_config,
        'store_name': store_config.store_name,
        'store_phone': store_config.phone,
        'store_email': store_config.email,
        'currency_symbol': store_config.currency_symbol,
    }

    html_message = render_to_string('receipt/receipt_email_template.html', context)
    pdf_html = render_to_string('receipt/receipt_pdf.html', context)

    pdf_file = BytesIO()
    try:
        HTML(string=pdf_html).write_pdf(pdf_file)
        pdf_content = pdf_file.getvalue()

        if not pdf_content or len(pdf_content) == 0:
            raise Exception("Generated PDF is empty.")

        # Validate PDF before sending
        from .pdf_validator import validate_receipt_pdf
        is_valid, error_msg = validate_receipt_pdf(pdf_content, receipt, sales, store_config)

        if not is_valid:
            raise Exception(f"PDF validation failed: {error_msg}")

        logger.info(f"✅ PDF validation passed for receipt {pk} - all required data present")

    except Exception as e:
        logger.error(f"❌ Error generating PDF for receipt {pk}: {e}")
        messages.error(request, f"❌ Error generating PDF: {e}")
        return redirect('receipt_list')

    subject = f"Your Receipt #{receipt.receipt_number}"

    logger.info(f"📧 Creating email message...")
    logger.info(f"   From: {settings.DEFAULT_FROM_EMAIL}")
    logger.info(f"   To: {receipt.customer.email}")
    logger.info(f"   Subject: {subject}")
    logger.info(f"   PDF size: {len(pdf_content)} bytes")

    email = EmailMessage(
        subject=subject,
        body=html_message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[receipt.customer.email],
        attachments=[
            (f'Receipt_{receipt.receipt_number}.pdf', pdf_content, 'application/pdf')
        ]
    )
    email.content_subtype = "html"

    try:
        logger.info(f"📧 Attempting to send email...")
        email.send()
        logger.info(f"✅ Receipt email sent successfully for receipt {pk} to {receipt.customer.email}")
        if loyalty_info:
            messages.success(request, f"✅ Receipt #{receipt.receipt_number} with {loyalty_info['points_earned']} loyalty points sent successfully to {receipt.customer.email}")
        else:
            messages.success(request, f"✅ Receipt #{receipt.receipt_number} sent successfully to {receipt.customer.email}")
    except Exception as e:
        logger.error(f"❌ Failed to send email for receipt {pk}: {str(e)}")
        messages.error(request, f"❌ Failed to send email: {str(e)}")

    # Redirect back to receipt list
    return redirect('receipt_list')



@login_required(login_url='login')
def download_receipt_pdf(request, pk):
    # Get receipt and related data
    receipt = get_object_or_404(Receipt, pk=pk)
    sales = receipt.sales.select_related('product').all()

    # Get payment (if exists)
    payment = None
    if sales.exists():
        first_sale = sales.first()
        if hasattr(first_sale, 'payment') and first_sale.payment:
            payment = first_sale.payment

    # Get customer (safe handling)
    customer = receipt.customer
    customer_name = customer.name if customer else "Walk-in Customer"

    # === Calculate Financials ===
    # Subtotal: sum of (selling_price * quantity) for all items
    total_price_before_discount = sum(
        (sale.product.selling_price * sale.quantity)
        for sale in sales
    )

    # Total item-level discounts: sum of (discount_amount * quantity)
    total_item_discount = sum(
        (sale.discount_amount or Decimal('0.00')) * sale.quantity
        for sale in sales
    )

    # Bill-level discount (from payment)
    total_bill_discount = payment.discount_amount if payment else Decimal('0.00')

    # Final subtotal after all discounts
    final_subtotal = total_price_before_discount - total_item_discount - total_bill_discount

    # === Delivery Fee ===
    delivery_cost = Decimal('0.00')
    delivery = None

    if customer:
        # Get the latest delivery for this customer (or filter by receipt if you have a relation)
        try:
            delivery = Delivery.objects.filter(customer=customer).latest('delivery_date')
            # Only apply delivery cost if delivery option is 'delivery'
            if delivery.delivery_option == 'delivery':
                delivery_cost = Decimal(str(delivery.delivery_cost))
        except Delivery.DoesNotExist:
            pass

    # Final total including delivery
    final_total_with_delivery = final_subtotal + delivery_cost

    # === Build logo URL ===
    domain = get_current_site(request).domain
    protocol = 'https' if request.is_secure() else 'http'
    logo_url = f'{protocol}://{domain}{static("img/Wlogo.png")}'

    # === Generate Location QR Code ===
    location_qr_code_url = None
    try:
        import qrcode
        from io import BytesIO
        import base64

        # Full address for Google Maps search
        full_address = "Wrighteous Wearhouse, Suit 10/11, Amma Centre, near AP Filling Station, opposite Old CBN, Garki, Abuja 900103, Federal Capital Territory"

        # Create Google Maps search URL
        import urllib.parse
        google_maps_url = f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote(full_address)}"

        # Generate QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(google_maps_url)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        # Convert to base64
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        qr_base64 = base64.b64encode(buffer.getvalue()).decode()
        location_qr_code_url = f'data:image/png;base64,{qr_base64}'
    except Exception as e:
        logger.error(f"Error generating location QR code: {e}")

    # === Get Payment Methods ===
    # Get unique payment IDs from sales in this receipt
    payment_ids = sales.values_list('payment_id', flat=True).distinct()
    payments = PaymentMethod.objects.filter(payment_id__in=payment_ids)

    # === Get Store Config ===
    store_config = StoreConfiguration.get_active_config()

    # === Get Loyalty Info ===
    loyalty_info = None
    if receipt.customer and hasattr(receipt.customer, 'loyalty_account'):
        try:
            config = LoyaltyConfiguration.get_active_config()
            if config and config.is_active:
                loyalty_transaction = LoyaltyTransaction.objects.filter(
                    customer=receipt.customer,
                    receipt=receipt
                ).order_by('-created_at').first()

                if loyalty_transaction:
                    loyalty_info = {
                        'program_name': config.program_name,
                        'points_earned': loyalty_transaction.points,
                        'previous_balance': loyalty_transaction.balance_after - loyalty_transaction.points,
                        'new_balance': loyalty_transaction.balance_after,
                        'redeemable_value': receipt.customer.loyalty_account.get_redeemable_value(),
                    }
        except Exception as e:
            logger.error(f"Error fetching loyalty info: {e}")

    # === Context for Template ===
    context = {
        'receipt': receipt,
        'sales': sales,
        'payment': payment,
        'payments': payments,
        'customer_name': customer_name,
        'user': receipt.user,
        'total_price_before_discount': total_price_before_discount,
        'total_item_discount': total_item_discount,
        'total_bill_discount': total_bill_discount,
        'final_total': final_subtotal,  # Final amount before delivery
        'final_total_with_delivery': final_total_with_delivery,
        'delivery': delivery,
        'logo_url': logo_url,
        'location_qr_code_url': location_qr_code_url,
        'store_config': store_config,
        'store_name': store_config.store_name,
        'store_phone': store_config.phone,
        'store_email': store_config.email,
        'currency_symbol': store_config.currency_symbol,
        'loyalty_info': loyalty_info,
    }

    # === Render HTML & Generate PDF ===
    html_string = render_to_string('receipt/receipt_pdf.html', context)
    pdf = HTML(string=html_string).write_pdf()

    # Log receipt download
    ActivityLog.log_activity(
        user=request.user,
        action='receipt_download',
        description=f'Downloaded receipt #{receipt.receipt_number} as PDF',
        model_name='Receipt',
        object_id=receipt.id,
        object_repr=f'Receipt #{receipt.receipt_number}',
        request=request
    )

    # === HTTP Response ===
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Receipt_{receipt.receipt_number}.pdf"'
    return response



@login_required(login_url='login')
def print_receipt(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)
    receipt = sale.receipt
    payment = sale.payment
    customer = sale.customer

    # Pass details to the template
    return render(request, 'sales/print_receipt.html', {
        'sale': sale,
        'receipt': receipt,
        'payment': payment,
        'customer': customer,
    })


@login_required(login_url='login')
def customer_list_view(request):
    # Get all customers
    customers = Customer.objects.all()
    return render(request, 'customer/customer_list.html', {'customers': customers})


@login_required(login_url='login')
def customer_receipt_history(request, customer_id):
    # Get the customer or return a 404 if not found
    customer = get_object_or_404(Customer, id=customer_id)

    # Get all receipts related to this customer
    receipts = Receipt.objects.filter(customer=customer).order_by('-date').prefetch_related('sales')

    # Apply date filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            receipts = receipts.filter(date__range=[start_date_obj, end_date_obj])
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")

    # Prepare receipt data with total amounts
    receipt_data = []
    for receipt in receipts:
        # Use total_price which already includes discount
        total_amount = sum(sale.total_price for sale in receipt.sales.all())

        # Add delivery cost if present
        if receipt.delivery_cost:
            total_amount += receipt.delivery_cost

        receipt_data.append({
            'receipt': receipt,
            'total_amount': total_amount
        })

    return render(request, 'customer/customer_receipt_history.html', {
        'customer': customer,
        'receipt_data': receipt_data
    })


@login_required(login_url='login')
def update_delivery_status(request, sale_id):
    sale = Sale.objects.get(id=sale_id)
    if request.method == 'POST':
        # Mark sale as delivered
        sale.delivery_status = 'delivered'
        sale.save()

        messages.success(request, f"{sale.product.brand} marked as delivered.")
        return redirect('delivery_list')

    return render(request, 'delivery/update_delivery_status.html', {'sale': sale})


@login_required(login_url='login')
def cancel_order(request, sale_id):
    sale = Sale.objects.get(id=sale_id)
    sale.product.quantity += sale.quantity  # Restore stock quantity
    sale.product.save()
    sale.delete()
    return redirect('sell_product')  # Redirect back to sales view


@login_required(login_url='login')
def customer_list(request):
    # Retrieve the search query from the GET request
    query = request.GET.get('search', '')
    customers = Customer.objects.all()

    if query:
        # Filter customers based on the search query
        customers = customers.filter(
            Q(name__icontains=query) | Q(phone_number__icontains=query) | Q(address__icontains=query)
        )

    # Add loyalty information to each customer
    from .loyalty_utils import get_customer_loyalty_summary
    frequent_count = 0
    loyalty_count = 0

    for customer in customers:
        customer.loyalty_info = get_customer_loyalty_summary(customer)
        if customer.frequent_customer:
            frequent_count += 1
        if customer.loyalty_info.get('has_account', False):
            loyalty_count += 1

    context = {
        'customers': customers,
        'search_query': query,
        'total_customers': customers.count(),
        'frequent_customers': frequent_count,
        'loyalty_members': loyalty_count,
    }

    return render(request, 'customer/customer_list.html', context)


@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def edit_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            customer = form.save()
            # Log customer update
            ActivityLog.log_activity(
                user=request.user,
                action='customer_update',
                description=f'Updated customer: {customer.name} - {customer.phone_number}',
                model_name='Customer',
                object_id=customer.id,
                object_repr=str(customer),
                request=request
            )
            return redirect('customer_list')
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'customer/edit_customer.html', {'form': form})


@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def delete_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        # Log customer deletion before deleting
        customer_info = str(customer)
        customer_id = customer.id
        ActivityLog.log_activity(
            user=request.user,
            action='customer_delete',
            description=f'Deleted customer: {customer_info}',
            model_name='Customer',
            object_id=customer_id,
            object_repr=customer_info,
            request=request
        )
        customer.delete()
        return redirect('customer_list')
    return render(request, 'customer/delete_customer.html', {'customer': customer})


@login_required(login_url='login')
def add_product(request):
    ProductFormSet = formset_factory(ProductForm, extra=1)

    if request.method == 'POST':
        formset = ProductFormSet(request.POST, request.FILES)
        if formset.is_valid():
            has_data = any(form.cleaned_data for form in formset)
            if not has_data:
                messages.error(request, "Please add at least one product.")
            else:
                invoice = Invoice.objects.create(user=request.user)

                for idx, form in enumerate(formset):
                    if form.cleaned_data:
                        product = form.save(commit=False)
                        product.invoice = invoice

                        # Explicitly assign these so calculate_selling_price is accurate
                        product.markup = form.cleaned_data.get('markup', 0)
                        product.markup_type = form.cleaned_data.get('markup_type', 'percentage')

                        # Handle design field (optional)
                        product.design = form.cleaned_data.get('design', 'plain')

                        # Handle image field (optional - will be None if not provided)
                        if 'image' in form.cleaned_data and form.cleaned_data['image']:
                            product.image = form.cleaned_data['image']

                        product.selling_price = product.calculate_selling_price()

                        product.save()

                        # Log product creation
                        ActivityLog.log_activity(
                            user=request.user,
                            action='product_create',
                            description=f'Created product: {product.brand} ({product.category}) - Qty: {product.quantity}',
                            model_name='Product',
                            object_id=product.id,
                            object_repr=str(product),
                            request=request
                        )

                        InvoiceProduct.objects.create(
                            invoice=invoice,
                            product_name=product.brand,
                            product_price=product.price,
                            product_color=product.color,
                            product_size=product.size,
                            product_category=product.category,
                            quantity=product.quantity,
                            total_price=product.price * product.quantity
                        )
                return redirect('invoice_list')
        else:
            messages.error(request, "There were errors in the form. Please correct them.")
    else:
        formset = ProductFormSet()

    return render(request, 'product/add_product.html', {'formset': formset})



def lookup_product_by_barcode(request):
    barcode = request.GET.get('barcode')
    context = request.GET.get('context')

    if not barcode:
        return JsonResponse({'found': False, 'error': 'No barcode provided'})

    try:
        product = Product.objects.get(barcode_number=barcode)

        # Build product data
        product_data = {
            'id': product.id,
            'brand': product.brand,
            'size': product.size,
            'color': product.color,
            'design': product.get_display_design() if product.design else None,
            'price': float(product.selling_price),
            'category': product.get_display_category(),
            'shop': product.get_shop_display(),
            'location': product.location,
            'quantity': product.quantity,
        }

        # Add image URL if image exists
        if product.image:
            product_data['image_url'] = product.image.url
        else:
            product_data['image_url'] = None

        # Context-based logic
        if context == 'transfer' or context == 'transfer_to_warehouse':
            # For transfer to warehouse - check if product is on shop floor
            if product.shop != 'STORE':
                return JsonResponse({
                    'found': False,
                    'error': f'Product is in {product.get_shop_display()}, not on Shop Floor.'
                })
            if product.quantity <= 0:
                return JsonResponse({
                    'found': False,
                    'error': 'This item is out of stock.'
                })
            product_data['stock'] = product.quantity

        elif context == 'transfer_from_warehouse':
            # Check if product exists in any completed transfers to this location
            # or if it has been transferred and received at the destination
            received_transfers = TransferItem.objects.filter(
                product=product,
                transfer__status='RECEIVED',
                transfer__to_location=product.location
            ).aggregate(
                total_received=models.Sum('quantity')
            )['total_received'] or 0

            # Also check for pending transfers that might be available
            pending_transfers = TransferItem.objects.filter(
                product=product,
                transfer__status__in=['PENDING', 'IN_TRANSIT'],
                transfer__from_location=product.location
            ).aggregate(
                total_pending=models.Sum('quantity')
            )['total_pending'] or 0

            available_quantity = received_transfers

            if available_quantity <= 0:
                return JsonResponse({
                    'found': False,
                    'error': f'This item has no stock available for transfer. Available: {available_quantity}'
                })

            product_data['stock'] = available_quantity

        elif context == 'sell':
            if product.quantity <= 0:
                return JsonResponse({
                    'found': False,
                    'error': f'This item is not available for sale. Current stock: {product.quantity}. Transfer from warehouse first.'
                })
            product_data['stock'] = product.quantity

        else:
            # Default fallback behavior — just return basic product info
            product_data['stock'] = product.quantity

        # Return product data wrapped in response
        return JsonResponse({
            'found': True,
            'product': product_data
        })

    except Product.DoesNotExist:
        return JsonResponse({'found': False, 'error': 'Product not found.'})
    except Product.MultipleObjectsReturned:
        return JsonResponse({
            'found': False,
            'error': 'Multiple items found with this barcode. Please fix duplicate.'
        })
    except Exception as e:
        # Add error logging for debugging — PRINT REMOVED
        return JsonResponse({
            'found': False,
            'error': f'System error occurred: {str(e)}'
        })


def barcode_lookup_page(request):
    barcode = request.GET.get('barcode')
    context = {}

    if barcode:
        try:
            product = Product.objects.get(barcode_number=barcode)
            context['product'] = product

            # Debug: Print raw values — PRINTS REMOVED
            context['data'] = {
                'id': product.id,
                'brand': product.brand,
                'size': product.size,
                'color': product.color,
                'color_display': product.get_display_color(),
                'design': product.design,
                'design_display': product.get_display_design() if product.design else 'N/A',
                'category': product.category,
                'category_display': product.get_display_category(),
                'shop': product.shop,
                'shop_display': product.get_shop_display(),
                'price': product.selling_price,
                'quantity': product.quantity,
                'image': product.image if product.image else None,
            }

            # Debug: Print display values — PRINTS REMOVED

        except Product.DoesNotExist:
            context['error'] = 'Product not found with this barcode.'
        except Product.MultipleObjectsReturned:
            context['error'] = 'Multiple products found with this barcode. Please contact admin.'

    return render(request, 'barcode/barcode_lookup.html', context)


def print_barcode(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    # Ensure barcode is generated and saved
    if not product.barcode_image or not product.barcode_number:
        product.generate_barcode()
        product.save()

    # Ensure barcode number is properly formatted with leading zeros
    if product.barcode_number and len(product.barcode_number) < 13:
        product.barcode_number = product.barcode_number.zfill(13)
        product.save(update_fields=['barcode_number'])

    # Build absolute URL for image
    barcode_url = request.build_absolute_uri(product.barcode_image.url) if product.barcode_image else None

    # Debug information — PRINTS REMOVED

    context = {
        'product': product,
        'barcode_url': barcode_url,
    }

    response = render(request, 'barcode/print_barcode.html', context)

    # Prevent caching
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'

    return response  # This line was missing


def add_temporary_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            product.status = 'temporary'  # Mark the product as temporary
            product.save()
            messages.success(request, "Product saved temporarily.")
            return redirect('temporary_product_list')
        else:
            messages.error(request, "Please correct the errors in the form.")
    else:
        form = ProductForm()
    return render(request, 'product/add_temporary_product.html', {'form': form})

@user_passes_test(lambda u: u.is_staff)
def finalize_products(request):
    from .models import Product
    if request.method == 'POST':
        product_ids = request.POST.getlist('product_ids')
        for product_id in product_ids:
            product = Product.objects.get(id=product_id)
            if product.markup and product.markup_type:
                product.status = 'finalized'  # Finalize the product
                product.save()
            else:
                messages.warning(request, f"Product {product.brand} requires markup.")
        messages.success(request, "Selected products have been finalized.")
        return redirect('finalize_products')
    else:
        products = Product.objects.filter(status='temporary')
    return render(request, 'product/finalize_products.html', {'products': products})


# views.py

@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        reason = request.POST.get('reason', '').strip()

        if not reason:
            messages.error(request, "A reason for editing the product is required.")
        elif form.is_valid():
            # Get the original quantity before editing
            old_quantity = product.quantity

            # Save the updated product (triggers save() → barcode regeneration)
            form.save()

            # Calculate the quantity change
            quantity_changed = product.quantity - old_quantity

            # Log the edit action with quantity change
            ProductHistory.objects.create(
                product=product,
                user=request.user,
                action='EDIT',
                reason=reason,
                quantity_changed=quantity_changed
            )

            # Log to activity log
            ActivityLog.log_activity(
                user=request.user,
                action='product_update',
                description=f'Updated product: {product.brand} - Reason: {reason} - Qty change: {quantity_changed:+d}',
                model_name='Product',
                object_id=product.id,
                object_repr=str(product),
                extra_data={'reason': reason, 'quantity_changed': quantity_changed},
                request=request
            )

            messages.success(request, "Product updated successfully.")
            return redirect('product_list')
        else:
            messages.error(request, "Error updating the product. Please correct the form.")
    else:
        form = ProductForm(instance=product)

    return render(request, 'product/edit_product.html', {'form': form, 'product': product})



@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        reason = request.POST.get('reason', '').strip()

        if not reason:
            messages.error(request, "A reason for deleting the product is required.")
        else:
            # Log the delete action with None as quantity_changed
            ProductHistory.objects.create(
                product=product,
                user=request.user,
                action='DELETE',
                reason=reason,
                quantity_changed=None  # No quantity change on deletion
            )

            # Log to activity log before deleting
            product_info = str(product)
            product_id = product.id
            ActivityLog.log_activity(
                user=request.user,
                action='product_delete',
                description=f'Deleted product: {product_info} - Reason: {reason}',
                model_name='Product',
                object_id=product_id,
                object_repr=product_info,
                extra_data={'reason': reason},
                request=request
            )

            product.delete()
            messages.success(request, "Product deleted successfully.")
            return redirect('product_list')

    return render(request, 'product/delete_product.html', {'product': product})


@login_required(login_url='login')
def transfer_to_warehouse_view(request):
    """Bulk transfer from shop floor to warehouse"""
    try:
        current_location = request.user.profile.location
    except:
        current_location = 'ABUJA'  # Fallback

    # Get filters
    search = request.GET.get('search', '')
    category = request.GET.get('category', '')
    size = request.GET.get('size', '')
    color = request.GET.get('color', '')
    design = request.GET.get('design', '')
    min_quantity = request.GET.get('min_quantity', '')
    max_quantity = request.GET.get('max_quantity', '')

    # Filter products at current location - ONLY SHOP FLOOR ITEMS
    products = Product.objects.filter(location=current_location, quantity__gt=0, shop='STORE')

    if search:
        products = products.filter(
            Q(brand__icontains=search) |
            Q(category__icontains=search) |
            Q(size__icontains=search) |
            Q(color__icontains=search) |
            Q(design__icontains=search)
        )

    if category:
        products = products.filter(category=category)
    if size:
        products = products.filter(size__icontains=size)
    if color:
        products = products.filter(color=color)
    if design:
        products = products.filter(design=design)
    if min_quantity:
        products = products.filter(quantity__gte=min_quantity)
    if max_quantity:
        products = products.filter(quantity__lte=max_quantity)

    products = products.order_by('brand', 'category', 'size')

    # Pagination
    paginator = Paginator(products, 100)
    page_number = request.GET.get('page')
    products_page = paginator.get_page(page_number)

    # Unique values for dropdowns
    categories = Product.objects.filter(location=current_location, shop='STORE').values_list('category', flat=True).distinct()
    sizes = Product.objects.filter(location=current_location, shop='STORE').values_list('size', flat=True).distinct()
    colors = Product.objects.filter(location=current_location, shop='STORE').values_list('color', flat=True).distinct()
    designs = Product.objects.filter(location=current_location, shop='STORE').values_list('design', flat=True).distinct()

    has_filters = any([search, category, size, color, design, min_quantity, max_quantity])

    # Handle POST
    if request.method == 'POST':
        selected_products = []
        has_errors = False

        # Get selected products data
        selected_products_data = request.POST.get('selected_products_data', '')
        notes = request.POST.get('notes', '')

        print(f"DEBUG: POST request received")
        print(f"DEBUG: selected_products_data = {selected_products_data}")
        print(f"DEBUG: notes = {notes}")

        if selected_products_data:
            try:
                import json
                frontend_selections = json.loads(selected_products_data)

                # Validate and collect products
                for product_id, item_data in frontend_selections.items():
                    try:
                        product = Product.objects.get(id=product_id, location=current_location, shop='STORE')
                        qty = int(item_data.get('quantity', 1))

                        if qty <= 0:
                            messages.error(request, f"Invalid quantity for {product.brand}.")
                            has_errors = True
                        elif qty > product.quantity:
                            messages.error(request,
                                           f"Not enough stock for {product.brand}. Available: {product.quantity}, Requested: {qty}")
                            has_errors = True
                        else:
                            selected_products.append({'product': product, 'quantity': qty})
                    except Product.DoesNotExist:
                        messages.error(request, f"Product with ID {product_id} not found on shop floor.")
                        has_errors = True
                    except (ValueError, TypeError):
                        messages.error(request, f"Invalid quantity data for product ID {product_id}.")
                        has_errors = True

            except json.JSONDecodeError:
                messages.error(request, "Invalid selection data format.")
                has_errors = True

        if not selected_products and not has_errors:
            print(f"DEBUG: No products selected")
            messages.error(request, "Please select at least one product to transfer to warehouse.")
            has_errors = True

        if not selected_products_data:
            print(f"DEBUG: selected_products_data is empty!")
            messages.error(request, "No selection data received. Please try selecting products again.")
            has_errors = True

        if not has_errors:
            try:
                with transaction.atomic():
                    # Create transfer record
                    transfer = LocationTransfer.objects.create(
                        transfer_type='internal',
                        from_shop='STORE',
                        to_shop='WAREHOUSE',
                        internal_location=current_location,
                        transfer_reference=LocationTransfer.generate_transfer_reference(
                            transfer_type='internal',
                            from_shop='STORE',
                            to_shop='WAREHOUSE'
                        ),
                        created_by=request.user,
                        notes=notes,
                        status='COMPLETED'
                    )

                    total_items = 0
                    total_value = 0

                    for item in selected_products:
                        product = item['product']
                        quantity = item['quantity']

                        # Create transfer item record
                        TransferItem.objects.create(
                            transfer=transfer,
                            product=product,
                            quantity=quantity,
                            unit_price=product.price
                        )

                        # Move quantity to warehouse using WarehouseInventory table
                        from .models import WarehouseInventory

                        # Find or create warehouse inventory
                        warehouse_item, created = WarehouseInventory.objects.get_or_create(
                            location=current_location,
                            brand=product.brand,
                            category=product.category,
                            size=product.size,
                            color=product.color,
                            design=product.design,
                            defaults={
                                'price': product.price,
                                'markup_type': product.markup_type,
                                'markup': product.markup,
                                'selling_price': product.selling_price,
                                'quantity': 0,
                                'original_barcode': product.barcode_number  # Store barcode for reference
                            }
                        )

                        # Add quantity to warehouse
                        warehouse_item.quantity += quantity
                        # Update barcode reference if not set
                        if not warehouse_item.original_barcode and product.barcode_number:
                            warehouse_item.original_barcode = product.barcode_number
                        warehouse_item.save()

                        # Subtract from shop floor product
                        product.quantity -= quantity
                        if product.quantity <= 0:
                            # Shop floor product depleted, can delete it
                            product.delete()
                        else:
                            product.save()

                        total_items += quantity
                        total_value += Decimal(str(product.price)) * quantity

                    # Update transfer totals
                    transfer.total_items = total_items
                    transfer.total_value = total_value
                    transfer.save()

                    # Log the transfer
                    ActivityLog.log_activity(
                        user=request.user,
                        action='transfer_to_warehouse',
                        description=f'Transferred to warehouse {transfer.transfer_reference} - Shop Floor → Warehouse - {total_items} items - ₦{total_value:,.2f}',
                        model_name='LocationTransfer',
                        object_id=transfer.id,
                        object_repr=transfer.transfer_reference,
                        extra_data={
                            'from': 'Shop Floor',
                            'to': 'Warehouse',
                            'total_items': total_items,
                            'total_value': float(total_value),
                            'location': current_location
                        },
                        request=request
                    )

                    success_message = f"Transfer {transfer.transfer_reference} completed! {total_items} items moved to warehouse (₦{total_value:,.2f})"
                    messages.success(request, success_message)

                    # Clear the session storage
                    request.session['transfer_created'] = True

                    return redirect('transfer_detail', transfer_id=transfer.id)

            except Exception as e:
                import traceback
                error_msg = f"Error completing warehouse transfer: {str(e)}"
                print(error_msg)
                print(traceback.format_exc())
                messages.error(request, f"An error occurred: {str(e)}")

    # Get store config for currency symbol
    from .models import StoreConfiguration
    store_config = StoreConfiguration.get_active_config()

    context = {
        'products': products_page,
        'categories': categories,
        'sizes': sizes,
        'colors': colors,
        'designs': designs,
        'current_location': current_location,
        'has_filters': has_filters,
        'current_filters': {
            'search': search,
            'category': category,
            'size': size,
            'color': color,
            'design': design,
            'min_quantity': min_quantity,
            'max_quantity': max_quantity,
        },
        'transfer_created': request.session.pop('transfer_created', False),
        'currency_symbol': store_config.currency_symbol if store_config else '₦',
    }

    return render(request, 'transfers/transfer_to_warehouse.html', context)



@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def product_history_report(request):
    history = ProductHistory.objects.select_related('product', 'user').order_by('-date')
    return render(request, 'product/product_history_report.html', {'history': history})


@login_required(login_url='login')
def pre_order(request):
    if request.method == 'POST':
        form = PreOrderForm(request.POST)
        if form.is_valid():
            form.save()
            # Add success message
            messages.success(request, "Pre-order created successfully!")
            # Redirect to the pre_order_list page after successful form submission
            return redirect('pre_order_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = PreOrderForm()

    return render(request, 'orders/pre_order.html', {'form': form})




@login_required(login_url='login')
def pre_order_list(request):
    # Retrieve search and filter parameters from the GET request
    query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')

    # Retrieve pre-orders and allow filtering by status
    pre_orders = PreOrder.objects.all().order_by('-order_date')

    if status_filter == 'pending':
        pre_orders = pre_orders.filter(converted_to_product=False)
    elif status_filter == 'converted':
        pre_orders = pre_orders.filter(converted_to_product=True)

    if query:
        # Search filter
        pre_orders = pre_orders.filter(
            Q(brand__icontains=query) |
            Q(customer__name__icontains=query) |
            Q(quantity__icontains=query)
        )

    return render(
        request,
        'orders/pre_order_list.html',
        {
            'pre_orders': pre_orders,
            'search_query': query,
            'status_filter': status_filter,
        },
    )


@login_required(login_url='login')
def toggle_delivered(request, pre_order_id):
    # Toggle the delivered status for the pre-order
    pre_order = get_object_or_404(PreOrder, id=pre_order_id)
    pre_order.delivered = not pre_order.delivered
    pre_order.save()
    return redirect('pre_order_list')




@login_required(login_url='login')
def pre_order_detail(request, pre_order_id):
    # Get the specific pre-order by ID
    pre_order = get_object_or_404(PreOrder, id=pre_order_id)

    if request.method == 'POST':
        # Use a form to update the delivery status
        form = PreOrderStatusForm(request.POST, instance=pre_order)
        if form.is_valid():
            form.save()
            return redirect('pre_order_list')
    else:
        form = PreOrderStatusForm(instance=pre_order)

    return render(request, 'orders/pre_order_detail.html', {'pre_order': pre_order, 'form': form})


@login_required(login_url='login')
def edit_pre_order(request, pre_order_id):
    """Edit an existing pre-order"""
    pre_order = get_object_or_404(PreOrder, id=pre_order_id)

    if request.method == 'POST':
        form = PreOrderForm(request.POST, instance=pre_order)
        if form.is_valid():
            form.save()
            messages.success(request, "Pre-order updated successfully!")
            return redirect('pre_order_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = PreOrderForm(instance=pre_order)

    return render(request, 'orders/edit_pre_order.html', {'form': form, 'pre_order': pre_order})


@login_required(login_url='login')
def delete_pre_order(request, pre_order_id):
    """Delete a pre-order"""
    pre_order = get_object_or_404(PreOrder, id=pre_order_id)

    if request.method == 'POST':
        brand = pre_order.brand
        pre_order.delete()
        messages.success(request, f"Pre-order for '{brand}' has been deleted successfully!")
        return redirect('pre_order_list')

    return render(request, 'orders/confirm_delete_pre_order.html', {'pre_order': pre_order})


@login_required(login_url='login')
def convert_preorder_to_product(request, pre_order_id):
    """Convert a pre-order to an actual product and create invoice"""
    from django.utils import timezone

    pre_order = get_object_or_404(PreOrder, id=pre_order_id)

    # Check if already converted
    if pre_order.converted_to_product:
        messages.warning(request, f"This pre-order has already been converted to a product on {pre_order.conversion_date.strftime('%Y-%m-%d %H:%M')}!")
        return redirect('pre_order_list')

    # Check if pre-order is ready for conversion
    is_ready, missing_fields = pre_order.is_ready_for_conversion()

    if not is_ready:
        field_labels = {
            'brand': 'Brand',
            'price': 'Buying Price',
            'size': 'Size',
            'category': 'Category',
            'markup_type': 'Markup Type',
            'markup': 'Markup',
            'shop': 'Shop'
        }
        missing_labels = [field_labels.get(f, f.title()) for f in missing_fields]
        messages.error(
            request,
            f"⚠️ Cannot convert to product yet! Please edit the pre-order and fill in these purchase details: {', '.join(missing_labels)}"
        )
        return redirect('edit_pre_order', pre_order_id=pre_order_id)

    try:
        # Create the product - match Product model exactly
        product = Product(
            brand=pre_order.brand,
            price=pre_order.price,  # This is the buying/cost price
            color=pre_order.color or '',
            design=pre_order.design or 'plain',
            size=pre_order.size,
            category=pre_order.category,
            quantity=pre_order.quantity,
            markup_type=pre_order.markup_type,
            markup=pre_order.markup or 0,
            selling_price=pre_order.selling_price,  # Can be null, will be calculated
            shop=pre_order.shop,
            barcode_number=pre_order.barcode_number or '',
            location=pre_order.location or 'ABUJA'
        )

        # Calculate selling price if not set
        if not product.selling_price:
            product.selling_price = product.calculate_selling_price()

        product.save()

        # Create the invoice
        invoice = Invoice(user=request.user)
        invoice.save()

        # Create invoice product entry
        invoice_product = InvoiceProduct(
            invoice=invoice,
            product_name=pre_order.brand,  # Use brand as product name in invoice
            product_price=pre_order.price,  # Buying price
            product_color=pre_order.color or '',
            product_size=pre_order.size,
            product_category=pre_order.category,
            quantity=pre_order.quantity,
            total_price=pre_order.price * pre_order.quantity
        )
        invoice_product.save()

        # Update pre-order to mark as converted
        pre_order.converted_to_product = True
        pre_order.conversion_date = timezone.now()
        pre_order.created_product = product
        pre_order.created_invoice = invoice
        pre_order.save()

        messages.success(
            request,
            f"✅ Pre-order converted successfully! Product '{product.brand}' added to inventory with {product.quantity} units. Invoice #{invoice.invoice_number} created."
        )

        # Redirect to invoice detail page
        return redirect('invoice_detail', invoice_id=invoice.id)

    except Exception as e:
        messages.error(request, f"Error converting pre-order to product: {str(e)}")
        return redirect('pre_order_list')



@login_required(login_url='login')
def invoice(request):
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        if form.is_valid():
            form.save()  # Now simply save the form without manually setting the user, since it's handled in add_product view
            return redirect('invoice_list')  # Redirect to the invoice list after saving
    else:
        form = InvoiceForm()

    return render(request, 'store/invoice.html', {'form': form})



@login_required(login_url='login')
def invoice_list(request):
    # Fetch all invoices, ordered by date descending
    invoices = Invoice.objects.select_related('user').all().order_by('-date')  # Using select_related for efficiency

    # Filter by invoice number if provided
    invoice_number = request.GET.get('invoice_number', '')
    if invoice_number:
        invoices = invoices.filter(invoice_number__icontains=invoice_number)

    # Filter by start date if provided
    start_date = request.GET.get('start_date', '')
    if start_date:
        invoices = invoices.filter(date__gte=start_date)

    # Filter by end date if provided
    end_date = request.GET.get('end_date', '')
    if end_date:
        invoices = invoices.filter(date__lte=end_date)

    return render(request, 'invoice/invoice_list.html', {'invoices': invoices})



@login_required(login_url='login')
def invoice_detail(request, pk):
    user = request.user
    invoice = get_object_or_404(Invoice, pk=pk)

    # Ensure the invoice has the current user set
    invoice_products = invoice.invoice_products.all()
    total_cost = sum(invoice_product.total_price for invoice_product in invoice_products)
    total_quantity = sum(invoice_product.quantity for invoice_product in invoice_products)  # Add this line

    return render(request, 'invoice/invoice_detail.html', {
        'invoice': invoice,
        'invoice_products': invoice_products,
        'total_cost': total_cost,
        'total_quantity': total_quantity,  # Add this line
        'user': user,
    })



@login_required(login_url='login')
def export_invoice_pdf(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    invoice_products = invoice.invoice_products.all()
    total_cost = sum(invoice_product.total_price for invoice_product in invoice_products)
    total_quantity = sum(invoice_product.quantity for invoice_product in invoice_products)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="invoice_{invoice.invoice_number}.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    # Add invoice details
    p.setFont("Helvetica-Bold", 16)
    p.drawString(1 * inch, height - 1 * inch, f"Invoice #{invoice.invoice_number}")
    p.setFont("Helvetica", 12)
    p.drawString(1 * inch, height - 1.25 * inch, f"Date: {invoice.date.strftime('%B %d, %Y')}")
    p.drawString(1 * inch, height - 1.5 * inch, f"Created by: {invoice.user.username}")

    # Add table headers
    y = height - 2 * inch
    p.setFont("Helvetica-Bold", 12)
    p.drawString(1 * inch, y, "Product")
    p.drawString(4 * inch, y, "Unit Price")
    p.drawString(5.5 * inch, y, "Quantity")
    p.drawString(6.5 * inch, y, "Total")

    # Add products
    p.setFont("Helvetica", 12)
    y -= 0.25 * inch
    for item in invoice_products:
        y -= 0.25 * inch
        if y < 1 * inch:  # Create new page if needed
            p.showPage()
            y = height - 1 * inch
        p.drawString(1 * inch, y, item.product_name)
        p.drawString(4 * inch, y, f"{item.product_price:.2f}")
        p.drawString(5.5 * inch, y, str(item.quantity))
        p.drawString(6.5 * inch, y, f"{item.total_price:.2f}")

    # Add totals
    y -= 0.5 * inch
    p.setFont("Helvetica-Bold", 12)
    p.drawString(1 * inch, y, f"Total Quantity: {total_quantity}")
    p.drawString(4 * inch, y, f"Total Cost: {total_cost:.2f}")

    p.showPage()
    p.save()
    return response


@login_required(login_url='login')
def export_invoice_excel(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    invoice_products = invoice.invoice_products.all()
    total_cost = sum(invoice_product.total_price for invoice_product in invoice_products)
    total_quantity = sum(invoice_product.quantity for invoice_product in invoice_products)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="invoice_{invoice.invoice_number}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active

    # Sanitize sheet title by removing invalid characters
    sheet_title = f"Invoice {invoice.invoice_number}"
    # Replace invalid characters with underscores
    invalid_chars = ['/', '\\', '*', '?', ':', '[', ']']
    for char in invalid_chars:
        sheet_title = sheet_title.replace(char, '_')
    # Truncate if longer than 31 characters (Excel limit)
    sheet_title = sheet_title[:31]
    ws.title = sheet_title

    # Add headers
    headers = ['Product', 'Unit Price', 'Quantity', 'Total']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Add data
    for row_num, item in enumerate(invoice_products, 2):
        ws.cell(row=row_num, column=1).value = item.product_name
        ws.cell(row=row_num, column=2).value = float(item.product_price)
        ws.cell(row=row_num, column=3).value = item.quantity
        ws.cell(row=row_num, column=4).value = float(item.total_price)

    # Add totals
    last_row = len(invoice_products) + 3
    ws.cell(row=last_row, column=2).value = "Total Quantity:"
    ws.cell(row=last_row, column=3).value = total_quantity
    ws.cell(row=last_row + 1, column=2).value = "Total Cost:"
    ws.cell(row=last_row + 1, column=3).value = float(total_cost)

    # Format currency cells
    for row in ws.iter_rows(min_row=2, max_row=len(invoice_products) + 1, min_col=2, max_col=4):
        for cell in row:
            if cell.column in [2, 4]:  # Price columns
                cell.number_format = '""#,##0.00'

    wb.save(response)
    return response

@login_required(login_url='login')
def goods_received(request):
    if request.method == 'POST':
        form = GoodsReceivedForm(request.POST)
        if form.is_valid():
            goods_received = form.save()
            # Optional: Update product stock
            product = goods_received.product
            product.quantity += goods_received.quantity_received
            product.save()

            messages.success(request, f"✅ {goods_received.quantity_received} units of {product.brand} received (Batch: {goods_received.batch_number}).")
            return redirect('goods_received')
    else:
        form = GoodsReceivedForm()

    # Get recent receipts (last 10)
    recent_receipts = GoodsReceived.objects.select_related('product').order_by('-received_date')[:10]

    return render(request, 'store/goods_received.html', {
        'form': form,
        'recent_receipts': recent_receipts
    })


@login_required(login_url='login')
def delivery(request):
    if request.method == 'POST':
        form = DeliveryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('delivery')
    else:
        form = DeliveryForm()
    return render(request, 'store/delivery.html', {'form': form})


def generate_barcodes_view(request):
    """Display products without barcodes and handle barcode generation"""
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'generate_all':
            # Generate barcodes for all products without them
            return _generate_barcodes_bulk(request)

        elif action == 'generate_selected':
            # Generate barcodes for selected products
            selected_ids = request.POST.getlist('selected_products')
            if not selected_ids:
                messages.warning(request, 'Please select at least one product.')
                return redirect('generate_barcodes')

            return _generate_barcodes_for_products(request, selected_ids)

    # GET request - display the form
    products_without_barcodes = Product.objects.filter(
        barcode_image__isnull=True
    ) | Product.objects.filter(
        barcode_number__isnull=True
    )

    context = {
        'products': products_without_barcodes,
        'total_count': products_without_barcodes.count()
    }

    return render(request, 'barcode/generate_barcodes.html', context)


def _generate_barcodes_bulk(request):
    """Generate barcodes for all products without them"""
    products = Product.objects.filter(
        barcode_image__isnull=True
    ) | Product.objects.filter(
        barcode_number__isnull=True
    )

    success_count = 0
    error_count = 0

    for product in products:
        if _generate_single_barcode(product):
            success_count += 1
        else:
            error_count += 1

    if success_count > 0:
        messages.success(request, f'Successfully generated {success_count} barcodes.')
    if error_count > 0:
        messages.error(request, f'Failed to generate {error_count} barcodes.')

    return redirect('generate_barcodes')


def _generate_barcodes_for_products(request, product_ids):
    """Generate barcodes for specific products"""
    success_count = 0
    error_count = 0

    for product_id in product_ids:
        try:
            product = Product.objects.get(id=product_id)
            if _generate_single_barcode(product):
                success_count += 1
            else:
                error_count += 1
        except Product.DoesNotExist:
            error_count += 1

    if success_count > 0:
        messages.success(request, f'Successfully generated {success_count} barcodes.')
    if error_count > 0:
        messages.error(request, f'Failed to generate {error_count} barcodes.')

    return redirect('generate_barcodes')


def _generate_single_barcode(product):
    """Generate barcode for a single product"""
    try:
        # Generate barcode number if it doesn't exist
        if not product.barcode_number:
            base_number = str(product.id).zfill(12)
            check_digit = product._calculate_ean13_check_digit(base_number)
            product.barcode_number = base_number + str(check_digit)

        # Generate barcode image
        code128 = barcode.Code128(product.barcode_number, writer=ImageWriter())
        buffer = BytesIO()
        code128.write(buffer)
        filename = f'{product.brand}_{product.barcode_number}.png'

        buffer.seek(0)
        product.barcode_image.save(filename, buffer, save=False)
        product.save(update_fields=['barcode_image', 'barcode_number'])

        return True

    except Exception as e:
        return False


# AJAX view for single product barcode generation
def generate_single_barcode_ajax(request, product_id):
    """AJAX view to generate barcode for a single product"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

    try:
        product = get_object_or_404(Product, id=product_id)

        if _generate_single_barcode(product):
            return JsonResponse({
                'success': True,
                'message': f'Barcode generated for {product.name}',
                'barcode_number': product.barcode_number,
                'barcode_url': product.barcode_image.url if product.barcode_image else None
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Failed to generate barcode'
            })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


# Alternative view for immediate redirect (your original approach)
def generate_barcodes_redirect_view(request):
    """Original view that generates all barcodes and redirects immediately"""
    products = Product.objects.filter(
        barcode_image__isnull=True
    ) | Product.objects.filter(
        barcode_number__isnull=True
    )

    for product in products:
        if not product.barcode_number:
            base_number = str(product.id).zfill(12)
            check_digit = product._calculate_ean13_check_digit(base_number)
            product.barcode_number = base_number + str(check_digit)

        try:
            code128 = barcode.Code128(product.barcode_number, writer=ImageWriter())
            buffer = BytesIO()
            code128.write(buffer)
            filename = f'{product.brand}_{product.barcode_number}.png'
            buffer.seek(0)
            product.barcode_image.save(filename, buffer, save=False)
            product.save(update_fields=['barcode_image', 'barcode_number'])
        except Exception as e:
            pass

    return redirect(reverse('product_list'))


@login_required(login_url='login')
def upload_products_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            overwrite = form.cleaned_data['overwrite_existing']

            try:
                # Read the Excel file
                df = pd.read_excel(excel_file)

                # Validate required columns
                required_columns = ['brand', 'price', 'category']
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    messages.error(request, f"Missing required columns: {', '.join(missing_columns)}")
                    return render(request, 'product/upload_products.html', {'form': form})

                # Helper function to get valid choices for validation
                def get_valid_categories():
                    return [choice[0] for choice in ProductChoices.CATEGORY_CHOICES]

                def get_valid_colors():
                    valid_colors = []
                    for family in ProductChoices.COLOR_CHOICES:
                        if isinstance(family[1], tuple):
                            valid_colors.extend([choice[0] for choice in family[1]])
                        else:
                            valid_colors.append(family[0])
                    return valid_colors

                def get_valid_designs():
                    valid_designs = []
                    for family in ProductChoices.DESIGN_CHOICES:
                        if isinstance(family[1], tuple):
                            valid_designs.extend([choice[0] for choice in family[1]])
                        else:
                            valid_designs.append(family[0])
                    return valid_designs

                # Process each row
                success_count = 0
                error_count = 0
                errors = []
                new_choices_added = {'colors': set(), 'designs': set(), 'categories': set()}

                with transaction.atomic():
                    for index, row in df.iterrows():
                        try:
                            # Check if product with barcode already exists
                            barcode = row.get('barcode_number', None)
                            product = None

                            if barcode and pd.notna(barcode):
                                barcode = str(barcode).strip()
                                if barcode:  # Only check if barcode is not empty
                                    existing_product = Product.objects.filter(barcode_number=barcode).first()
                                    if existing_product:
                                        if overwrite:
                                            product = existing_product
                                        else:
                                            errors.append(
                                                f"Row {index + 2}: Product with barcode {barcode} already exists")
                                            error_count += 1
                                            continue

                            if not product:
                                product = Product()

                            # Map Excel columns to model fields with flexible validation
                            product.brand = str(row['brand']).strip() if pd.notna(row.get('brand')) else ''
                            if not product.brand:
                                errors.append(f"Row {index + 2}: Brand is required")
                                error_count += 1
                                continue

                            # Price validation
                            try:
                                product.price = float(row['price']) if pd.notna(row.get('price')) else 0.0
                                if product.price < 0:
                                    errors.append(f"Row {index + 2}: Price cannot be negative")
                                    error_count += 1
                                    continue
                            except (ValueError, TypeError):
                                errors.append(f"Row {index + 2}: Invalid price format")
                                error_count += 1
                                continue

                            # FLEXIBLE Category validation - accept existing or new values
                            category = str(row['category']).strip() if pd.notna(row.get('category')) else 'Apparel'
                            valid_categories = get_valid_categories()
                            if category not in valid_categories:
                                # Accept new category value and track it
                                new_choices_added['categories'].add(category)
                            product.category = category

                            # FLEXIBLE Color validation - accept existing or new values
                            if pd.notna(row.get('color')):
                                color = str(row['color']).strip()
                                valid_colors = get_valid_colors()
                                if color and color not in valid_colors:
                                    # Accept new color value and track it
                                    new_choices_added['colors'].add(color)
                                product.color = color if color else None
                            else:
                                product.color = None

                            # FLEXIBLE Design validation - accept existing or new values
                            if pd.notna(row.get('design')):
                                design = str(row['design']).strip()
                                valid_designs = get_valid_designs()
                                if design and design not in valid_designs:
                                    # Accept new design value and track it
                                    new_choices_added['designs'].add(design)
                                product.design = design if design else 'plain'
                            else:
                                product.design = 'plain'

                            product.size = str(row['size']).strip() if pd.notna(row.get('size')) else ''

                            # Quantity validation
                            try:
                                product.quantity = int(row['quantity']) if pd.notna(row.get('quantity')) else 0
                                if product.quantity < 0:
                                    errors.append(f"Row {index + 2}: Quantity cannot be negative")
                                    error_count += 1
                                    continue
                            except (ValueError, TypeError):
                                errors.append(f"Row {index + 2}: Invalid quantity format")
                                error_count += 1
                                continue

                            # Markup type validation (still strict as it affects calculations)
                            markup_type = str(row['markup_type']).strip() if pd.notna(
                                row.get('markup_type')) else 'percentage'
                            valid_markup_types = [choice[0] for choice in ProductChoices.MARKUP_TYPE_CHOICES]
                            if markup_type not in valid_markup_types:
                                errors.append(
                                    f"Row {index + 2}: Invalid markup_type '{markup_type}'. Valid options: {', '.join(valid_markup_types)}")
                                error_count += 1
                                continue
                            product.markup_type = markup_type

                            # Markup validation
                            try:
                                product.markup = float(row['markup']) if pd.notna(row.get('markup')) else 0.0
                                if product.markup < 0:
                                    errors.append(f"Row {index + 2}: Markup cannot be negative")
                                    error_count += 1
                                    continue
                            except (ValueError, TypeError):
                                errors.append(f"Row {index + 2}: Invalid markup format")
                                error_count += 1
                                continue

                            # Shop validation (still strict as it affects business logic)
                            shop = str(row['shop']).strip() if pd.notna(row.get('shop')) else 'STORE'
                            valid_shops = [choice[0] for choice in ProductChoices.SHOP_TYPE]
                            if shop not in valid_shops:
                                errors.append(
                                    f"Row {index + 2}: Invalid shop '{shop}'. Valid options: {', '.join(valid_shops)}")
                                error_count += 1
                                continue
                            product.shop = shop

                            # Handle optional barcode field
                            if pd.notna(row.get('barcode_number')):
                                barcode_value = str(row['barcode_number']).strip()
                                product.barcode_number = barcode_value if barcode_value else None
                            else:
                                product.barcode_number = None

                            # Save without full_clean() to bypass Django's choice validation
                            # We'll handle our own validation above
                            product.save()
                            success_count += 1

                        except Exception as e:
                            errors.append(f"Row {index + 2}: {str(e)}")
                            error_count += 1

                # Prepare results message
                if success_count > 0:
                    messages.success(request, f"Successfully processed {success_count} products")

                # Show new choices that were added
                if any(new_choices_added.values()):
                    new_items = []
                    if new_choices_added['categories']:
                        new_items.append(f"Categories: {', '.join(new_choices_added['categories'])}")
                    if new_choices_added['colors']:
                        new_items.append(f"Colors: {', '.join(new_choices_added['colors'])}")
                    if new_choices_added['designs']:
                        new_items.append(f"Designs: {', '.join(new_choices_added['designs'])}")

                    messages.info(request, f"New choices added from Excel: {' | '.join(new_items)}")
                    messages.warning(request,
                                     "Note: These new choices are not in your predefined lists. Consider updating your model choices if you want them to appear in dropdowns.")

                if error_count > 0:
                    messages.warning(request, f"Failed to process {error_count} products. Check errors below.")
                    for error in errors[:10]:  # Limit to first 10 errors to avoid overwhelming
                        messages.error(request, error)
                    if len(errors) > 10:
                        messages.warning(request, f"... and {len(errors) - 10} more errors")

                return redirect('product_list')

            except Exception as e:
                messages.error(request, f"Error processing Excel file: {str(e)}")
    else:
        form = ExcelUploadForm()

    return render(request, 'product/upload_products.html', {'form': form})


@login_required(login_url='login')
def download_excel_template(request):
    """Download Excel template with sample data and validation comments"""
    # Create a dataframe with comprehensive sample data
    sample_data = {
        'brand': ['Nike Air Max', 'Adidas Ultraboost', 'Puma Classic'],
        'price': [99.99, 129.99, 79.99],
        'color': ['black', 'navy', 'white'],
        'design': ['plain', 'geometric', 'solid'],
        'size': ['M', 'L', '42'],
        'category': ['Footwear', 'Footwear', 'Footwear'],
        'quantity': [25, 15, 30],
        'markup_type': ['percentage', 'fixed', 'percentage'],
        'markup': [20.0, 15.0, 25.0],
        'shop': ['STORE', 'STORE', 'STORE'],
        'barcode_number': ['', '1234567890123', '']
    }

    df = pd.DataFrame(sample_data)

    # Create Excel file in memory with formatting
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write main data
        df.to_excel(writer, sheet_name='Products', index=False, startrow=1)

        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Products']

        # Add title
        worksheet['A1'] = 'Product Upload Template - Replace sample data with your products'
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet['A1'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # Format headers
        header_row = 2
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=header_row, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Add validation sheet
        validation_data = {
            'Field': ['brand', 'price', 'color', 'design', 'size', 'category', 'quantity', 'markup_type', 'markup',
                      'shop', 'barcode_number'],
            'Required': ['Yes', 'Yes', 'No', 'No', 'No', 'Yes', 'No', 'No', 'No', 'No', 'No'],
            'Valid_Values': [
                'Any text',
                'Positive number',
                'black, navy, white, etc. (see Color Options sheet)',
                'plain, solid, geometric, etc. (see Design Options sheet)',
                'Any text (e.g., S, M, L, XL, or shoe sizes)',
                'Apparel, Footwear, Accessories',
                'Non-negative integer',
                'percentage, fixed',
                'Non-negative number',
                'STORE',
                'Optional unique barcode number'
            ]
        }

        validation_df = pd.DataFrame(validation_data)
        validation_df.to_excel(writer, sheet_name='Field_Validation', index=False)

        # Format validation sheet
        val_sheet = writer.sheets['Field_Validation']
        for col_num, column_title in enumerate(validation_df.columns, 1):
            cell = val_sheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

        # Add color options sheet
        color_options = []
        for family in Product.COLOR_CHOICES:
            if isinstance(family[1], tuple):
                for color in family[1]:
                    color_options.append({'Family': family[0], 'Color_Code': color[0], 'Display_Name': color[1]})

        color_df = pd.DataFrame(color_options)
        color_df.to_excel(writer, sheet_name='Color_Options', index=False)

        # Add design options sheet
        design_options = []
        for family in Product.DESIGN_CHOICES:
            if isinstance(family[1], tuple):
                for design in family[1]:
                    design_options.append({'Family': family[0], 'Design_Code': design[0], 'Display_Name': design[1]})

        design_df = pd.DataFrame(design_options)
        design_df.to_excel(writer, sheet_name='Design_Options', index=False)

        # Auto-adjust column widths
        for sheet_name in writer.sheets:
            sheet = writer.sheets[sheet_name]
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Prepare response
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=product_upload_template.xlsx'
    return response


@login_required(login_url='login')
def export_products_excel(request):
    """Export filtered products to Excel"""
    # Get the same filters as in product_list view
    products = filter_products(request, Product.objects.all())

    # Apply all the same filters from product_list
    query = request.GET.get('search', '')
    category = request.GET.get('category', '')
    shop = request.GET.get('shop', '')
    size = request.GET.get('size', '')
    color = request.GET.get('color', '')
    design = request.GET.get('design', '')
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')
    min_quantity = request.GET.get('min_quantity', '')
    max_quantity = request.GET.get('max_quantity', '')

    # Apply filters (same logic as product_list)
    if query:
        products = products.filter(
            Q(brand__icontains=query) |
            Q(color__icontains=query) |
            Q(category__icontains=query) |
            Q(design__icontains=query) |
            Q(size__icontains=query)
        )
    if category:
        products = products.filter(category=category)
    if shop:
        products = products.filter(shop=shop)
    if size:
        products = products.filter(size__icontains=size)
    if color:
        products = products.filter(color=color)
    if design:
        products = products.filter(design=design)
    if min_price:
        products = products.filter(price__gte=min_price)
    if max_price:
        products = products.filter(price__lte=max_price)
    if min_quantity:
        products = products.filter(quantity__gte=min_quantity)
    if max_quantity:
        products = products.filter(quantity__lte=max_quantity)

    # Convert to DataFrame
    data = []
    for product in products:
        data.append({
            'Brand': product.brand,
            'Category': product.category,
            'Color': product.get_display_color() if product.color else '',
            'Design': product.get_display_design() if product.design else '',
            'Size': product.size,
            'Price': product.price,
            'Markup Type': product.get_markup_type_display(),
            'Markup': product.markup,
            'Selling Price': product.selling_price,
            'Quantity': product.quantity,
            'Shop': product.shop,
            'Barcode': product.barcode_number or '',
        })

    df = pd.DataFrame(data)

    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Products', index=False)

        # Format the Excel file
        workbook = writer.book
        worksheet = writer.sheets['Products']

        # Add title and timestamp
        worksheet.insert_rows(1, 2)
        worksheet['A1'] = f'Product Export - Generated on {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet['A2'] = f'Total Products: {len(data)}'
        worksheet['A2'].font = Font(bold=True)

        # Format headers (now in row 3)
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=3, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    output.seek(0)

    # Generate filename with timestamp
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f'products_export_{timestamp}.xlsx'

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@login_required(login_url='login')
def export_products_pdf(request):
    """Export filtered products to PDF"""
    # Get filtered products (same logic as product_list)
    products = filter_products(request, Product.objects.all())

    # Apply additional filters
    query = request.GET.get('search', '')
    category = request.GET.get('category', '')
    shop = request.GET.get('shop', '')
    size = request.GET.get('size', '')
    color = request.GET.get('color', '')
    design = request.GET.get('design', '')
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')
    min_quantity = request.GET.get('min_quantity', '')
    max_quantity = request.GET.get('max_quantity', '')

    if query:
        products = products.filter(
            Q(brand__icontains=query) |
            Q(color__icontains=query) |
            Q(category__icontains=query) |
            Q(design__icontains=query) |
            Q(size__icontains=query)
        )
    if category:
        products = products.filter(category=category)
    if shop:
        products = products.filter(shop=shop)
    if size:
        products = products.filter(size__icontains=size)
    if color:
        products = products.filter(color=color)
    if design:
        products = products.filter(design=design)
    if min_price:
        products = products.filter(price__gte=min_price)
    if max_price:
        products = products.filter(price__lte=max_price)
    if min_quantity:
        products = products.filter(quantity__gte=min_quantity)
    if max_quantity:
        products = products.filter(quantity__lte=max_quantity)

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=50,
        bottomMargin=50,
        title="Product Inventory Report"
    )

    styles = getSampleStyleSheet()
    story = []

    # --- Title ---
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue,
        fontName='Helvetica-Bold'
    )
    story.append(Paragraph("Product Inventory Report", title_style))
    story.append(Spacer(1, 6))

    # --- Summary ---
    total_products = products.count()
    total_value = sum(p.price * p.quantity for p in products)
    total_selling_value = sum(p.selling_price * p.quantity for p in products)

    summary_data = [
        ['Report Generated:', timezone.now().strftime("%Y-%m-%d %H:%M:%S")],
        ['Total Products:', str(total_products)],
        ['Total Inventory Value (Cost):', f'{total_value:,.2f}'],
        ['Total Inventory Value (Selling):', f'{total_selling_value:,.2f}'],
    ]

    summary_table = Table(summary_data, colWidths=[2.8 * inch, 3.0 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 24))

    # --- Products Table ---
    if products.exists():
        # Prepare table data with headers
        data = [['Brand', 'Category', 'Size', 'Color', 'Design', 'Price', 'Qty', 'Sell Price', 'Total Value']]

        for product in products:  # Removed limit
            total_value = product.selling_price * product.quantity
            data.append([
                Paragraph(str(product.brand), styles['Normal']),
                Paragraph(str(product.category), styles['Normal']),
                Paragraph(str(product.size), styles['Normal']),
                Paragraph(str(product.color), styles['Normal']),
                Paragraph(str(product.design), styles['Normal']),
                Paragraph(f'{product.price:,.2f}', styles['Normal']),
                Paragraph(str(product.quantity), styles['Normal']),
                Paragraph(f'{product.selling_price:,.2f}', styles['Normal']),
                Paragraph(f'{total_value:,.2f}', styles['Normal']),
            ])

        # Calculate available width
        available_width = A4[0] - 80  # page width minus margins
        col_widths = [
            available_width * 0.18,  # Brand
            available_width * 0.14,  # Category
            available_width * 0.10,  # Size
            available_width * 0.10,  # Color
            available_width * 0.12,  # Design
            available_width * 0.09,  # Price
            available_width * 0.07,  # Qty
            available_width * 0.10,  # Sell Price
            available_width * 0.10,  # Total Value
        ]

        table = Table(data, colWidths=col_widths, repeatRows=1)

        # Style the table
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, 0), 8),

            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 1), (-1, -1), 6),

            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),

            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ]))

        story.append(table)

    else:
        no_products = Paragraph("No products found matching the current filters.", styles['Italic'])
        story.append(no_products)

    # --- Build PDF ---
    try:
        doc.build(story)
    except Exception as e:
        logger.error(f"PDF generation error: {e}")
        return HttpResponse("Error generating PDF.", status=500)

    buffer.seek(0)

    # Generate filename
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f'products_report_{timestamp}.pdf'

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    buffer.close()

    return response


@login_required(login_url='login')
def reports_menu(request):
    """Reports menu page showing all available reports"""
    return render(request, 'reports/reports_menu.html')

@login_required(login_url='login')
def user_menu(request):
    """User management menu page"""
    return render(request, 'users/user_menu.html')

@login_required(login_url='login')
def tools_menu(request):
    """Tools and utilities menu page"""
    return render(request, 'tools/tools_menu.html')

@login_required(login_url='login')
def inventory_menu(request):
    """Inventory management menu page"""
    return render(request, 'inventory/inventory_menu.html')

'''========================REPORT=============================='''
@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def sales_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    payment_method_filter = request.GET.get('payment_method')
    shop_type = request.GET.get('shop_type')
    export_format = request.GET.get('export')

    # Base queryset with select/prefetch
    sales = Sale.objects.select_related(
        'product', 'receipt', 'payment', 'delivery', 'customer'
    ).prefetch_related(
        'payment__payment_methods'
    ).order_by('-sale_date')

    # Apply filters - default to today if no dates provided
    start_date_obj = None
    end_date_obj = None
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            # Include the entire end date by adding 1 day and subtracting 1 second
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            sales = sales.filter(sale_date__range=[start_date_obj, end_date_obj])
        except ValueError:
            start_date_obj = None
            end_date_obj = None
    else:
        # Default to today's date if no filters applied
        today = timezone.now().date()
        start_date_obj = datetime.combine(today, datetime.min.time())
        end_date_obj = datetime.combine(today, datetime.max.time())
        sales = sales.filter(sale_date__range=[start_date_obj, end_date_obj])
        start_date = today.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    if shop_type:
        sales = sales.filter(product__shop=shop_type)

    if payment_method_filter:
        # Filter by specific payment method
        sales = sales.filter(payment__payment_methods__payment_method=payment_method_filter).distinct()

    # Group sales by receipt
    grouped_sales = defaultdict(list)
    for sale in sales:
        grouped_sales[sale.receipt].append(sale)

    # Convert to list of tuples for easy template iteration
    grouped_sales = sorted(
        grouped_sales.items(),
        key=lambda x: x[1][0].sale_date if x[1] else timezone.now(),
        reverse=True
    )

    # Calculate total sales using receipt.total_with_delivery (avoid double counting)
    # Get unique receipts from filtered sales
    unique_receipts = Receipt.objects.filter(sales__in=sales).distinct()
    total_sales = sum(receipt.total_with_delivery for receipt in unique_receipts if receipt.total_with_delivery)

    payment_methods = PaymentMethod.PAYMENT_METHODS
    shop_types = ProductChoices.SHOP_TYPE

    if export_format:
        if export_format == 'excel':
            return export_to_excel(grouped_sales, start_date_obj, end_date_obj, total_sales)
        elif export_format == 'pdf':
            return export_to_pdf(grouped_sales, start_date_obj, end_date_obj, total_sales)

    # Calculate discount and delivery totals for the report
    total_payment_discounts = 0
    total_line_discounts = 0
    total_delivery_fees = 0

    for receipt, sale_list in grouped_sales:
        # Payment discounts
        first_sale = sale_list[0] if sale_list else None
        if first_sale and first_sale.payment and first_sale.payment.discount_amount:
            total_payment_discounts += first_sale.payment.discount_amount

        # Line discounts
        for sale in sale_list:
            if sale.discount_amount:
                total_line_discounts += sale.discount_amount

        # Delivery fees
        if receipt and receipt.delivery_cost:
            total_delivery_fees += receipt.delivery_cost

    return render(request, 'sales/sales_report.html', {
        'grouped_sales': grouped_sales,
        'start_date': start_date,
        'end_date': end_date,
        'total_sales': total_sales,
        'payment_methods': payment_methods,
        'selected_payment_method': payment_method_filter,
        'shop_types': shop_types,
        'selected_shop_type': shop_type,
        'total_payment_discounts': total_payment_discounts,
        'total_line_discounts': total_line_discounts,
        'total_delivery_fees': total_delivery_fees,
        'total_all_discounts': total_payment_discounts + total_line_discounts,
        'currency_symbol': '₦'
    })


def export_to_pdf(grouped_sales, start_date, end_date, total_sales):
    buffer = io.BytesIO()
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch

    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(1 * inch, height - 1 * inch, "Sales Report")

    # Date range
    p.setFont("Helvetica", 12)
    date_text = "All Dates"
    if start_date and end_date:
        date_text = f"From {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    p.drawString(1 * inch, height - 1.2 * inch, date_text)

    # Table headers
    data = [['Receipt', 'Item', 'Qty', 'Item Total', 'Subtotal/Disc/Deliv', 'Payment', 'Date']]

    # Track which rows are receipt headers
    receipt_header_rows = []  # Will store row indices

    row_index = 1  # Start after header
    for receipt, sale_list in grouped_sales:
        sale = sale_list[0]

        # Get receipt totals
        subtotal = receipt.subtotal if receipt.subtotal else 0
        discount = sale.payment.discount_amount if sale.payment and sale.payment.discount_amount else 0
        delivery = receipt.delivery_cost if receipt.delivery_cost else 0
        receipt_total = receipt.total_with_delivery if receipt.total_with_delivery else 0

        # Add receipt header row
        data.append([
            f"Receipt #{receipt.receipt_number}",
            "",
            "",
            "",
            f"Total: ₦{receipt_total:.2f}",
            "",
            sale.sale_date.strftime("%m/%d %H:%M")
        ])
        receipt_header_rows.append(row_index)
        row_index += 1

        # Add each item
        for item in sale_list:
            payment_method = ""
            if item.payment:
                payment_methods = item.payment.payment_methods.all()
                payment_method = ", ".join([pm.get_payment_method_display() for pm in payment_methods])

            data.append([
                "",
                item.product.brand,
                str(item.quantity),
                f"₦{item.total_price:.2f}",
                "",  # Breakdown only on summary rows
                payment_method,
                ""
            ])
            row_index += 1

        # Add receipt breakdown rows
        data.append(["", "", "", "Subtotal:", f"₦{subtotal:.2f}", "", ""])
        row_index += 1

        if discount > 0:
            discount_pct = sale.payment.discount_percentage if sale.payment else 0
            data.append(["", "", "", f"Discount ({discount_pct}%):", f"-₦{discount:.2f}", "", ""])
            row_index += 1

        if delivery > 0:
            data.append(["", "", "", "Delivery:", f"₦{delivery:.2f}", "", ""])
            row_index += 1

        # Add blank spacer row
        data.append(["", "", "", "", "", "", ""])
        row_index += 1

    # Total row
    data.append([
        "TOTAL SALES",
        "",
        "",
        "",
        f"₦{total_sales:.2f}",
        "",
        ""
    ])
    row_index += 1

    # Create table with updated column widths
    # [Receipt, Item, Qty, Item Total, Receipt Total, Payment, Date]
    col_widths = [90, 120, 35, 70, 70, 90, 65]
    table = Table(data, colWidths=col_widths, repeatRows=1)

    # Build style commands
    style_commands = [
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ]

    # Highlight receipt header rows
    for row in receipt_header_rows:
        style_commands.append(('FONTNAME', (0, row), (6, row), 'Helvetica-Bold'))
        style_commands.append(('BACKGROUND', (0, row), (6, row), colors.beige))
        style_commands.append(('BOTTOMPADDING', (0, row), (6, row), 6))

    # Apply all styles
    table.setStyle(TableStyle(style_commands))

    # Position table
    table.wrapOn(p, width, height)
    y = height - 2 * inch
    table.drawOn(p, 0.5 * inch, y - table._height)

    # Finalize
    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


def export_to_excel(grouped_sales, start_date, end_date, total_sales):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    headers = ['Receipt', 'Date', 'Customer', 'Product', 'Qty', 'Item Total', 'Subtotal', 'Discount', 'Delivery', 'Total', 'Payment Methods', 'Status']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    for receipt, sale_list in grouped_sales:
        # Get first sale for metadata
        sale = sale_list[0]

        # Get receipt breakdown
        subtotal = float(receipt.subtotal) if receipt.subtotal else 0
        discount = float(sale.payment.discount_amount) if sale.payment and sale.payment.discount_amount else 0
        discount_pct = float(sale.payment.discount_percentage) if sale.payment and sale.payment.discount_percentage else 0
        delivery = float(receipt.delivery_cost) if receipt.delivery_cost else 0
        receipt_total = float(receipt.total_with_delivery) if receipt.total_with_delivery else 0

        # Format payment methods
        if sale.payment:
            payment_text = ", ".join([
                f"{pm.get_payment_method_display()}(₦{pm.amount:.2f})"
                for pm in sale.payment.payment_methods.all()
            ])
            status = sale.payment.get_payment_status_display()
        else:
            payment_text = "No Payment"
            status = "N/A"

        customer_name = sale.customer.name if sale.customer else "N/A"

        # Write each item in the receipt
        for idx, item in enumerate(sale_list):
            ws.cell(row=row_num, column=1, value=f"#{receipt.receipt_number}")
            ws.cell(row=row_num, column=2, value=item.sale_date.strftime("%Y-%m-%d %H:%M"))
            ws.cell(row=row_num, column=3, value=customer_name)
            ws.cell(row=row_num, column=4, value=item.product.brand)
            ws.cell(row=row_num, column=5, value=item.quantity)
            ws.cell(row=row_num, column=6, value=float(item.total_price))  # Individual item total
            ws.cell(row=row_num, column=7, value=subtotal if idx == 0 else "")  # Subtotal only on first row
            ws.cell(row=row_num, column=8, value=f"-{discount} ({discount_pct}%)" if idx == 0 and discount > 0 else "")  # Discount
            ws.cell(row=row_num, column=9, value=delivery if idx == 0 else "")  # Delivery
            ws.cell(row=row_num, column=10, value=receipt_total if idx == 0 else "")  # Total
            ws.cell(row=row_num, column=11, value=payment_text if idx == 0 else "")
            ws.cell(row=row_num, column=12, value=status if idx == 0 else "")

            row_num += 1

        # Add blank row between receipts
        row_num += 1

    # Add total sales
    ws.cell(row=row_num + 1, column=6, value="Total Sales:")
    ws.cell(row=row_num + 1, column=7, value=float(total_sales))
    ws.cell(row=row_num + 1, column=6).font = Font(bold=True)
    ws.cell(row=row_num + 1, column=7).font = Font(bold=True)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Create HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def discount_report(request):
    """
    Discount Report - Shows all sales with discounts applied (both payment and line-level)
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Get all payments with discounts (payment-level discounts)
    payments_with_discounts = Payment.objects.filter(
        discount_amount__gt=0
    ).prefetch_related('sale_set', 'sale_set__product', 'sale_set__receipt')

    # Get all sales with line discounts
    sales_with_line_discounts = Sale.objects.filter(
        discount_amount__gt=0
    ).select_related('product', 'receipt', 'payment').prefetch_related('payment__payment_methods')

    # Apply date filters - default to today if no dates provided
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            payments_with_discounts = payments_with_discounts.filter(payment_date__range=[start_date_obj, end_date_obj])
            sales_with_line_discounts = sales_with_line_discounts.filter(sale_date__range=[start_date_obj, end_date_obj])
        except ValueError:
            pass
    else:
        # Default to today's date if no filters applied
        today = timezone.now().date()
        start_date_obj = datetime.combine(today, datetime.min.time())
        end_date_obj = datetime.combine(today, datetime.max.time())
        payments_with_discounts = payments_with_discounts.filter(payment_date__range=[start_date_obj, end_date_obj])
        sales_with_line_discounts = sales_with_line_discounts.filter(sale_date__range=[start_date_obj, end_date_obj])
        start_date = today.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    # Calculate payment-level discount totals
    payment_discount_total = payments_with_discounts.aggregate(total=Sum('discount_amount'))['total'] or 0
    payment_transactions = payments_with_discounts.count()

    # Calculate line-level discount totals
    line_discount_total = sales_with_line_discounts.aggregate(total=Sum('discount_amount'))['total'] or 0
    line_transactions = sales_with_line_discounts.count()

    # Total discounts (both types)
    total_discount_amount = payment_discount_total + line_discount_total
    total_transactions = payment_transactions + line_transactions

    context = {
        'payments_with_discounts': payments_with_discounts,
        'sales_with_line_discounts': sales_with_line_discounts,
        'payment_discount_total': payment_discount_total,
        'line_discount_total': line_discount_total,
        'total_discount_amount': total_discount_amount,
        'payment_transactions': payment_transactions,
        'line_transactions': line_transactions,
        'total_transactions': total_transactions,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'reports/discount_report.html', context)


@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def delivery_report(request):
    """
    Delivery Report - Shows all receipts with delivery fees
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Get all receipts with delivery costs
    receipts = Receipt.objects.filter(
        delivery_cost__gt=0
    ).select_related('customer', 'user').prefetch_related('sales', 'sales__product')

    # Apply date filters - default to today if no dates provided
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            receipts = receipts.filter(date__range=[start_date_obj, end_date_obj])
        except ValueError:
            pass
    else:
        # Default to today's date if no filters applied
        today = timezone.now().date()
        start_date_obj = datetime.combine(today, datetime.min.time())
        end_date_obj = datetime.combine(today, datetime.max.time())
        receipts = receipts.filter(date__range=[start_date_obj, end_date_obj])
        start_date = today.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    # Calculate totals
    total_delivery_fees = receipts.aggregate(total=Sum('delivery_cost'))['total'] or 0
    total_deliveries = receipts.count()
    avg_delivery_fee = total_delivery_fees / total_deliveries if total_deliveries > 0 else 0

    context = {
        'receipts': receipts,
        'total_delivery_fees': total_delivery_fees,
        'total_deliveries': total_deliveries,
        'avg_delivery_fee': avg_delivery_fee,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'reports/delivery_report.html', context)


@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def financial_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    report_type = request.GET.get('report_type', 'revenue')

    sales = Sale.objects.select_related('payment', 'product', 'delivery').prefetch_related('payment__payment_methods')

    # Apply date filters - default to today if no dates provided
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            sales = sales.filter(sale_date__range=[start_date_obj, end_date_obj])
        except:
            pass
    else:
        # Default to today's date if no filters applied
        today = timezone.now().date()
        start_date_obj = datetime.combine(today, datetime.min.time())
        end_date_obj = datetime.combine(today, datetime.max.time())
        sales = sales.filter(sale_date__range=[start_date_obj, end_date_obj])
        start_date = today.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    # Get unique payments to avoid double counting
    unique_payments = Payment.objects.filter(sale__in=sales).distinct()

    # ENHANCED FINANCIAL METRICS

    # 1. Revenue Analysis
    gross_revenue = sum(sale.product.selling_price * sale.quantity for sale in sales)
    total_revenue = sum(payment.total_amount for payment in unique_payments if payment.total_amount)

    # 2. Discount Analysis
    total_item_discounts = sum(sale.discount_amount or 0 for sale in sales)
    total_payment_discounts = sum(payment.discount_amount or 0 for payment in unique_payments)
    total_discounts = total_item_discounts + total_payment_discounts

    # 3. Delivery Fee Analysis
    total_delivery_fees = sum(payment.sale_set.first().delivery.delivery_cost or 0
                              for payment in unique_payments
                              if payment.sale_set.exists() and payment.sale_set.first().delivery)

    # 4. Cost and Profit Analysis
    total_cost = sum(sale.product.price * sale.quantity for sale in sales)
    net_revenue = total_revenue - total_delivery_fees  # Revenue without delivery fees
    total_profit = net_revenue - total_cost
    profit_margin = (total_profit / net_revenue * 100) if net_revenue > 0 else 0

    # 5. DETAILED PAYMENT METHOD BREAKDOWN
    payment_method_breakdown = {}
    payment_method_stats = {}

    for payment in unique_payments:
        for pm in payment.payment_methods.filter(status='completed'):
            method_name = pm.get_payment_method_display()

            if method_name not in payment_method_breakdown:
                payment_method_breakdown[method_name] = {
                    'total_amount': 0,
                    'transaction_count': 0,
                    'avg_transaction': 0
                }

            payment_method_breakdown[method_name]['total_amount'] += float(pm.amount)
            payment_method_breakdown[method_name]['transaction_count'] += 1

    # Calculate averages
    for method, data in payment_method_breakdown.items():
        if data['transaction_count'] > 0:
            data['avg_transaction'] = data['total_amount'] / data['transaction_count']

    # Sort by total amount
    payment_method_breakdown = dict(sorted(
        payment_method_breakdown.items(),
        key=lambda x: x[1]['total_amount'],
        reverse=True
    ))

    # 6. Payment Status Analysis
    payment_status_breakdown = {
        'completed': unique_payments.filter(payment_status='completed').count(),
        'partial': unique_payments.filter(payment_status='partial').count(),
        'pending': unique_payments.filter(payment_status='pending').count(),
        'failed': unique_payments.filter(payment_status='failed').count(),
    }

    completed_amount = sum(p.total_amount for p in unique_payments.filter(payment_status='completed'))
    partial_amount = sum(p.total_paid for p in unique_payments.filter(payment_status='partial'))
    pending_amount = sum(p.total_amount for p in unique_payments.filter(payment_status='pending'))

    # 7. Monthly Revenue Trend with detailed breakdown
    monthly_data = []
    monthly_raw = (
        Payment.objects
        .filter(sale__in=sales)
        .annotate(month=TruncMonth('sale__sale_date'))
        .values('month')
        .annotate(
            revenue=Sum('total_amount'),
            discount_amount=Sum('discount_amount'),
            transaction_count=Count('id', distinct=True)
        )
        .order_by('month')
    )

    for month in monthly_raw:
        month_sales = sales.filter(sale_date__month=month['month'].month, sale_date__year=month['month'].year)
        delivery_fees = sum(s.delivery.delivery_cost or 0 for s in month_sales if s.delivery)

        monthly_data.append({
            'month': month['month'],
            'revenue': month['revenue'],
            'discount_amount': month['discount_amount'] or 0,
            'delivery_fees': delivery_fees,
            'transaction_count': month['transaction_count'],
            'net_revenue': (month['revenue'] or 0) - delivery_fees
        })

    # 8. Discount Analysis by Type
    discount_breakdown = {
        'item_level_discounts': total_item_discounts,
        'payment_level_discounts': total_payment_discounts,
        'total_discounts': total_discounts,
        'discount_rate': (total_discounts / gross_revenue * 100) if gross_revenue > 0 else 0
    }

    context = {
        # Revenue Metrics
        'gross_revenue': gross_revenue,
        'total_revenue': total_revenue,
        'net_revenue': net_revenue,
        'total_cost': total_cost,
        'total_profit': total_profit,
        'profit_margin': profit_margin,

        # Discount & Fees
        'discount_breakdown': discount_breakdown,
        'total_delivery_fees': total_delivery_fees,

        # Payment Analysis
        'payment_method_breakdown': payment_method_breakdown,
        'payment_status_breakdown': payment_status_breakdown,
        'completed_amount': completed_amount,
        'partial_amount': partial_amount,
        'pending_amount': pending_amount,

        # Trends
        'monthly_data': monthly_data,

        # Filters
        'start_date': start_date,
        'end_date': end_date,
        'report_type': report_type,

        # Summary Stats
        'total_transactions': unique_payments.count(),
        'avg_transaction_value': total_revenue / unique_payments.count() if unique_payments.count() > 0 else 0,
    }

    # Check for export format
    export_format = request.GET.get('export')
    if export_format == 'excel':
        return export_financial_to_excel(context, start_date, end_date)
    elif export_format == 'pdf':
        return export_financial_to_pdf(context, start_date, end_date)

    return render(request, 'reports/financial_report.html', context)


def export_financial_to_excel(context, start_date=None, end_date=None):
    """Export financial report data to Excel (CSV format)"""
    response = HttpResponse(content_type='text/csv')
    filename = f"financial_report_{start_date or 'all'}_to_{end_date or 'now'}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # Report Title
    writer.writerow(['FINANCIAL REPORT'])
    if start_date and end_date:
        writer.writerow([f'Date Range: {start_date} to {end_date}'])
    writer.writerow([])

    # Summary Section
    writer.writerow(['SUMMARY METRICS'])
    writer.writerow(['Metric', 'Value'])
    writer.writerow(['Gross Revenue', f"₦{context['gross_revenue']:,.2f}"])
    writer.writerow(['Net Revenue', f"₦{context['net_revenue']:,.2f}"])
    writer.writerow(['Total Profit', f"₦{context['total_profit']:,.2f}"])
    writer.writerow(['Profit Margin', f"{context['profit_margin']:.2f}%"])
    writer.writerow(['Total Transactions', context['total_transactions']])
    writer.writerow(['Avg Transaction Value', f"₦{context['avg_transaction_value']:,.2f}"])
    writer.writerow(['Total Delivery Fees', f"₦{context['total_delivery_fees']:,.2f}"])
    writer.writerow([])

    # Payment Method Breakdown
    writer.writerow(['PAYMENT METHOD BREAKDOWN'])
    writer.writerow(['Method', 'Total Amount', 'Transactions', 'Average Amount'])
    for method, data in context['payment_method_breakdown'].items():
        writer.writerow([
            method,
            f"₦{data['total_amount']:,.2f}",
            data['transaction_count'],
            f"₦{data['avg_transaction']:,.2f}"
        ])
    writer.writerow([])

    # Discount Analysis
    writer.writerow(['DISCOUNT ANALYSIS'])
    writer.writerow(['Type', 'Amount'])
    db = context['discount_breakdown']
    writer.writerow(['Item-Level Discounts', f"₦{db['item_level_discounts']:,.2f}"])
    writer.writerow(['Payment-Level Discounts', f"₦{db['payment_level_discounts']:,.2f}"])
    writer.writerow(['Total Discounts', f"₦{db['total_discounts']:,.2f}"])
    writer.writerow(['Discount Rate', f"{db['discount_rate']:.2f}%"])
    writer.writerow([])

    # Payment Status
    writer.writerow(['PAYMENT STATUS BREAKDOWN'])
    writer.writerow(['Status', 'Count', 'Amount'])
    psb = context['payment_status_breakdown']
    writer.writerow(['Completed', psb['completed'], f"₦{context['completed_amount']:,.2f}"])
    writer.writerow(['Partial', psb['partial'], f"₦{context['partial_amount']:,.2f}"])
    writer.writerow(['Pending', psb['pending'], f"₦{context['pending_amount']:,.2f}"])
    writer.writerow(['Failed', psb['failed'], "₦0.00"])
    writer.writerow([])

    # Monthly Trends
    writer.writerow(['MONTHLY FINANCIAL TRENDS'])
    writer.writerow(['Month', 'Revenue', 'Discounts', 'Delivery Fees', 'Net Revenue', 'Transactions'])
    for month in context['monthly_data']:
        writer.writerow([
            month['month'].strftime('%b %Y'),
            f"₦{month['revenue']:,.2f}",
            f"₦{month['discount_amount']:,.2f}",
            f"₦{month['delivery_fees']:,.2f}",
            f"₦{month['net_revenue']:,.2f}",
            month['transaction_count']
        ])

    return response


def export_financial_to_pdf(context, start_date=None, end_date=None):
    """Generate PDF for financial report"""
    # Create a file-like buffer to receive PDF data
    buffer = io.BytesIO()

    # Create the PDF object
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=30
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=12,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    normal_style = styles['Normal']

    # Title
    title = "Financial Report"
    if start_date and end_date:
        title += f" ({start_date} to {end_date})"
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))

    # Summary Stats
    summary_data = [
        ['Metric', 'Value'],
        ['Gross Revenue', f"₦{context['gross_revenue']:,.2f}"],
        ['Net Revenue', f"₦{context['net_revenue']:,.2f}"],
        ['Total Profit', f"₦{context['total_profit']:,.2f}"],
        ['Profit Margin', f"{context['profit_margin']:.1f}%"],
        ['Total Transactions', str(context['total_transactions'])],
        ['Avg Transaction Value', f"₦{context['avg_transaction_value']:,.2f}"],
        ['Total Delivery Fees', f"₦{context['total_delivery_fees']:,.2f}"],
    ]

    summary_table = Table(summary_data, colWidths=[200, 150])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
    ]))
    elements.append(Paragraph("Summary", subtitle_style))
    elements.append(summary_table)
    elements.append(Spacer(1, 24))

    # Payment Method Breakdown
    if context['payment_method_breakdown']:
        pm_data = [['Payment Method', 'Total Amount', 'Transactions', 'Avg Amount']]
        for method, data in context['payment_method_breakdown'].items():
            pm_data.append([
                method,
                f"₦{data['total_amount']:,.2f}",
                str(data['transaction_count']),
                f"₦{data['avg_transaction']:,.2f}"
            ])

        pm_table = Table(pm_data, colWidths=[150, 120, 100, 120])
        pm_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (1, 1), (3, -1), 'RIGHT'),
        ]))
        elements.append(Paragraph("Payment Method Breakdown", subtitle_style))
        elements.append(pm_table)
        elements.append(Spacer(1, 24))

    # Discount Analysis
    discount_data = [
        ['Discount Type', 'Amount'],
        ['Item-Level Discounts', f"₦{context['discount_breakdown']['item_level_discounts']:,.2f}"],
        ['Payment-Level Discounts', f"₦{context['discount_breakdown']['payment_level_discounts']:,.2f}"],
        ['Total Discounts', f"₦{context['discount_breakdown']['total_discounts']:,.2f}"],
        ['Discount Rate', f"{context['discount_breakdown']['discount_rate']:.1f}%"],
    ]

    discount_table = Table(discount_data, colWidths=[200, 150])
    discount_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
    ]))
    elements.append(Paragraph("Discount Analysis", subtitle_style))
    elements.append(discount_table)
    elements.append(Spacer(1, 24))

    # Payment Status
    status_data = [
        ['Status', 'Count', 'Amount'],
        ['Completed', str(context['payment_status_breakdown']['completed']), f"₦{context['completed_amount']:,.2f}"],
        ['Partial', str(context['payment_status_breakdown']['partial']), f"₦{context['partial_amount']:,.2f}"],
        ['Pending', str(context['payment_status_breakdown']['pending']), f"₦{context['pending_amount']:,.2f}"],
        ['Failed', str(context['payment_status_breakdown']['failed']), "₦0.00"],
    ]

    status_table = Table(status_data, colWidths=[100, 80, 120])
    status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
    ]))
    elements.append(Paragraph("Payment Status Overview", subtitle_style))
    elements.append(status_table)
    elements.append(Spacer(1, 24))

    # Monthly Data
    if context['monthly_data']:
        monthly_data = [['Month', 'Revenue', 'Discounts', 'Delivery Fees', 'Net Revenue', 'Transactions']]
        for month in context['monthly_data']:
            monthly_data.append([
                month['month'].strftime('%b %Y'),
                f"₦{month['revenue']:,.2f}",
                f"₦{month['discount_amount']:,.2f}",
                f"₦{month['delivery_fees']:,.2f}",
                f"₦{month['net_revenue']:,.2f}",
                str(month['transaction_count'])
            ])

        monthly_table = Table(monthly_data, colWidths=[80, 100, 100, 100, 100, 80])
        monthly_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (1, 1), (4, -1), 'RIGHT'),
        ]))
        elements.append(Paragraph("Monthly Financial Trends", subtitle_style))
        elements.append(monthly_table)

    # Build PDF
    doc.build(elements)

    # FileResponse sets the Content-Disposition header
    buffer.seek(0)
    filename = f"financial_report_{start_date or 'all'}_to_{end_date or 'now'}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response  # ✅ Added missing return statement
    return response


@login_required(login_url='login')
@user_passes_test(lambda u: u.is_superuser or (hasattr(u, 'profile') and u.profile.access_level in ['md', 'accountant', 'admin']), login_url='access_denied')
def tax_report(request):
    """Tax Report - Show tax breakdown for all receipts"""
    import json
    from datetime import datetime, timedelta
    from django.db.models import Sum, Q

    # Date filtering
    filter_type = request.GET.get('filter', 'today')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    now = timezone.now()

    # Set date range based on filter
    if filter_type == 'today':
        start_date = now.date()
        end_date = now.date()
    elif filter_type == 'yesterday':
        start_date = (now - timedelta(days=1)).date()
        end_date = (now - timedelta(days=1)).date()
    elif filter_type == 'this_week':
        start_date = (now - timedelta(days=now.weekday())).date()
        end_date = now.date()
    elif filter_type == 'this_month':
        start_date = now.replace(day=1).date()
        end_date = now.date()
    elif filter_type == 'last_month':
        last_month = now.replace(day=1) - timedelta(days=1)
        start_date = last_month.replace(day=1).date()
        end_date = last_month.date()
    elif filter_type == 'custom' and start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        start_date = now.date()
        end_date = now.date()

    # Get receipts within date range
    receipts = Receipt.objects.filter(
        date__date__gte=start_date,
        date__date__lte=end_date
    ).order_by('-date')

    # Calculate tax summary
    tax_summary = {}
    total_sales = Decimal('0')
    total_tax_collected = Decimal('0')
    inclusive_tax_total = Decimal('0')
    exclusive_tax_total = Decimal('0')

    receipt_details = []

    for receipt in receipts:
        if receipt.tax_details:
            try:
                tax_data = json.loads(receipt.tax_details) if isinstance(receipt.tax_details, str) else receipt.tax_details

                receipt_info = {
                    'receipt': receipt,
                    'total': receipt.total_with_delivery,
                    'tax_amount': receipt.tax_amount,
                    'taxes': []
                }

                for tax_code, tax_info in tax_data.items():
                    tax_amount = Decimal(str(tax_info['amount']))
                    tax_method = tax_info['method']

                    # Add to summary by tax code
                    if tax_code not in tax_summary:
                        tax_summary[tax_code] = {
                            'name': tax_info['name'],
                            'rate': tax_info['rate'],
                            'type': tax_info.get('type', 'percentage'),
                            'inclusive_amount': Decimal('0'),
                            'exclusive_amount': Decimal('0'),
                            'total_amount': Decimal('0'),
                            'count': 0
                        }

                    tax_summary[tax_code]['total_amount'] += tax_amount
                    tax_summary[tax_code]['count'] += 1

                    if tax_method == 'inclusive':
                        tax_summary[tax_code]['inclusive_amount'] += tax_amount
                        inclusive_tax_total += tax_amount
                    else:
                        tax_summary[tax_code]['exclusive_amount'] += tax_amount
                        exclusive_tax_total += tax_amount

                    receipt_info['taxes'].append({
                        'name': tax_info['name'],
                        'code': tax_code,
                        'rate': tax_info['rate'],
                        'amount': tax_amount,
                        'method': tax_method
                    })

                total_sales += receipt.total_with_delivery
                total_tax_collected += receipt.tax_amount
                receipt_details.append(receipt_info)
            except (json.JSONDecodeError, KeyError, TypeError) as e:
                logger.error(f"Error parsing tax details for receipt {receipt.receipt_number}: {e}")
                continue

    # Get active tax configurations for reference
    active_taxes = TaxConfiguration.get_active_taxes()

    context = {
        'filter_type': filter_type,
        'start_date': start_date,
        'end_date': end_date,
        'tax_summary': tax_summary,
        'receipt_details': receipt_details,
        'total_sales': total_sales,
        'total_tax_collected': total_tax_collected,
        'inclusive_tax_total': inclusive_tax_total,
        'exclusive_tax_total': exclusive_tax_total,
        'active_taxes': active_taxes,
        'receipts_count': receipts.count(),
    }

    return render(request, 'Reports/tax_report.html', context)


@user_passes_test(lambda u: u.is_superuser or u.is_md, login_url='access_denied')
@login_required(login_url='login')
def reports_dashboard(request):
    # Get date filter from GET parameters
    filter_type = request.GET.get('filter', 'today')
    now = timezone.now()
    start_date = None
    end_date = now

    # Define date ranges
    if filter_type == 'today':
        start_date = now.date()
        sales = Sale.objects.filter(sale_date__date=start_date)
    elif filter_type == 'this_week':
        start_date = now - timedelta(days=now.weekday())
        sales = Sale.objects.filter(sale_date__gte=start_date)
    elif filter_type == 'this_month':
        start_date = now.replace(day=1)
        sales = Sale.objects.filter(sale_date__gte=start_date)
    elif filter_type == 'this_year':
        start_date = now.replace(month=1, day=1)
        sales = Sale.objects.filter(sale_date__gte=start_date)
    elif filter_type == 'custom':
        start_str = request.GET.get('start_date')
        end_str = request.GET.get('end_date')
        if start_str and end_str:
            try:
                start_date = timezone.datetime.strptime(start_str, '%Y-%m-%d').date()
                end_date = timezone.datetime.strptime(end_str, '%Y-%m-%d').date()
                sales = Sale.objects.filter(sale_date__date__range=[start_date, end_date])
            except:
                start_date = now.date()
                sales = Sale.objects.filter(sale_date__date=start_date)
        else:
            start_date = now.date()
            sales = Sale.objects.filter(sale_date__date=start_date)
    else:
        start_date = now.date()
        sales = Sale.objects.filter(sale_date__date=start_date)

    # Prefetch related payment methods
    sales = sales.select_related('payment', 'product', 'customer', 'delivery').prefetch_related(
        'payment__payment_methods'
    )

    # Use Payment.total_amount and avoid double counting
    unique_payments = Payment.objects.filter(sale__in=sales).distinct()
    total_revenue = sum(payment.total_amount for payment in unique_payments if payment.total_amount)

    total_sales_count = sales.count()

    # Average Order Value (AOV) - use unique receipts to avoid inflating AOV
    unique_receipts = sales.values('receipt').distinct().count()
    avg_order_value = total_revenue / unique_receipts if unique_receipts > 0 else 0

    # Counts
    total_products = Product.objects.count()
    total_customers = Customer.objects.count()
    low_stock_products = Product.objects.filter(quantity__lt=10).count()
    low_stock_items = Product.objects.filter(quantity__lt=10).order_by('quantity')[:5]

    # Recent sales (with enhanced payment method info)
    recent_sales = []
    for sale in sales.order_by('-sale_date')[:5]:
        if sale.payment:
            methods = []
            for pm in sale.payment.payment_methods.all():
                methods.append(f"{pm.get_payment_method_display()} ({pm.amount:.2f})")
            payment_summary = ", ".join(methods) or "Unknown"
            payment_status = sale.payment.get_payment_status_display()
            total_paid = sale.payment.total_paid
            balance_due = sale.payment.balance_due
        else:
            payment_summary = "No Payment"
            payment_status = "N/A"
            total_paid = 0
            balance_due = 0

        recent_sales.append({
            'sale': sale,
            'payment_summary': payment_summary,
            'payment_status': payment_status,
            'total_paid': total_paid,
            'balance_due': balance_due,
        })

    # Monthly revenue trend - use Payment.total_amount
    monthly_data = (
        Payment.objects
        .filter(sale__sale_date__gte=now - timedelta(days=365))
        .annotate(month=TruncMonth('sale__sale_date'))
        .values('month')
        .annotate(
            revenue=Sum('total_amount'),
            discount_amount=Sum('discount_amount'),
            transaction_count=Count('id', distinct=True)
        )
        .order_by('month')[:12]
    )

    # Enhanced Category performance with payment data
    category_performance = {}
    for payment in unique_payments:
        for sale in payment.sale_set.all():
            category = sale.product.category
            if category not in category_performance:
                category_performance[category] = {
                    'product__category': category,
                    'total_revenue': 0,
                    'total_units': 0,
                    'total_discount': 0
                }

            # Proportional split of payment across items
            payment_per_sale = payment.total_amount / payment.sale_set.count() if payment.sale_set.count() > 0 else 0
            discount_per_sale = (
                                            payment.discount_amount or 0) / payment.sale_set.count() if payment.sale_set.count() > 0 else 0

            category_performance[category]['total_revenue'] += payment_per_sale
            category_performance[category]['total_units'] += sale.quantity
            category_performance[category]['total_discount'] += discount_per_sale

    category_performance = sorted(category_performance.values(), key=lambda x: x['total_revenue'], reverse=True)

    # Enhanced Top Selling Products with payment data
    product_revenue = {}
    for payment in unique_payments:
        sales_in_payment = payment.sale_set.all()
        payment_per_sale = payment.total_amount / sales_in_payment.count() if sales_in_payment.count() > 0 else 0

        for sale in sales_in_payment:
            key = (sale.product.brand, sale.product.category)
            if key not in product_revenue:
                product_revenue[key] = {'total_qty': 0, 'total_rev': 0}

            product_revenue[key]['total_qty'] += sale.quantity
            product_revenue[key]['total_rev'] += payment_per_sale

    top_products = []
    for (brand, category), data in product_revenue.items():
        top_products.append({
            'product__brand': brand,
            'product__category': category,
            'total_qty': data['total_qty'],
            'total_rev': data['total_rev']
        })

    top_products = sorted(top_products, key=lambda x: x['total_qty'], reverse=True)[:5]

    # New Customers by Month
    new_customers = (
        Customer.objects
        .filter(created_at__gte=timezone.now() - timedelta(days=365))
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    # Enhanced Payment Status Summary
    all_payments = Payment.objects.filter(sale__in=sales).distinct()
    completed_payments = all_payments.filter(payment_status='completed').count()
    partial_payments = all_payments.filter(payment_status='partial').count()
    pending_payments = all_payments.filter(payment_status='pending').count()
    failed_payments = all_payments.filter(payment_status='failed').count()

    # Enhanced Payment Method Breakdown with statistics
    payment_method_breakdown = {}
    for payment in unique_payments:
        for pm in payment.payment_methods.filter(status='completed'):
            method_name = pm.get_payment_method_display()
            if method_name not in payment_method_breakdown:
                payment_method_breakdown[method_name] = {
                    'total_amount': 0,
                    'transaction_count': 0,
                    'avg_amount': 0
                }

            payment_method_breakdown[method_name]['total_amount'] += float(pm.amount)
            payment_method_breakdown[method_name]['transaction_count'] += 1

    # Calculate averages and sort
    for method, data in payment_method_breakdown.items():
        if data['transaction_count'] > 0:
            data['avg_amount'] = data['total_amount'] / data['transaction_count']

    payment_method_breakdown = dict(sorted(
        payment_method_breakdown.items(),
        key=lambda x: x[1]['total_amount'],
        reverse=True
    ))

    # Delivery Analysis
    delivery_stats = {
        'total_delivery_fees': sum(s.delivery.delivery_cost or 0 for s in sales if s.delivery),
        'delivery_orders': sales.filter(delivery__isnull=False).count(),
        'pickup_orders': sales.filter(delivery__delivery_option='pickup').count(),
        'delivery_orders_count': sales.filter(delivery__delivery_option='delivery').count(),
    }

    # Discount Analysis - both payment-level and line-level
    # Payment-level discounts
    payment_discounts_total = sum(p.discount_amount or 0 for p in unique_payments)
    payment_discounts_count = unique_payments.filter(discount_amount__gt=0).count()

    # Line-level discounts
    line_discounts_total = sum(s.discount_amount or 0 for s in sales if s.discount_amount and s.discount_amount > 0)
    line_discounts_count = sales.filter(discount_amount__gt=0).count()

    discount_stats = {
        'total_discounts': payment_discounts_total + line_discounts_total,
        'payment_discounts': payment_discounts_total,
        'line_discounts': line_discounts_total,
        'discount_transactions': payment_discounts_count + line_discounts_count,
    }

    # Chart colors
    colors = ['#4e73df', '#1cc88a', '#36b9cc', '#f6c23e', '#e74a3b', '#858796', '#5a5c69']
    for i, cat in enumerate(category_performance):
        cat['color'] = colors[i % len(colors)]
        cat['hover_color'] = colors[(i + 2) % len(colors)]

    context = {
        'today': timezone.now().date(),
        'filter_type': filter_type,
        'start_date': start_date,
        'end_date': end_date,
        'total_revenue': total_revenue,
        'total_sales_count': total_sales_count,
        'avg_order_value': round(avg_order_value, 2),
        'total_products': total_products,
        'total_customers': total_customers,
        'low_stock_products': low_stock_products,
        'low_stock_items': low_stock_items,
        'recent_sales': recent_sales,
        'monthly_data': monthly_data,
        'category_performance': category_performance,
        'top_products': top_products,
        'new_customers': new_customers,
        'completed_payments': completed_payments,
        'discount_stats': discount_stats,
        'partial_payments': partial_payments,
        'pending_payments': pending_payments,
        'failed_payments': failed_payments,
        'payment_method_breakdown': payment_method_breakdown,
        'delivery_stats': delivery_stats,
    }

    return render(request, 'reports/dashboard.html', context)


@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def inventory_report(request):
    category = request.GET.get('category')
    low_stock = request.GET.get('low_stock')

    products = Product.objects.all()

    if category:
        products = products.filter(category=category)

    if low_stock:
        products = products.filter(quantity__lt=10)

    # Calculate enhanced inventory metrics
    total_value = sum(product.selling_price * product.quantity for product in products)
    total_cost_value = sum(product.price * product.quantity for product in products)
    potential_profit = total_value - total_cost_value

    # Additional inventory statistics
    low_stock_count = products.filter(quantity__lt=10).count()
    critical_stock_count = products.filter(quantity__lt=5).count()
    avg_markup = products.aggregate(
        avg_markup=models.Avg('markup')
    )['avg_markup'] or 0

    # Annotate products with additional fields for template
    products = products.annotate(
        total_value=models.F('selling_price') * models.F('quantity')
    )

    context = {
        'products': products,
        'categories': ProductChoices.CATEGORY_CHOICES,
        'total_value': total_value,
        'total_cost_value': total_cost_value,
        'potential_profit': potential_profit,
        'selected_category': category,
        'show_low_stock': bool(low_stock),
        'low_stock_count': low_stock_count,
        'critical_stock_count': critical_stock_count,
        'avg_markup': avg_markup,
    }

    return render(request, 'reports/inventory_report.html', context)

'''========================BARCODE PRINTS=============================='''

@csrf_exempt
@require_POST
def print_multiple_barcodes_directly(request):
    """Print multiple barcodes directly to thermal printer with individual quantities"""
    try:
        data = json.loads(request.body)
        products_data = data.get('products', [])  # [{product_id: 1, quantity: 3}, ...]

        if not products_data:
            return JsonResponse({
                'success': False,
                'error': 'No products specified for printing'
            })

        printer_name = request.session.get('selected_printer') or win32print.GetDefaultPrinter()
        if not printer_name:
            return JsonResponse({
                'success': False,
                'error': 'No printer selected. Please select a printer first.'
            })

        results = []
        total_printed = 0

        for item in products_data:
            product_id = item.get('product_id')
            quantity = item.get('quantity', 1)

            try:
                product = get_object_or_404(Product, id=product_id)

                # Ensure barcode is generated
                if not product.barcode_image or not product.barcode_number:
                    product.generate_barcode()
                    product.save()

                # Get the path to the barcode image
                barcode_path = product.barcode_image.path

                # Print the required number of copies
                copies_printed = 0
                for i in range(quantity):
                    success = print_image(printer_name, barcode_path)
                    if success:
                        copies_printed += 1
                        total_printed += 1

                    # Small delay between copies
                    time.sleep(0.3)

                results.append({
                    'product_id': product_id,
                    'product_name': product.brand,
                    'requested_quantity': quantity,
                    'printed_quantity': copies_printed,
                    'success': copies_printed == quantity
                })

            except Exception as e:
                logger.error(f"Error printing barcode for product {product_id}: {str(e)}")
                results.append({
                    'product_id': product_id,
                    'product_name': f'Product {product_id}',
                    'requested_quantity': quantity,
                    'printed_quantity': 0,
                    'success': False,
                    'error': str(e)
                })

        # Check success rate
        successful_products = sum(1 for result in results if result['success'])
        total_products = len(results)

        return JsonResponse({
            'success': successful_products > 0,
            'message': f'Printed {total_printed} barcodes for {successful_products}/{total_products} products',
            'total_printed': total_printed,
            'successful_products': successful_products,
            'total_products': total_products,
            'results': results,
            'printer_name': printer_name
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        })
    except Exception as e:
        logger.error(f"Multiple direct print failed: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Multiple direct print failed: {str(e)}'
        })


@csrf_exempt
@require_POST
def print_single_barcode_directly(request, product_id):
    """Print a single barcode with specified quantity"""
    try:
        data = json.loads(request.body)
        quantity = data.get('quantity', 1)

        product = get_object_or_404(Product, id=product_id)

        # Ensure barcode is generated
        if not product.barcode_image or not product.barcode_number:
            product.generate_barcode()
            product.save()

        printer_name = request.session.get('selected_printer') or win32print.GetDefaultPrinter()
        if not printer_name:
            return JsonResponse({
                'success': False,
                'error': 'No printer selected. Please select a printer first.'
            })

        # Get the path to the barcode image
        barcode_path = product.barcode_image.path

        # Print the required number of copies
        copies_printed = 0
        for i in range(quantity):
            success = print_image(printer_name, barcode_path)
            if success:
                copies_printed += 1

            # Small delay between copies
            time.sleep(0.3)

        return JsonResponse({
            'success': copies_printed > 0,
            'message': f'Printed {copies_printed}/{quantity} copies of barcode for {product.brand}',
            'product_name': product.brand,
            'requested_quantity': quantity,
            'printed_quantity': copies_printed,
            'printer_name': printer_name
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        })
    except Exception as e:
        logger.error(f"Single direct print failed: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Single direct print failed: {str(e)}'
        })


def print_image(printer_name, image_path):
    """Print an image directly to a printer using Windows GDI"""
    try:
        from PIL import Image, ImageWin
        import win32ui

        # Open the image
        image = Image.open(image_path)

        # Convert to monochrome if needed (better for thermal printers)
        if image.mode != "1":
            image = image.convert("1")

        # Get the printer
        hprinter = win32print.OpenPrinter(printer_name)

        try:
            # Get printer properties
            printer_info = win32print.GetPrinter(hprinter, 2)

            # Create a device context for the printer
            hdc = win32ui.CreateDC()
            hdc.CreatePrinterDC(printer_name)

            # Start document
            hdc.StartDoc("Barcode Print")
            hdc.StartPage()

            # Calculate scaling to fit the page
            printable_area = hdc.GetDeviceCaps(110), hdc.GetDeviceCaps(111)  # HORZRES, VERTRES

            # Scale image to fit printable area
            img_width, img_height = image.size
            scaling_x = printable_area[0] / img_width
            scaling_y = printable_area[1] / img_height
            scaling = min(scaling_x, scaling_y)

            # Calculate position to center image
            x = (printable_area[0] - img_width * scaling) / 2
            y = (printable_area[1] - img_height * scaling) / 2

            # Draw the image
            dib = ImageWin.Dib(image)
            dib.draw(hdc.GetHandleOutput(), (
                int(x),
                int(y),
                int(x + img_width * scaling),
                int(y + img_height * scaling)
            ))

            # End document
            hdc.EndPage()
            hdc.EndDoc()

            return True

        finally:
            win32print.ClosePrinter(hprinter)

    except Exception as e:
        logger.error(f"Error printing image: {str(e)}")
        return False


# Activity Log Views
@login_required(login_url='login')
def activity_log_list(request):
    """
    Display a paginated list of activity logs with filtering options
    """
    # Get all logs ordered by most recent
    logs = ActivityLog.objects.all().select_related('user').order_by('-created_at')

    # Get filter parameters
    action_filter = request.GET.get('action', '')
    user_filter = request.GET.get('user', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    success_filter = request.GET.get('success', '')
    search_query = request.GET.get('search', '')

    # Apply filters
    if action_filter:
        logs = logs.filter(action=action_filter)

    if user_filter:
        logs = logs.filter(username=user_filter)

    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            logs = logs.filter(created_at__gte=date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            # Add one day to include the entire end date
            date_to_obj = date_to_obj + timedelta(days=1)
            logs = logs.filter(created_at__lt=date_to_obj)
        except ValueError:
            pass

    if success_filter:
        logs = logs.filter(success=(success_filter == 'true'))

    if search_query:
        logs = logs.filter(
            Q(description__icontains=search_query) |
            Q(username__icontains=search_query) |
            Q(ip_address__icontains=search_query) |
            Q(object_repr__icontains=search_query)
        )

    # Get unique values for filters
    action_choices = ActivityLog.ACTION_CHOICES
    unique_users = ActivityLog.objects.values_list('username', flat=True).distinct().order_by('username')

    # Pagination
    paginator = Paginator(logs, 50)  # Show 50 logs per page
    page = request.GET.get('page')

    try:
        logs = paginator.page(page)
    except PageNotAnInteger:
        logs = paginator.page(1)
    except EmptyPage:
        logs = paginator.page(paginator.num_pages)

    context = {
        'logs': logs,
        'action_choices': action_choices,
        'unique_users': unique_users,
        'action_filter': action_filter,
        'user_filter': user_filter,
        'date_from': date_from,
        'date_to': date_to,
        'success_filter': success_filter,
        'search_query': search_query,
    }

    return render(request, 'activity/activity_log_list.html', context)


@login_required(login_url='login')
def activity_log_detail(request, log_id):
    """
    Display detailed information about a specific activity log entry
    """
    log = get_object_or_404(ActivityLog, id=log_id)

    context = {
        'log': log,
    }

    return render(request, 'activity/activity_log_detail.html', context)


# ==================== LOYALTY PROGRAM ENDPOINTS ====================

@login_required(login_url='login')
@require_http_methods(["GET"])
def get_customer_loyalty_info(request, customer_id):
    """
    AJAX endpoint to get customer loyalty information
    Returns JSON with customer's loyalty points, balance, and redemption eligibility
    """
    try:
        customer = get_object_or_404(Customer, id=customer_id)

        from .loyalty_utils import get_customer_loyalty_summary
        from .models import LoyaltyConfiguration

        # Get loyalty configuration
        try:
            config = LoyaltyConfiguration.get_active_config()
        except Exception:
            return JsonResponse({
                'success': False,
                'error': 'Loyalty program is not configured'
            })

        if not config.is_active:
            return JsonResponse({
                'success': False,
                'error': 'Loyalty program is not active'
            })

        # Get customer loyalty summary
        loyalty_info = get_customer_loyalty_summary(customer)

        if not loyalty_info['has_account']:
            return JsonResponse({
                'success': True,
                'has_account': False,
                'message': 'Customer does not have a loyalty account'
            })

        return JsonResponse({
            'success': True,
            'has_account': True,
            'is_active': loyalty_info['is_active'],
            'current_balance': loyalty_info['current_balance'],
            'total_earned': loyalty_info['total_earned'],
            'total_redeemed': loyalty_info['total_redeemed'],
            'redeemable_value': float(loyalty_info['redeemable_value']),
            'can_redeem': loyalty_info['can_redeem'],
            'tier': loyalty_info.get('tier', ''),
            'minimum_points_for_redemption': config.minimum_points_for_redemption,
            'points_to_currency_rate': float(config.points_to_currency_rate),
            'maximum_discount_percentage': float(config.maximum_discount_percentage)
        })

    except Customer.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Customer not found'
        })
    except Exception as e:
        logger.error(f"Error fetching loyalty info for customer {customer_id}: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required(login_url='login')
@require_http_methods(["POST"])
@csrf_exempt
def apply_loyalty_discount(request):
    """
    AJAX endpoint to calculate loyalty discount before applying it to a receipt
    This is called during POS to preview the discount
    """
    try:
        data = json.loads(request.body)
        customer_id = data.get('customer_id')
        points_to_redeem = int(data.get('points_to_redeem', 0))
        transaction_total = Decimal(str(data.get('transaction_total', 0)))

        if not customer_id or not points_to_redeem or not transaction_total:
            return JsonResponse({
                'success': False,
                'error': 'Missing required parameters'
            })

        customer = get_object_or_404(Customer, id=customer_id)

        from .models import LoyaltyConfiguration, CustomerLoyaltyAccount

        # Get loyalty configuration
        try:
            config = LoyaltyConfiguration.get_active_config()
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error loading loyalty configuration: {str(e)}'
            })

        if not config.is_active:
            return JsonResponse({
                'success': False,
                'error': 'Loyalty program is not active'
            })

        # Get loyalty account
        try:
            loyalty_account = customer.loyalty_account
        except CustomerLoyaltyAccount.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Customer does not have a loyalty account'
            })

        # Validate points redemption
        if not loyalty_account.can_redeem_points(points_to_redeem):
            return JsonResponse({
                'success': False,
                'error': f'Cannot redeem {points_to_redeem} points. '
                         f'Customer has {loyalty_account.current_balance} points. '
                         f'Minimum redemption: {config.minimum_points_for_redemption} points.'
            })

        # Calculate discount amount
        discount_amount = config.calculate_discount_from_points(points_to_redeem)

        # Check maximum discount percentage
        max_discount = config.get_maximum_redeemable_amount(transaction_total)

        if discount_amount > max_discount:
            return JsonResponse({
                'success': False,
                'error': f'Discount amount (₦{discount_amount}) exceeds maximum allowed '
                         f'(₦{max_discount}, {config.maximum_discount_percentage}% of transaction)'
            })

        if discount_amount > transaction_total:
            return JsonResponse({
                'success': False,
                'error': f'Discount amount (₦{discount_amount}) exceeds transaction total (₦{transaction_total})'
            })

        # Return discount preview
        return JsonResponse({
            'success': True,
            'points_to_redeem': points_to_redeem,
            'discount_amount': float(discount_amount),
            'remaining_balance': loyalty_account.current_balance - points_to_redeem,
            'new_total': float(transaction_total - discount_amount)
        })

    except Customer.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Customer not found'
        })
    except Exception as e:
        logger.error(f"Error applying loyalty discount: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required(login_url='login')
@require_http_methods(["POST"])
@csrf_exempt
def enroll_customer_in_loyalty(request):
    """
    AJAX endpoint to enroll a customer in the loyalty program
    """
    try:
        data = json.loads(request.body)
        customer_id = data.get('customer_id')

        if not customer_id:
            return JsonResponse({
                'success': False,
                'error': 'Customer ID is required'
            })

        customer = get_object_or_404(Customer, id=customer_id)

        # Check if customer already has a loyalty account
        from .models import CustomerLoyaltyAccount
        if hasattr(customer, 'loyalty_account'):
            return JsonResponse({
                'success': False,
                'error': 'Customer is already enrolled in the loyalty program'
            })

        # Check if loyalty program is configured and active
        from .models import LoyaltyConfiguration
        try:
            config = LoyaltyConfiguration.get_active_config()
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': 'Loyalty program is not configured'
            })

        if not config.is_active:
            return JsonResponse({
                'success': False,
                'error': 'Loyalty program is not active'
            })

        # Create loyalty account
        from .loyalty_utils import get_or_create_loyalty_account
        loyalty_account = get_or_create_loyalty_account(customer)

        logger.info(f"Customer {customer.name} (ID: {customer.id}) enrolled in loyalty program by user {request.user.username}")

        return JsonResponse({
            'success': True,
            'message': f'{customer.name} has been enrolled in the loyalty program',
            'loyalty_account_id': loyalty_account.id,
            'current_balance': loyalty_account.current_balance
        })

    except Customer.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Customer not found'
        })
    except Exception as e:
        logger.error(f"Error enrolling customer in loyalty program: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required(login_url='login')
def customer_detail(request, pk):
    """
    View customer details including loyalty information
    """
    customer = get_object_or_404(Customer, pk=pk)

    # Get loyalty information
    from .loyalty_utils import get_customer_loyalty_summary
    from .models import LoyaltyTransaction

    loyalty_info = get_customer_loyalty_summary(customer)

    # Get recent loyalty transactions if enrolled
    loyalty_transactions = []
    if loyalty_info['has_account']:
        loyalty_transactions = LoyaltyTransaction.objects.filter(
            loyalty_account=customer.loyalty_account
        ).order_by('-created_at')[:10]

    # Get recent receipts
    recent_receipts = Receipt.objects.filter(
        customer=customer
    ).order_by('-date')[:5]

    context = {
        'customer': customer,
        'loyalty_info': loyalty_info,
        'loyalty_transactions': loyalty_transactions,
        'recent_receipts': recent_receipts,
    }

    return render(request, 'customer/customer_detail.html', context)


# ============================================================================
# RETURN MANAGEMENT VIEWS
# ============================================================================

@login_required
def return_list(request):
    """List all returns with filtering"""
    from .models import Return

    returns = Return.objects.all().select_related('customer', 'receipt', 'processed_by')

    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter:
        returns = returns.filter(status=status_filter)

    # Filter by customer if provided
    customer_id = request.GET.get('customer')
    if customer_id:
        returns = returns.filter(customer_id=customer_id)

    context = {
        'returns': returns,
        'status_choices': Return.STATUS_CHOICES,
    }
    return render(request, 'returns/return_list.html', context)


@login_required
def return_detail(request, return_id):
    """View details of a specific return"""
    from .models import Return, StoreCredit
    import logging

    logger = logging.getLogger(__name__)
    logger.info(f"=== RETURN DETAIL VIEW ===")
    logger.info(f"Return ID: {return_id}")

    return_obj = get_object_or_404(
        Return.objects.select_related('customer', 'receipt', 'processed_by', 'approved_by'),
        id=return_id
    )

    logger.info(f"Return Number: {return_obj.return_number}")
    logger.info(f"Status: {return_obj.status}")
    logger.info(f"Refund Amount: {return_obj.refund_amount}")
    logger.info(f"Customer: {return_obj.customer}")

    return_items = return_obj.return_items.all().select_related('product', 'original_sale')
    logger.info(f"Return Items Count: {return_items.count()}")

    # Check for associated store credit
    store_credit = None
    if return_obj.customer:
        store_credit = StoreCredit.objects.filter(return_transaction=return_obj).first()
        logger.info(f"Store Credit: {store_credit.credit_number if store_credit else 'None'}")

    context = {
        'return_obj': return_obj,
        'return': return_obj,  # Keep both for compatibility
        'return_items': return_items,
        'store_credit': store_credit,
    }
    logger.info("=== END RETURN DETAIL VIEW ===")
    return render(request, 'returns/return_detail.html', context)


@login_required
def return_search(request):
    """Search for receipts to create returns"""
    from .models import Receipt

    receipts = None
    query = request.GET.get('q', '').strip()

    if query:
        receipts = Receipt.objects.filter(
            models.Q(receipt_number__icontains=query) |
            models.Q(customer__name__icontains=query) |
            models.Q(customer__phone__icontains=query)
        ).select_related('customer')[:20]

    context = {
        'receipts': receipts,
        'query': query,
    }
    return render(request, 'returns/return_search.html', context)


@login_required
def return_select_items(request, receipt_id):
    """Select items from a receipt to return"""
    from .models import Receipt, Sale, ReturnItem
    from django.db.models import Sum
    from datetime import timedelta

    receipt = get_object_or_404(Receipt.objects.select_related('customer'), id=receipt_id)

    # Check if receipt is within return period (7 days)
    receipt_age = timezone.now() - receipt.date
    days_since_purchase = receipt_age.days
    days_remaining = max(0, 7 - days_since_purchase)

    if days_remaining == 0:
        messages.error(request, "This receipt is beyond the 7-day return period and cannot be returned.")
        return redirect('receipt_detail', pk=receipt_id)

    sales = receipt.sales.all().select_related('product')

    # Calculate already returned quantities for each sale
    for sale in sales:
        # Get total quantity already returned for this sale
        returned_qty = ReturnItem.objects.filter(
            original_sale=sale
        ).aggregate(total=Sum('quantity_returned'))['total'] or 0

        sale.already_returned = returned_qty
        sale.max_returnable = sale.quantity - returned_qty
        sale.has_returnable = sale.max_returnable > 0

    if request.method == 'POST':
        # Process the return creation
        from .models import Return, ReturnItem
        from decimal import Decimal

        selected_items = []
        # Get list of selected item IDs from checkboxes
        selected_sale_ids = request.POST.getlist('selected_items')

        for sale in sales:
            # Check if this sale was selected via checkbox
            if str(sale.id) in selected_sale_ids:
                qty_str = request.POST.get(f'return_quantity_{sale.id}', '0').strip()
                try:
                    qty = int(qty_str) if qty_str else 0
                except ValueError:
                    qty = 0

                if qty > 0:
                    # Verify quantity doesn't exceed max returnable
                    if qty > sale.max_returnable:
                        messages.error(request, f"Cannot return {qty} of {sale.product.product_name} - only {sale.max_returnable} available")
                        return redirect('return_select_items', receipt_id=receipt_id)

                    selected_items.append({
                        'sale': sale,
                        'quantity': qty,
                        'new_price': request.POST.get(f'new_price_{sale.id}'),
                        'condition': request.POST.get(f'item_condition_{sale.id}', 'new'),
                        'restock': f'restock_{sale.id}' in request.POST,
                        'notes': request.POST.get(f'item_notes_{sale.id}', ''),
                    })

        if not selected_items:
            messages.error(request, "Please select at least one item to return")
            return redirect('return_select_items', receipt_id=receipt_id)

        # Create the return
        return_obj = Return.objects.create(
            receipt=receipt,
            customer=receipt.customer,
            processed_by=request.user,
            return_reason=request.POST.get('return_reason', 'other'),
            reason_notes=request.POST.get('reason_notes', ''),
        )

        # Create return items
        subtotal = Decimal('0.00')
        for item_data in selected_items:
            sale = item_data['sale']
            qty = item_data['quantity']

            # Calculate refund amount (proportional to quantity)
            refund_amount = (sale.total_price / sale.quantity) * qty

            # Use new price if provided (handle empty strings)
            new_price = item_data.get('new_price', '').strip()
            if new_price:
                try:
                    new_price = Decimal(new_price)
                except (ValueError, Exception):
                    new_price = None
            else:
                new_price = None

            ReturnItem.objects.create(
                return_transaction=return_obj,
                original_sale=sale,
                product=sale.product,
                quantity_sold=sale.quantity,
                quantity_returned=qty,
                original_selling_price=sale.product.selling_price,
                new_selling_price=new_price,
                original_total=sale.total_price,
                refund_amount=refund_amount,
                item_condition=item_data['condition'],
                restock_to_inventory=item_data['restock'],
                notes=item_data.get('notes', '').strip(),
            )

            subtotal += refund_amount

        # Update return totals
        return_obj.subtotal = subtotal
        return_obj.refund_amount = subtotal  # Can be adjusted later
        return_obj.save()

        messages.success(request, f"Return {return_obj.return_number} created successfully")
        return redirect('return_detail', return_id=return_obj.id)

    context = {
        'receipt': receipt,
        'sales': sales,
        'days_remaining': days_remaining,
        'days_since_purchase': days_since_purchase,
    }
    return render(request, 'returns/return_select_items.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def return_approve(request, return_id):
    """Approve a return"""
    from .models import Return

    return_obj = get_object_or_404(Return, id=return_id)

    if request.method == 'POST':
        return_obj.status = 'approved'
        return_obj.approved_by = request.user
        return_obj.approved_date = timezone.now()
        return_obj.save()
        messages.success(request, f"Return {return_obj.return_number} approved successfully")
        return redirect('return_detail', return_id=return_obj.id)

    return redirect('return_detail', return_id=return_obj.id)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def return_reject(request, return_id):
    """Reject a return"""
    from .models import Return

    return_obj = get_object_or_404(Return, id=return_id)

    if request.method == 'POST':
        return_obj.status = 'rejected'
        return_obj.reason_notes = request.POST.get('rejection_reason', return_obj.reason_notes)
        return_obj.save()
        messages.warning(request, f"Return {return_obj.return_number} rejected")
        return redirect('return_detail', return_id=return_obj.id)

    return redirect('return_detail', return_id=return_obj.id)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def return_cancel(request, return_id):
    """Cancel a return"""
    from .models import Return

    return_obj = get_object_or_404(Return, id=return_id)

    if request.method == 'POST':
        return_obj.status = 'cancelled'
        return_obj.save()
        messages.info(request, f"Return {return_obj.return_number} has been cancelled")
        return redirect('return_detail', return_id=return_obj.id)

    return redirect('return_detail', return_id=return_obj.id)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def return_complete_form(request, return_id):
    """Complete/approve a return and process refund"""
    from .models import Return, StoreCredit
    from decimal import Decimal
    import logging

    logger = logging.getLogger(__name__)
    return_obj = get_object_or_404(Return, id=return_id)

    if request.method == 'POST':
        logger.info(f"=== RETURN COMPLETE DEBUG ===")
        logger.info(f"Return ID: {return_id}")
        logger.info(f"Return Number: {return_obj.return_number}")
        logger.info(f"POST Data: {dict(request.POST)}")

        action = request.POST.get('action')
        logger.info(f"Action parameter: {action}")

        if action == 'approve':
            logger.info("Processing APPROVE action")
            return_obj.status = 'approved'
            return_obj.approved_by = request.user
            return_obj.approved_date = timezone.now()
            return_obj.save()
            messages.success(request, "Return approved successfully")

        elif action == 'complete':
            logger.info("Processing COMPLETE action")
            # Process the refund
            refund_type = request.POST.get('refund_type')
            logger.info(f"Refund type: {refund_type}")
            logger.info(f"Refund amount: {return_obj.refund_amount}")

            return_obj.refund_type = refund_type
            return_obj.status = 'completed'
            return_obj.refunded_date = timezone.now()

            if refund_type == 'store_credit':
                logger.info(f"Creating store credit for customer: {return_obj.customer}")
                logger.info(f"Customer ID: {return_obj.customer.id if return_obj.customer else 'None'}")

                if not return_obj.customer:
                    logger.error("ERROR: Cannot create store credit - no customer associated with return")
                    messages.error(request, "Cannot create store credit: No customer associated with this return")
                    return redirect('return_detail', return_id=return_obj.id)

                # Check if store credit already exists for this return
                existing_credit = StoreCredit.objects.filter(return_transaction=return_obj).first()
                if existing_credit:
                    logger.warning(f"Store credit already exists: {existing_credit.credit_number}")
                    messages.warning(request, f"Store credit {existing_credit.credit_number} already exists for this return")
                else:
                    try:
                        # Create store credit
                        store_credit = StoreCredit.objects.create(
                            customer=return_obj.customer,
                            original_amount=return_obj.refund_amount,
                            remaining_balance=return_obj.refund_amount,
                            return_transaction=return_obj,
                            issued_by=request.user,
                            notes=f"Store credit from return {return_obj.return_number}",
                        )
                        logger.info(f"✓ Store credit created successfully!")
                        logger.info(f"  - Credit Number: {store_credit.credit_number}")
                        logger.info(f"  - Amount: ₦{store_credit.original_amount}")
                        logger.info(f"  - Customer: {store_credit.customer.name}")
                        logger.info(f"  - Issued By: {request.user.username}")
                        messages.success(request, f"Store credit {store_credit.credit_number} created for ₦{return_obj.refund_amount}")
                    except Exception as e:
                        logger.error(f"ERROR creating store credit: {str(e)}")
                        logger.exception("Full traceback:")
                        messages.error(request, f"Failed to create store credit: {str(e)}")
                        return redirect('return_detail', return_id=return_obj.id)
            else:
                # Cash refund
                return_obj.refund_method = request.POST.get('refund_method', 'Cash')
                logger.info(f"Cash refund processed via {return_obj.refund_method}")
                messages.success(request, f"Cash refund of ₦{return_obj.refund_amount} processed")

            # Restock items if needed
            logger.info("Processing inventory restocking...")
            restocked_count = 0
            for return_item in return_obj.return_items.all():
                logger.info(f"Item: {return_item.product.brand} - Qty: {return_item.quantity_returned}, Restock: {return_item.restock_to_inventory}, Already Restocked: {return_item.restocked}")

                if return_item.restock_to_inventory and not return_item.restocked:
                    product = return_item.product
                    old_quantity = product.quantity
                    product.quantity += return_item.quantity_returned
                    product.save()
                    logger.info(f"Restocked {return_item.product.brand}: {old_quantity} -> {product.quantity}")

                    return_item.restocked = True
                    return_item.restocked_date = timezone.now()
                    return_item.save()
                    restocked_count += 1

            logger.info(f"Total items restocked: {restocked_count}")
            return_obj.save()

            # Verify store credit creation
            if refund_type == 'store_credit' and return_obj.customer:
                verified_credit = StoreCredit.objects.filter(return_transaction=return_obj).first()
                if verified_credit:
                    logger.info(f"✓ VERIFIED: Store credit {verified_credit.credit_number} exists in database")
                    logger.info(f"  - Balance: ₦{verified_credit.remaining_balance}")
                    logger.info(f"  - Active: {verified_credit.is_active}")
                else:
                    logger.error("✗ ERROR: Store credit was NOT saved to database!")
                    messages.error(request, "Warning: Store credit may not have been created properly")

            messages.success(request, f"Return completed successfully. {restocked_count} item(s) restocked.")

        elif action == 'reject':
            logger.info("Processing REJECT action")
            return_obj.status = 'rejected'
            return_obj.reason_notes = request.POST.get('rejection_reason', '')
            return_obj.save()
            messages.warning(request, "Return rejected")

        else:
            logger.warning(f"Unknown or missing action parameter: '{action}'")
            messages.error(request, f"Invalid action. Please try again.")

        logger.info("=== END RETURN COMPLETE DEBUG ===")
        return redirect('return_detail', return_id=return_obj.id)

    context = {
        'return_obj': return_obj,
        'return': return_obj,  # Keep both for compatibility
        'return_items': return_obj.return_items.all(),
    }
    return render(request, 'returns/return_complete_form.html', context)


# ============================================================================
# STORE CREDIT VIEWS
# ============================================================================

@login_required
def store_credit_list(request):
    """List all store credits"""
    from .models import StoreCredit
    from django.db.models import Sum

    store_credits = StoreCredit.objects.all().select_related('customer', 'issued_by')

    # Filter by active status
    is_active = request.GET.get('active')
    if is_active == '1':
        store_credits = store_credits.filter(is_active=True, remaining_balance__gt=0)
    elif is_active == '0':
        store_credits = store_credits.filter(is_active=False)

    # Filter by customer
    customer_id = request.GET.get('customer')
    if customer_id:
        store_credits = store_credits.filter(customer_id=customer_id)

    # Calculate totals
    total_credits = store_credits.count()
    total_balance = store_credits.aggregate(Sum('remaining_balance'))['remaining_balance__sum'] or 0

    context = {
        'credits': store_credits,  # Changed from 'store_credits' to 'credits' to match template
        'total_credits': total_credits,
        'total_balance': total_balance,
    }
    return render(request, 'store_credits/store_credit_list.html', context)


@login_required
def store_credit_detail(request, credit_id):
    """View details of a specific store credit"""
    from .models import StoreCredit

    store_credit = get_object_or_404(
        StoreCredit.objects.select_related('customer', 'issued_by', 'return_transaction'),
        id=credit_id
    )

    usages = store_credit.usages.all().select_related('receipt', 'used_by')

    context = {
        'store_credit': store_credit,
        'usages': usages,
    }
    return render(request, 'store_credits/store_credit_detail.html', context)


@login_required
def get_customer_store_credit(request, customer_id):
    """API endpoint to get customer's store credit information"""
    from .models import StoreCredit, Customer
    from django.http import JsonResponse
    from decimal import Decimal

    try:
        customer = Customer.objects.get(id=customer_id)

        # Get all active store credits for this customer
        active_credits = StoreCredit.objects.filter(
            customer=customer,
            is_active=True,
            remaining_balance__gt=0
        )

        # Calculate total available balance
        total_balance = sum([credit.remaining_balance for credit in active_credits])

        # Get credit details
        credits_list = []
        for credit in active_credits:
            credits_list.append({
                'credit_number': credit.credit_number,
                'remaining_balance': float(credit.remaining_balance),
                'original_amount': float(credit.original_amount),
                'issued_date': credit.issued_date.strftime('%Y-%m-%d'),
            })

        return JsonResponse({
            'success': True,
            'customer_id': customer.id,
            'customer_name': customer.name,
            'total_balance': float(total_balance),
            'credits_count': active_credits.count(),
            'credits': credits_list
        })

    except Customer.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Customer not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ============================================================================
# PARTIAL PAYMENT VIEWS
# ============================================================================

@login_required
def add_partial_payment(request, receipt_id):
    """Add a partial payment to a receipt"""
    from .models import Receipt, PartialPayment
    from decimal import Decimal

    receipt = get_object_or_404(Receipt, id=receipt_id)

    if request.method == 'POST':
        amount = Decimal(request.POST.get('amount', '0'))
        payment_method = request.POST.get('payment_method', 'Cash')
        notes = request.POST.get('notes', '')

        if amount <= 0:
            messages.error(request, "Payment amount must be greater than 0")
            return redirect('receipt_detail', pk=receipt_id)

        if amount > receipt.balance_remaining:
            messages.error(request, "Payment amount cannot exceed remaining balance")
            return redirect('receipt_detail', pk=receipt_id)

        # Create the partial payment
        PartialPayment.objects.create(
            receipt=receipt,
            amount=amount,
            payment_method=payment_method,
            notes=notes,
            received_by=request.user,
        )

        # Update receipt balances
        receipt.amount_paid += amount
        receipt.balance_remaining -= amount

        if receipt.balance_remaining <= 0:
            receipt.payment_status = 'paid'
        else:
            receipt.payment_status = 'partial'

        receipt.save()

        messages.success(request, f"Payment of {amount} recorded successfully")
        return redirect('receipt_detail', pk=receipt_id)

    return redirect('receipt_detail', pk=receipt_id)


@login_required
def customer_debt_dashboard(request):
    """View all customers with outstanding balances"""
    from .models import Receipt
    from django.db.models import Sum, Q

    # Get all receipts with outstanding balances
    outstanding_receipts = Receipt.objects.filter(
        payment_status__in=['partial', 'pending'],
        balance_remaining__gt=0
    ).select_related('customer').order_by('-date')

    # Calculate totals
    total_outstanding = outstanding_receipts.aggregate(
        total=Sum('balance_remaining')
    )['total'] or 0

    context = {
        'outstanding_receipts': outstanding_receipts,
        'total_outstanding': total_outstanding,
    }
    return render(request, 'sales/customer_debt_dashboard.html', context)


@login_required
def gift_report(request):
    """Report on all items given as gifts"""
    from .models import Sale
    from django.db.models import Sum, Count

    # Get date filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Get all gift sales
    gift_sales = Sale.objects.filter(is_gift=True).select_related(
        'product', 'customer', 'receipt'
    ).order_by('-sale_date')

    if start_date:
        gift_sales = gift_sales.filter(sale_date__gte=start_date)
    if end_date:
        gift_sales = gift_sales.filter(sale_date__lte=end_date)

    # Calculate statistics
    stats = gift_sales.aggregate(
        total_items=Count('id'),
        total_quantity=Sum('quantity'),
        total_value=Sum('original_value'),
    )

    context = {
        'gift_sales': gift_sales,
        'stats': stats,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'reports/gift_report.html', context)


