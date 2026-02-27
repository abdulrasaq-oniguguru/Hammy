# Standard library
import csv
import json
import logging
from decimal import Decimal
from io import BytesIO

# Third-party libraries
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
)

# Django imports
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db import models, transaction
from django.db.models import Q, F, Sum, DecimalField, ExpressionWrapper
from django.forms import formset_factory
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

# Local app imports
from ..choices import ProductChoices
from ..forms import (
    ProductForm, ExcelUploadForm
)
from ..models import (
    Product, Invoice, InvoiceProduct, ProductHistory, ProductDraft,
    ActivityLog, TransferItem
)
from ..utils import get_cached_choices, get_product_stats
from .auth import is_md, is_cashier, is_superuser, user_required_access
from .invoices import _check_duplicate_invoice

logger = logging.getLogger(__name__)


@login_required(login_url='login')
@user_passes_test(is_superuser, login_url='login')
@require_POST
@csrf_exempt
def update_product_quantity(request):
    """Update product quantity via AJAX for superusers only"""
    try:
        data = json.loads(request.body)
        product_id = data.get('product_id')
        action = data.get('action')  # 'increase' or 'decrease'

        if not product_id or not action:
            return JsonResponse({'success': False, 'error': 'Missing product_id or action'})

        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Product not found'})

        # Update quantity based on action
        if action == 'increase':
            product.quantity += 1
        elif action == 'decrease':
            if product.quantity > 0:
                product.quantity -= 1
            else:
                return JsonResponse({'success': False, 'error': 'Quantity cannot be negative'})
        else:
            return JsonResponse({'success': False, 'error': 'Invalid action'})

        # Save the product
        product.save()

        # Return updated data
        return JsonResponse({
            'success': True,
            'new_quantity': product.quantity,
            'product_brand': product.brand
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def filter_products(request, queryset):
    """Apply filters to product queryset based on request parameters"""
    query = request.GET.get('search', '')
    category = request.GET.get('category', '')
    shop = request.GET.get('shop', '')
    size = request.GET.get('size', '')
    color = request.GET.get('color', '')
    design = request.GET.get('design', '')
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')
    min_quantity = request.GET.get('min_quantity', '')
    max_quantity = request.GET.get('max_quantity', '')

    if query:
        queryset = queryset.filter(
            Q(brand__icontains=query) |
            Q(color__icontains=query) |
            Q(category__icontains=query) |
            Q(design__icontains=query) |
            Q(size__icontains=query)
        )
    if category:
        queryset = queryset.filter(category=category)
    if shop:
        queryset = queryset.filter(shop=shop)
    if size:
        queryset = queryset.filter(size__icontains=size)
    if color:
        queryset = queryset.filter(color=color)
    if design:
        queryset = queryset.filter(design=design)
    if min_price:
        queryset = queryset.filter(price__gte=min_price)
    if max_price:
        queryset = queryset.filter(price__lte=max_price)
    if min_quantity:
        queryset = queryset.filter(quantity__gte=min_quantity)
    if max_quantity:
        queryset = queryset.filter(quantity__lte=max_quantity)

    return queryset


@login_required(login_url='login')
def product_list(request):
    # Get all filter parameters
    query = request.GET.get('search', '')
    barcode = request.GET.get('barcode', '')
    category = request.GET.get('category', '')
    shop = request.GET.get('shop', '')
    size = request.GET.get('size', '')
    color = request.GET.get('color', '')
    design = request.GET.get('design', '')
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')
    min_quantity = request.GET.get('min_quantity', '')
    max_quantity = request.GET.get('max_quantity', '')
    sort_by_name = request.GET.get('sort_by_name', '')
    sort_by_price = request.GET.get('sort_by_price', '')
    sort_by_quantity = request.GET.get('sort_by_quantity', '')

    # Build a single Q object for all filters
    filters = Q()

    # Always exclude products with zero or negative quantities
    filters &= Q(quantity__gt=0)

    # Default to showing only STORE (shop floor) items unless user explicitly selects warehouse
    if shop:
        filters &= Q(shop=shop)
    else:
        filters &= Q(shop='STORE')  # Default to shop floor items

    if query:
        filters &= (
                Q(brand__icontains=query) |
                Q(color__icontains=query) |
                Q(category__icontains=query) |
                Q(design__icontains=query) |
                Q(size__icontains=query)
        )

    if barcode:
        filters &= Q(barcode_number__icontains=barcode)

    if category:
        filters &= Q(category=category)
    if size:
        filters &= Q(size__icontains=size)
    if color:
        filters &= Q(color=color)
    if design:
        filters &= Q(design=design)
    if min_price:
        filters &= Q(price__gte=min_price)
    if max_price:
        filters &= Q(price__lte=max_price)
    if min_quantity:
        filters &= Q(quantity__gte=min_quantity)
    if max_quantity:
        filters &= Q(quantity__lte=max_quantity)

    # Apply all filters at once
    products = Product.objects.filter(filters)

    # Apply sorting
    if sort_by_name:
        products = products.order_by(sort_by_name)
    elif sort_by_price:
        products = products.order_by(sort_by_price)
    elif sort_by_quantity:
        products = products.order_by(sort_by_quantity)
    else:
        products = products.order_by('brand')  # Default sorting

    # GET CACHED STATS (FAST!)
    stats = get_product_stats()
    total_items = stats['total_items']
    total_quantity = stats['total_quantity']
    store_quantity = stats['store_quantity']
    warehouse_quantity = stats['warehouse_quantity']
    total_inventory_value = stats['total_inventory_value']
    store_inventory_value = stats['store_inventory_value']
    warehouse_inventory_value = stats['warehouse_inventory_value']

    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(products, 25)
    try:
        products = paginator.page(page)
    except PageNotAnInteger:
        products = paginator.page(1)
    except EmptyPage:
        products = paginator.page(paginator.num_pages)

    # Check if any filters are applied
    has_filters = any([query, barcode, category, shop, size, color, design, min_price, max_price, min_quantity, max_quantity])

    # GET CACHED CHOICES (NO MORE FLATTENING IN VIEW!)
    color_choices = get_cached_choices('color')
    design_choices = get_cached_choices('design')
    category_choices = get_cached_choices('category')

    # Process shop choices - ensure they're in tuple format
    shop_choices = []
    for shop in ProductChoices.SHOP_TYPE:
        if isinstance(shop, (tuple, list)) and len(shop) >= 2:
            shop_choices.append((shop[0], shop[1]))
        else:
            shop_choices.append((shop, shop))

    return render(request, 'product/product_list.html', {
        'products': products,
        'query': query,
        'category_choices': category_choices,
        'total_items': total_items,
        'total_quantity': total_quantity,
        'store_quantity': store_quantity,
        'warehouse_quantity': warehouse_quantity,
        'total_inventory_value': total_inventory_value,
        'store_inventory_value': store_inventory_value,
        'warehouse_inventory_value': warehouse_inventory_value,
        'shop_choices': shop_choices,
        'has_filters': has_filters,
        'COLOR_CHOICES': color_choices,
        'DESIGN_CHOICES': design_choices,
        'is_superuser': request.user.is_superuser,
        'current_filters': {
            'barcode': barcode,
            'category': category,
            'shop': shop,
            'size': size,
            'color': color,
            'design': design,
            'min_price': min_price,
            'max_price': max_price,
            'min_quantity': min_quantity,
            'max_quantity': max_quantity,
            'sort_by_name': sort_by_name,
            'sort_by_price': sort_by_price,
            'sort_by_quantity': sort_by_quantity,
        }
    })


@login_required(login_url='login')
def add_product(request):
    ProductFormSet = formset_factory(ProductForm, extra=1)

    if request.method == 'POST':
        formset = ProductFormSet(request.POST, request.FILES)
        draft_id = request.POST.get('current_draft_id')
        if formset.is_valid():
            has_data = any(form.cleaned_data for form in formset)
            if not has_data:
                messages.error(request, "Please add at least one product.")
            else:
                invoice = Invoice.objects.create(user=request.user)
                created_product_ids = []

                for idx, form in enumerate(formset):
                    if form.cleaned_data:
                        product = form.save(commit=False)
                        product.invoice = invoice

                        # Explicitly assign these so calculate_selling_price is accurate
                        product.markup = form.cleaned_data.get('markup', 0)
                        product.markup_type = form.cleaned_data.get('markup_type', 'percentage')

                        # Handle design field (optional)
                        product.design = form.cleaned_data.get('design', 'plain')

                        # Handle image field (optional - will be None if not provided)
                        if 'image' in form.cleaned_data and form.cleaned_data['image']:
                            product.image = form.cleaned_data['image']

                        product.selling_price = product.calculate_selling_price()

                        product.save()
                        created_product_ids.append(product.id)

                        # Log product creation
                        ActivityLog.log_activity(
                            user=request.user,
                            action='product_create',
                            description=f'Created product: {product.brand} ({product.category}) - Qty: {product.quantity}',
                            model_name='Product',
                            object_id=product.id,
                            object_repr=str(product),
                            request=request
                        )

                        InvoiceProduct.objects.create(
                            invoice=invoice,
                            product_name=product.brand,
                            product_price=product.price,
                            product_color=product.color,
                            product_size=product.size,
                            product_category=product.category,
                            quantity=product.quantity,
                            total_price=product.price * product.quantity
                        )

                # Delete draft if one was active
                if draft_id:
                    ProductDraft.objects.filter(id=draft_id, user=request.user).delete()

                # Warn if a matching invoice was created in the last 48 hours
                duplicate = _check_duplicate_invoice(invoice)
                if duplicate:
                    messages.warning(
                        request,
                        f"Invoice {duplicate.invoice_number} (created within the last 48 hours) "
                        f"contains the same items as this one. Please verify this is not a duplicate."
                    )

                # Store created product IDs in session and redirect to success page
                request.session['new_product_ids'] = created_product_ids
                return redirect('add_product_success')
        else:
            messages.error(request, "There were errors in the form. Please correct them.")
    else:
        formset = ProductFormSet()

    # Load user drafts for the draft resume modal
    user_drafts = ProductDraft.objects.filter(user=request.user)
    context = {
        'formset': formset,
        'user_drafts': user_drafts,
        # Only auto-open the draft modal on the initial GET visit.
        # On a POST re-render (validation failure) keep it closed so the
        # user can see the form errors without the modal appearing.
        'show_draft_modal': request.method == 'GET',
    }
    return render(request, 'product/add_product.html', context)


# =====================================
# PRODUCT DRAFT VIEWS
# =====================================

@csrf_exempt
@login_required(login_url='login')
def save_product_draft(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'}, status=405)
    try:
        data = json.loads(request.body)
        draft_id = data.get('draft_id')
        name = data.get('name', 'Draft')
        form_data = data.get('form_data', {})

        if draft_id:
            draft = ProductDraft.objects.filter(id=draft_id, user=request.user).first()
            if draft:
                draft.name = name
                draft.form_data = form_data
                draft.save()
            else:
                draft = ProductDraft.objects.create(user=request.user, name=name, form_data=form_data)
        else:
            draft = ProductDraft.objects.create(user=request.user, name=name, form_data=form_data)

        return JsonResponse({'success': True, 'draft_id': draft.id, 'message': 'Draft saved successfully.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required(login_url='login')
def load_product_draft(request, draft_id):
    draft = get_object_or_404(ProductDraft, id=draft_id, user=request.user)
    return JsonResponse({'success': True, 'form_data': draft.form_data, 'name': draft.name})


@login_required(login_url='login')
def delete_product_draft(request, draft_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'}, status=405)
    ProductDraft.objects.filter(id=draft_id, user=request.user).delete()
    return JsonResponse({'success': True})


@login_required(login_url='login')
def list_product_drafts(request):
    drafts = ProductDraft.objects.filter(user=request.user).values('id', 'name', 'updated_at')
    return JsonResponse({'drafts': [
        {'id': d['id'], 'name': d['name'], 'updated_at': d['updated_at'].strftime('%Y-%m-%d %H:%M')}
        for d in drafts
    ]})


# =====================================
# ADD PRODUCT SUCCESS VIEW
# =====================================

@login_required(login_url='login')
def add_product_success(request):
    product_ids = request.session.pop('new_product_ids', [])
    if not product_ids:
        return redirect('invoice_list')
    products = Product.objects.filter(id__in=product_ids)
    return render(request, 'product/add_product_success.html', {
        'products': products,
        'product_ids_json': json.dumps([
            {'product_id': p.id, 'quantity': p.quantity} for p in products
        ]),
    })


# =====================================
# TEMPORARY PRODUCT VIEWS
# =====================================

@login_required(login_url='login')
def add_temporary_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            product.status = 'temporary'
            product.save()
            messages.success(request, "Temporary product added successfully.")
            return redirect('finalize_products')
    else:
        form = ProductForm()
    return render(request, 'product/add_temporary_product.html', {'form': form})


@login_required(login_url='login')
def finalize_products(request):
    if request.method == 'POST':
        selected_product_ids = request.POST.getlist('products')
        for product_id in selected_product_ids:
            product = Product.objects.get(id=product_id)
            if product.markup and product.markup_type:
                product.status = 'finalized'  # Finalize the product
                product.save()
            else:
                messages.warning(request, f"Product {product.brand} requires markup.")
        messages.success(request, "Selected products have been finalized.")
        return redirect('finalize_products')
    else:
        products = Product.objects.filter(status='temporary')
    return render(request, 'product/finalize_products.html', {'products': products})


@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        reason = request.POST.get('reason', '').strip()

        if not reason:
            messages.error(request, "A reason for editing the product is required.")
        elif form.is_valid():
            # Get the original quantity before editing
            old_quantity = product.quantity

            # Save the updated product (triggers save() → barcode regeneration)
            form.save()

            # Calculate the quantity change
            quantity_changed = product.quantity - old_quantity

            # Log the edit action with quantity change
            ProductHistory.objects.create(
                product=product,
                user=request.user,
                action='EDIT',
                reason=reason,
                quantity_changed=quantity_changed
            )

            # Log to activity log
            ActivityLog.log_activity(
                user=request.user,
                action='product_update',
                description=f'Updated product: {product.brand} - Reason: {reason} - Qty change: {quantity_changed:+d}',
                model_name='Product',
                object_id=product.id,
                object_repr=str(product),
                extra_data={'reason': reason, 'quantity_changed': quantity_changed},
                request=request
            )

            messages.success(request, "Product updated successfully.")
            return redirect('product_list')
        else:
            messages.error(request, "Error updating the product. Please correct the form.")
    else:
        form = ProductForm(instance=product)

    return render(request, 'product/edit_product.html', {'form': form, 'product': product})


@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        reason = request.POST.get('reason', '').strip()

        if not reason:
            messages.error(request, "A reason for deleting the product is required.")
        else:
            # Log the delete action with None as quantity_changed
            ProductHistory.objects.create(
                product=product,
                user=request.user,
                action='DELETE',
                reason=reason,
                quantity_changed=None  # No quantity change on deletion
            )

            # Log to activity log before deleting
            product_info = str(product)
            product_id = product.id
            ActivityLog.log_activity(
                user=request.user,
                action='product_delete',
                description=f'Deleted product: {product_info} - Reason: {reason}',
                model_name='Product',
                object_id=product_id,
                object_repr=product_info,
                extra_data={'reason': reason},
                request=request
            )

            product.delete()
            messages.success(request, "Product deleted successfully.")
            return redirect('product_list')

    return render(request, 'product/delete_product.html', {'product': product})


@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def product_history_report(request):
    history = ProductHistory.objects.select_related('product', 'user').order_by('-date')
    return render(request, 'product/product_history_report.html', {'history': history})


@login_required(login_url='login')
def upload_products_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            overwrite = form.cleaned_data['overwrite_existing']

            try:
                # Read the Excel file
                df = pd.read_excel(excel_file)

                # Validate required columns
                required_columns = ['brand', 'price', 'category']
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    messages.error(request, f"Missing required columns: {', '.join(missing_columns)}")
                    return render(request, 'product/upload_products.html', {'form': form})

                # Helper function to get valid choices for validation
                def get_valid_categories():
                    return [choice[0] for choice in ProductChoices.CATEGORY_CHOICES]

                def get_valid_colors():
                    valid_colors = []
                    for family in ProductChoices.COLOR_CHOICES:
                        if isinstance(family[1], tuple):
                            valid_colors.extend([choice[0] for choice in family[1]])
                        else:
                            valid_colors.append(family[0])
                    return valid_colors

                def get_valid_designs():
                    valid_designs = []
                    for family in ProductChoices.DESIGN_CHOICES:
                        if isinstance(family[1], tuple):
                            valid_designs.extend([choice[0] for choice in family[1]])
                        else:
                            valid_designs.append(family[0])
                    return valid_designs

                # Process each row
                success_count = 0
                error_count = 0
                errors = []
                new_choices_added = {'colors': set(), 'designs': set(), 'categories': set()}
                saved_products = []   # Track products saved this upload for invoice creation
                invoice = None        # Will be created after the loop if any products succeed

                with transaction.atomic():
                    for index, row in df.iterrows():
                        try:
                            # Check if product with barcode already exists
                            barcode = row.get('barcode_number', None)
                            product = None

                            if barcode and pd.notna(barcode):
                                barcode = str(barcode).strip()
                                if barcode:  # Only check if barcode is not empty
                                    existing_product = Product.objects.filter(barcode_number=barcode).first()
                                    if existing_product:
                                        if overwrite:
                                            product = existing_product
                                        else:
                                            errors.append(
                                                f"Row {index + 2}: Product with barcode {barcode} already exists")
                                            error_count += 1
                                            continue

                            if not product:
                                product = Product()

                            # Map Excel columns to model fields with flexible validation
                            product.brand = str(row['brand']).strip() if pd.notna(row.get('brand')) else ''
                            if not product.brand:
                                errors.append(f"Row {index + 2}: Brand is required")
                                error_count += 1
                                continue

                            # Price validation
                            try:
                                product.price = float(row['price']) if pd.notna(row.get('price')) else 0.0
                                if product.price < 0:
                                    errors.append(f"Row {index + 2}: Price cannot be negative")
                                    error_count += 1
                                    continue
                            except (ValueError, TypeError):
                                errors.append(f"Row {index + 2}: Invalid price format")
                                error_count += 1
                                continue

                            # FLEXIBLE Category validation - accept existing or new values
                            category = str(row['category']).strip() if pd.notna(row.get('category')) else 'Apparel'
                            valid_categories = get_valid_categories()
                            if category not in valid_categories:
                                # Accept new category value and track it
                                new_choices_added['categories'].add(category)
                            product.category = category

                            # FLEXIBLE Color validation - accept existing or new values
                            if pd.notna(row.get('color')):
                                color = str(row['color']).strip()
                                valid_colors = get_valid_colors()
                                if color and color not in valid_colors:
                                    # Accept new color value and track it
                                    new_choices_added['colors'].add(color)
                                product.color = color if color else None
                            else:
                                product.color = None

                            # FLEXIBLE Design validation - accept existing or new values
                            if pd.notna(row.get('design')):
                                design = str(row['design']).strip()
                                valid_designs = get_valid_designs()
                                if design and design not in valid_designs:
                                    # Accept new design value and track it
                                    new_choices_added['designs'].add(design)
                                product.design = design if design else 'plain'
                            else:
                                product.design = 'plain'

                            product.size = str(row['size']).strip() if pd.notna(row.get('size')) else ''

                            # Quantity validation
                            try:
                                product.quantity = int(row['quantity']) if pd.notna(row.get('quantity')) else 0
                                if product.quantity < 0:
                                    errors.append(f"Row {index + 2}: Quantity cannot be negative")
                                    error_count += 1
                                    continue
                            except (ValueError, TypeError):
                                errors.append(f"Row {index + 2}: Invalid quantity format")
                                error_count += 1
                                continue

                            # Markup type validation (still strict as it affects calculations)
                            markup_type = str(row['markup_type']).strip() if pd.notna(
                                row.get('markup_type')) else 'percentage'
                            valid_markup_types = [choice[0] for choice in ProductChoices.MARKUP_TYPE_CHOICES]
                            if markup_type not in valid_markup_types:
                                errors.append(
                                    f"Row {index + 2}: Invalid markup_type '{markup_type}'. Valid options: {', '.join(valid_markup_types)}")
                                error_count += 1
                                continue
                            product.markup_type = markup_type

                            # Markup validation
                            try:
                                product.markup = float(row['markup']) if pd.notna(row.get('markup')) else 0.0
                                if product.markup < 0:
                                    errors.append(f"Row {index + 2}: Markup cannot be negative")
                                    error_count += 1
                                    continue
                            except (ValueError, TypeError):
                                errors.append(f"Row {index + 2}: Invalid markup format")
                                error_count += 1
                                continue

                            # Shop validation (still strict as it affects business logic)
                            shop = str(row['shop']).strip() if pd.notna(row.get('shop')) else 'STORE'
                            valid_shops = [choice[0] for choice in ProductChoices.SHOP_TYPE]
                            if shop not in valid_shops:
                                errors.append(
                                    f"Row {index + 2}: Invalid shop '{shop}'. Valid options: {', '.join(valid_shops)}")
                                error_count += 1
                                continue
                            product.shop = shop

                            # Handle optional barcode field
                            if pd.notna(row.get('barcode_number')):
                                barcode_value = str(row['barcode_number']).strip()
                                product.barcode_number = barcode_value if barcode_value else None
                            else:
                                product.barcode_number = None

                            # Save without full_clean() to bypass Django's choice validation
                            # We'll handle our own validation above
                            product.save()
                            saved_products.append(product)
                            success_count += 1

                        except Exception as e:
                            errors.append(f"Row {index + 2}: {str(e)}")
                            error_count += 1

                    # Create invoice for this upload (mirrors add_product behaviour)
                    if saved_products:
                        invoice = Invoice.objects.create(user=request.user)
                        for product in saved_products:
                            InvoiceProduct.objects.create(
                                invoice=invoice,
                                product_name=product.brand,
                                product_price=product.price,
                                product_color=product.color,
                                product_size=product.size,
                                product_category=product.category,
                                quantity=product.quantity,
                                total_price=product.price * product.quantity,
                            )
                        ActivityLog.log_activity(
                            user=request.user,
                            action='product_create',
                            description=f'Bulk uploaded {len(saved_products)} product(s) via Excel — Invoice #{invoice.invoice_number}',
                            model_name='Invoice',
                            object_id=invoice.id,
                            object_repr=f'Invoice #{invoice.invoice_number}',
                            request=request
                        )

                # Prepare results message
                if success_count > 0:
                    messages.success(request, f"Successfully uploaded {success_count} product(s).")
                    if invoice:
                        messages.success(request, f"Invoice #{invoice.invoice_number} created with {success_count} item(s).")

                # Show new choices that were added
                if any(new_choices_added.values()):
                    new_items = []
                    if new_choices_added['categories']:
                        new_items.append(f"Categories: {', '.join(new_choices_added['categories'])}")
                    if new_choices_added['colors']:
                        new_items.append(f"Colors: {', '.join(new_choices_added['colors'])}")
                    if new_choices_added['designs']:
                        new_items.append(f"Designs: {', '.join(new_choices_added['designs'])}")

                    messages.info(request, f"New choices added from Excel: {' | '.join(new_items)}")
                    messages.warning(request,
                                     "Note: These new choices are not in your predefined lists. Consider updating your model choices if you want them to appear in dropdowns.")

                if error_count > 0:
                    messages.warning(request, f"Failed to process {error_count} products. Check errors below.")
                    for error in errors[:10]:  # Limit to first 10 errors to avoid overwhelming
                        messages.error(request, error)
                    if len(errors) > 10:
                        messages.warning(request, f"... and {len(errors) - 10} more errors")

                if invoice:
                    return redirect('invoice_list')
                return redirect('product_list')

            except Exception as e:
                messages.error(request, f"Error processing Excel file: {str(e)}")
    else:
        form = ExcelUploadForm()

    return render(request, 'product/upload_products.html', {'form': form})


@login_required(login_url='login')
def download_excel_template(request):
    """Download Excel template with sample data and validation comments"""
    # Create a dataframe with comprehensive sample data
    sample_data = {
        'brand': ['Nike Air Max', 'Adidas Ultraboost', 'Puma Classic'],
        'price': [99.99, 129.99, 79.99],
        'color': ['black', 'navy', 'white'],
        'design': ['plain', 'geometric', 'solid'],
        'size': ['M', 'L', '42'],
        'category': ['Footwear', 'Footwear', 'Footwear'],
        'quantity': [25, 15, 30],
        'markup_type': ['percentage', 'fixed', 'percentage'],
        'markup': [20.0, 15.0, 25.0],
        'shop': ['STORE', 'STORE', 'STORE'],
        'barcode_number': ['', '1234567890123', '']
    }

    df = pd.DataFrame(sample_data)

    # Create Excel file in memory with formatting
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write main data
        df.to_excel(writer, sheet_name='Products', index=False, startrow=1)

        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Products']

        # Add title
        worksheet['A1'] = 'Product Upload Template - Replace sample data with your products'
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet['A1'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # Format headers
        header_row = 2
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=header_row, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Add validation sheet
        validation_data = {
            'Field': ['brand', 'price', 'color', 'design', 'size', 'category', 'quantity', 'markup_type', 'markup',
                      'shop', 'barcode_number'],
            'Required': ['Yes', 'Yes', 'No', 'No', 'No', 'Yes', 'No', 'No', 'No', 'No', 'No'],
            'Valid_Values': [
                'Any text',
                'Positive number',
                'black, navy, white, etc. (see Color Options sheet)',
                'plain, solid, geometric, etc. (see Design Options sheet)',
                'Any text (e.g., S, M, L, XL, or shoe sizes)',
                'Apparel, Footwear, Accessories',
                'Non-negative integer',
                'percentage, fixed',
                'Non-negative number',
                'STORE',
                'Optional unique barcode number'
            ]
        }

        validation_df = pd.DataFrame(validation_data)
        validation_df.to_excel(writer, sheet_name='Field_Validation', index=False)

        # Format validation sheet
        val_sheet = writer.sheets['Field_Validation']
        for col_num, column_title in enumerate(validation_df.columns, 1):
            cell = val_sheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

        # Add color options sheet
        color_options = []
        for family in Product.COLOR_CHOICES:
            if isinstance(family[1], tuple):
                for color in family[1]:
                    color_options.append({'Family': family[0], 'Color_Code': color[0], 'Display_Name': color[1]})

        color_df = pd.DataFrame(color_options)
        color_df.to_excel(writer, sheet_name='Color_Options', index=False)

        # Add design options sheet
        design_options = []
        for family in Product.DESIGN_CHOICES:
            if isinstance(family[1], tuple):
                for design in family[1]:
                    design_options.append({'Family': family[0], 'Design_Code': design[0], 'Display_Name': design[1]})

        design_df = pd.DataFrame(design_options)
        design_df.to_excel(writer, sheet_name='Design_Options', index=False)

        # Auto-adjust column widths
        for sheet_name in writer.sheets:
            sheet = writer.sheets[sheet_name]
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Prepare response
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=product_upload_template.xlsx'
    return response


@login_required(login_url='login')
def export_products_excel(request):
    """Export filtered products to Excel"""
    # Get the same filters as in product_list view
    products = filter_products(request, Product.objects.all())

    # Apply all the same filters from product_list
    query = request.GET.get('search', '')
    category = request.GET.get('category', '')
    shop = request.GET.get('shop', '')
    size = request.GET.get('size', '')
    color = request.GET.get('color', '')
    design = request.GET.get('design', '')
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')
    min_quantity = request.GET.get('min_quantity', '')
    max_quantity = request.GET.get('max_quantity', '')

    # Apply filters (same logic as product_list)
    if query:
        products = products.filter(
            Q(brand__icontains=query) |
            Q(color__icontains=query) |
            Q(category__icontains=query) |
            Q(design__icontains=query) |
            Q(size__icontains=query)
        )
    if category:
        products = products.filter(category=category)
    if shop:
        products = products.filter(shop=shop)
    if size:
        products = products.filter(size__icontains=size)
    if color:
        products = products.filter(color=color)
    if design:
        products = products.filter(design=design)
    if min_price:
        products = products.filter(price__gte=min_price)
    if max_price:
        products = products.filter(price__lte=max_price)
    if min_quantity:
        products = products.filter(quantity__gte=min_quantity)
    if max_quantity:
        products = products.filter(quantity__lte=max_quantity)

    # Convert to DataFrame
    data = []
    for product in products:
        data.append({
            'Brand': product.brand,
            'Category': product.category,
            'Color': product.get_display_color() if product.color else '',
            'Design': product.get_display_design() if product.design else '',
            'Size': product.size,
            'Price': product.price,
            'Markup Type': product.get_markup_type_display(),
            'Markup': product.markup,
            'Selling Price': product.selling_price,
            'Quantity': product.quantity,
            'Shop': product.shop,
            'Barcode': product.barcode_number or '',
        })

    df = pd.DataFrame(data)

    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Products', index=False)

        # Format the Excel file
        workbook = writer.book
        worksheet = writer.sheets['Products']

        # Add title and timestamp
        worksheet.insert_rows(1, 2)
        worksheet['A1'] = f'Product Export - Generated on {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet['A2'] = f'Total Products: {len(data)}'
        worksheet['A2'].font = Font(bold=True)

        # Format headers (now in row 3)
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=3, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    output.seek(0)

    # Generate filename with timestamp
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f'products_export_{timestamp}.xlsx'

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@login_required(login_url='login')
def export_products_pdf(request):
    """Export filtered products to PDF"""
    # Get filtered products (same logic as product_list)
    products = filter_products(request, Product.objects.all())

    # Apply additional filters
    query = request.GET.get('search', '')
    category = request.GET.get('category', '')
    shop = request.GET.get('shop', '')
    size = request.GET.get('size', '')
    color = request.GET.get('color', '')
    design = request.GET.get('design', '')
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')
    min_quantity = request.GET.get('min_quantity', '')
    max_quantity = request.GET.get('max_quantity', '')

    if query:
        products = products.filter(
            Q(brand__icontains=query) |
            Q(color__icontains=query) |
            Q(category__icontains=query) |
            Q(design__icontains=query) |
            Q(size__icontains=query)
        )
    if category:
        products = products.filter(category=category)
    if shop:
        products = products.filter(shop=shop)
    if size:
        products = products.filter(size__icontains=size)
    if color:
        products = products.filter(color=color)
    if design:
        products = products.filter(design=design)
    if min_price:
        products = products.filter(price__gte=min_price)
    if max_price:
        products = products.filter(price__lte=max_price)
    if min_quantity:
        products = products.filter(quantity__gte=min_quantity)
    if max_quantity:
        products = products.filter(quantity__lte=max_quantity)

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=50,
        bottomMargin=50,
        title="Product Inventory Report"
    )

    styles = getSampleStyleSheet()
    story = []

    # --- Title ---
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue,
        fontName='Helvetica-Bold'
    )
    story.append(Paragraph("Product Inventory Report", title_style))
    story.append(Spacer(1, 6))

    # --- Summary ---
    total_products = products.count()
    total_value = sum(p.price * p.quantity for p in products)
    total_selling_value = sum(p.selling_price * p.quantity for p in products)

    summary_data = [
        ['Report Generated:', timezone.now().strftime("%Y-%m-%d %H:%M:%S")],
        ['Total Products:', str(total_products)],
        ['Total Inventory Value (Cost):', f'{total_value:,.2f}'],
        ['Total Inventory Value (Selling):', f'{total_selling_value:,.2f}'],
    ]

    summary_table = Table(summary_data, colWidths=[2.8 * inch, 3.0 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 24))

    # --- Products Table ---
    if products.exists():
        # Prepare table data with headers
        data = [['Brand', 'Category', 'Size', 'Color', 'Design', 'Price', 'Qty', 'Sell Price', 'Total Value']]

        for product in products:  # Removed limit
            total_value = product.selling_price * product.quantity
            data.append([
                Paragraph(str(product.brand), styles['Normal']),
                Paragraph(str(product.category), styles['Normal']),
                Paragraph(str(product.size), styles['Normal']),
                Paragraph(str(product.color), styles['Normal']),
                Paragraph(str(product.design), styles['Normal']),
                Paragraph(f'{product.price:,.2f}', styles['Normal']),
                Paragraph(str(product.quantity), styles['Normal']),
                Paragraph(f'{product.selling_price:,.2f}', styles['Normal']),
                Paragraph(f'{total_value:,.2f}', styles['Normal']),
            ])

        # Calculate available width
        available_width = A4[0] - 80  # page width minus margins
        col_widths = [
            available_width * 0.18,  # Brand
            available_width * 0.14,  # Category
            available_width * 0.10,  # Size
            available_width * 0.10,  # Color
            available_width * 0.12,  # Design
            available_width * 0.09,  # Price
            available_width * 0.07,  # Qty
            available_width * 0.10,  # Sell Price
            available_width * 0.10,  # Total Value
        ]

        table = Table(data, colWidths=col_widths, repeatRows=1)

        # Style the table
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, 0), 8),

            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 1), (-1, -1), 6),

            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),

            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ]))

        story.append(table)

    else:
        no_products = Paragraph("No products found matching the current filters.", styles['Italic'])
        story.append(no_products)

    # --- Build PDF ---
    try:
        doc.build(story)
    except Exception as e:
        logger.error(f"PDF generation error: {e}")
        return HttpResponse("Error generating PDF.", status=500)

    buffer.seek(0)

    # Generate filename
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    filename = f'products_report_{timestamp}.pdf'

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    buffer.close()

    return response
