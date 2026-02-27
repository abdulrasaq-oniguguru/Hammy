# Standard library
import io
import json
import logging
import sys
import time
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO

# Third-party libraries
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors as rl_colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
)
from weasyprint import HTML
import win32print

# Django imports
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.db import models
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.db.models import (
    Q, F, Sum, Avg, Count, FloatField, DecimalField, ExpressionWrapper
)
from django.db.models.functions import (
    Coalesce, TruncMonth, TruncWeek, TruncDay
)
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.utils.timezone import make_aware

# Local app imports
from ..choices import ProductChoices
from ..models import (
    Product, Customer, Sale, Receipt, Payment, PaymentMethod, Delivery,
    ActivityLog, StoreConfiguration, LoyaltyConfiguration, TaxConfiguration,
    PartialPayment, PrinterTaskMapping,
)
from .auth import is_md, is_cashier, is_superuser, user_required_access

logger = logging.getLogger(__name__)

# Alias colors to avoid conflict with local colors variable in dashboard
colors = rl_colors

@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def sales_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    payment_method_filter = request.GET.get('payment_method')
    shop_type = request.GET.get('shop_type')
    export_format = request.GET.get('export')

    # Base queryset with select/prefetch
    sales = Sale.objects.select_related(
        'product', 'receipt', 'payment', 'delivery', 'customer'
    ).prefetch_related(
        'payment__payment_methods'
    ).order_by('-sale_date')

    # Apply filters - default to today if no dates provided
    start_date_obj = None
    end_date_obj = None
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            # Include the entire end date by adding 1 day and subtracting 1 second
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            sales = sales.filter(sale_date__range=[start_date_obj, end_date_obj])
        except ValueError:
            start_date_obj = None
            end_date_obj = None
    else:
        # Default to today's date if no filters applied
        today = timezone.now().date()
        start_date_obj = datetime.combine(today, datetime.min.time())
        end_date_obj = datetime.combine(today, datetime.max.time())
        sales = sales.filter(sale_date__range=[start_date_obj, end_date_obj])
        start_date = today.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    if shop_type:
        sales = sales.filter(product__shop=shop_type)

    if payment_method_filter:
        # Filter by specific payment method
        sales = sales.filter(payment__payment_methods__payment_method=payment_method_filter).distinct()

    # Group sales by receipt
    grouped_sales = defaultdict(list)
    for sale in sales:
        grouped_sales[sale.receipt].append(sale)

    # Convert to list of tuples for easy template iteration
    grouped_sales = sorted(
        grouped_sales.items(),
        key=lambda x: x[1][0].sale_date if x[1] else timezone.now(),
        reverse=True
    )

    # Calculate total sales using receipt.total_with_delivery (avoid double counting)
    # Get unique receipts from filtered sales
    unique_receipts = Receipt.objects.filter(sales__in=sales).distinct()
    total_sales = sum(receipt.total_with_delivery for receipt in unique_receipts if receipt.total_with_delivery)

    payment_methods = PaymentMethod.PAYMENT_METHODS
    shop_types = ProductChoices.SHOP_TYPE

    if export_format:
        if export_format == 'excel':
            return export_to_excel(grouped_sales, start_date_obj, end_date_obj, total_sales)
        elif export_format == 'pdf':
            return export_to_pdf(grouped_sales, start_date_obj, end_date_obj, total_sales)

    # Calculate discount and delivery totals for the report
    total_payment_discounts = 0
    total_line_discounts = 0
    total_delivery_fees = 0

    for receipt, sale_list in grouped_sales:
        # Payment discounts
        first_sale = sale_list[0] if sale_list else None
        if first_sale and first_sale.payment and first_sale.payment.discount_amount:
            total_payment_discounts += first_sale.payment.discount_amount

        # Line discounts
        for sale in sale_list:
            if sale.discount_amount:
                total_line_discounts += sale.discount_amount

        # Delivery fees
        if receipt and receipt.delivery_cost:
            total_delivery_fees += receipt.delivery_cost

    return render(request, 'sales/sales_report.html', {
        'grouped_sales': grouped_sales,
        'start_date': start_date,
        'end_date': end_date,
        'total_sales': total_sales,
        'payment_methods': payment_methods,
        'selected_payment_method': payment_method_filter,
        'shop_types': shop_types,
        'selected_shop_type': shop_type,
        'total_payment_discounts': total_payment_discounts,
        'total_line_discounts': total_line_discounts,
        'total_delivery_fees': total_delivery_fees,
        'total_all_discounts': total_payment_discounts + total_line_discounts,
        'currency_symbol': '₦'
    })


def export_to_pdf(grouped_sales, start_date, end_date, total_sales):
    buffer = io.BytesIO()
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch

    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(1 * inch, height - 1 * inch, "Sales Report")

    # Date range
    p.setFont("Helvetica", 12)
    date_text = "All Dates"
    if start_date and end_date:
        date_text = f"From {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    p.drawString(1 * inch, height - 1.2 * inch, date_text)

    # Table headers
    data = [['Receipt', 'Item', 'Qty', 'Item Total', 'Subtotal/Disc/Deliv', 'Payment', 'Date']]

    # Track which rows are receipt headers
    receipt_header_rows = []  # Will store row indices

    row_index = 1  # Start after header
    for receipt, sale_list in grouped_sales:
        sale = sale_list[0]

        # Get receipt totals
        subtotal = receipt.subtotal if receipt.subtotal else 0
        discount = sale.payment.discount_amount if sale.payment and sale.payment.discount_amount else 0
        delivery = receipt.delivery_cost if receipt.delivery_cost else 0
        receipt_total = receipt.total_with_delivery if receipt.total_with_delivery else 0

        # Add receipt header row
        data.append([
            f"Receipt #{receipt.receipt_number}",
            "",
            "",
            "",
            f"Total: ₦{receipt_total:.2f}",
            "",
            sale.sale_date.strftime("%m/%d %H:%M")
        ])
        receipt_header_rows.append(row_index)
        row_index += 1

        # Add each item
        for item in sale_list:
            payment_method = ""
            if item.payment:
                payment_methods = item.payment.payment_methods.all()
                payment_method = ", ".join([pm.get_payment_method_display() for pm in payment_methods])

            data.append([
                "",
                item.product.brand,
                str(item.quantity),
                f"₦{item.total_price:.2f}",
                "",  # Breakdown only on summary rows
                payment_method,
                ""
            ])
            row_index += 1

        # Add receipt breakdown rows
        data.append(["", "", "", "Subtotal:", f"₦{subtotal:.2f}", "", ""])
        row_index += 1

        if discount > 0:
            discount_pct = sale.payment.discount_percentage if sale.payment else 0
            data.append(["", "", "", f"Discount ({discount_pct}%):", f"-₦{discount:.2f}", "", ""])
            row_index += 1

        if delivery > 0:
            data.append(["", "", "", "Delivery:", f"₦{delivery:.2f}", "", ""])
            row_index += 1

        # Add blank spacer row
        data.append(["", "", "", "", "", "", ""])
        row_index += 1

    # Total row
    data.append([
        "TOTAL SALES",
        "",
        "",
        "",
        f"₦{total_sales:.2f}",
        "",
        ""
    ])
    row_index += 1

    # Create table with updated column widths
    # [Receipt, Item, Qty, Item Total, Receipt Total, Payment, Date]
    col_widths = [90, 120, 35, 70, 70, 90, 65]
    table = Table(data, colWidths=col_widths, repeatRows=1)

    # Build style commands
    style_commands = [
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ]

    # Highlight receipt header rows
    for row in receipt_header_rows:
        style_commands.append(('FONTNAME', (0, row), (6, row), 'Helvetica-Bold'))
        style_commands.append(('BACKGROUND', (0, row), (6, row), colors.beige))
        style_commands.append(('BOTTOMPADDING', (0, row), (6, row), 6))

    # Apply all styles
    table.setStyle(TableStyle(style_commands))

    # Position table
    table.wrapOn(p, width, height)
    y = height - 2 * inch
    table.drawOn(p, 0.5 * inch, y - table._height)

    # Finalize
    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


def export_to_excel(grouped_sales, start_date, end_date, total_sales):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    headers = ['Receipt', 'Date', 'Customer', 'Product', 'Qty', 'Item Total', 'Subtotal', 'Discount', 'Delivery', 'Total', 'Payment Methods', 'Status']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    for receipt, sale_list in grouped_sales:
        # Get first sale for metadata
        sale = sale_list[0]

        # Get receipt breakdown
        subtotal = float(receipt.subtotal) if receipt.subtotal else 0
        discount = float(sale.payment.discount_amount) if sale.payment and sale.payment.discount_amount else 0
        discount_pct = float(sale.payment.discount_percentage) if sale.payment and sale.payment.discount_percentage else 0
        delivery = float(receipt.delivery_cost) if receipt.delivery_cost else 0
        receipt_total = float(receipt.total_with_delivery) if receipt.total_with_delivery else 0

        # Format payment methods
        if sale.payment:
            payment_text = ", ".join([
                f"{pm.get_payment_method_display()}(₦{pm.amount:.2f})"
                for pm in sale.payment.payment_methods.all()
            ])
            status = sale.payment.get_payment_status_display()
        else:
            payment_text = "No Payment"
            status = "N/A"

        customer_name = sale.customer.name if sale.customer else "N/A"

        # Write each item in the receipt
        for idx, item in enumerate(sale_list):
            ws.cell(row=row_num, column=1, value=f"#{receipt.receipt_number}")
            ws.cell(row=row_num, column=2, value=item.sale_date.strftime("%Y-%m-%d %H:%M"))
            ws.cell(row=row_num, column=3, value=customer_name)
            ws.cell(row=row_num, column=4, value=item.product.brand)
            ws.cell(row=row_num, column=5, value=item.quantity)
            ws.cell(row=row_num, column=6, value=float(item.total_price))  # Individual item total
            ws.cell(row=row_num, column=7, value=subtotal if idx == 0 else "")  # Subtotal only on first row
            ws.cell(row=row_num, column=8, value=f"-{discount} ({discount_pct}%)" if idx == 0 and discount > 0 else "")  # Discount
            ws.cell(row=row_num, column=9, value=delivery if idx == 0 else "")  # Delivery
            ws.cell(row=row_num, column=10, value=receipt_total if idx == 0 else "")  # Total
            ws.cell(row=row_num, column=11, value=payment_text if idx == 0 else "")
            ws.cell(row=row_num, column=12, value=status if idx == 0 else "")

            row_num += 1

        # Add blank row between receipts
        row_num += 1

    # Add total sales
    ws.cell(row=row_num + 1, column=6, value="Total Sales:")
    ws.cell(row=row_num + 1, column=7, value=float(total_sales))
    ws.cell(row=row_num + 1, column=6).font = Font(bold=True)
    ws.cell(row=row_num + 1, column=7).font = Font(bold=True)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Create HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def discount_report(request):
    """
    Discount Report - Shows all sales with discounts applied (both payment and line-level)
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Get all payments with discounts (payment-level discounts)
    payments_with_discounts = Payment.objects.filter(
        discount_amount__gt=0
    ).prefetch_related('sale_set', 'sale_set__product', 'sale_set__receipt')

    # Get all sales with line discounts
    sales_with_line_discounts = Sale.objects.filter(
        discount_amount__gt=0
    ).select_related('product', 'receipt', 'payment').prefetch_related('payment__payment_methods')

    # Apply date filters - default to today if no dates provided
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            payments_with_discounts = payments_with_discounts.filter(payment_date__range=[start_date_obj, end_date_obj])
            sales_with_line_discounts = sales_with_line_discounts.filter(sale_date__range=[start_date_obj, end_date_obj])
        except ValueError:
            pass
    else:
        # Default to today's date if no filters applied
        today = timezone.now().date()
        start_date_obj = datetime.combine(today, datetime.min.time())
        end_date_obj = datetime.combine(today, datetime.max.time())
        payments_with_discounts = payments_with_discounts.filter(payment_date__range=[start_date_obj, end_date_obj])
        sales_with_line_discounts = sales_with_line_discounts.filter(sale_date__range=[start_date_obj, end_date_obj])
        start_date = today.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    # Calculate payment-level discount totals
    payment_discount_total = payments_with_discounts.aggregate(total=Sum('discount_amount'))['total'] or 0
    payment_transactions = payments_with_discounts.count()

    # Calculate line-level discount totals
    line_discount_total = sales_with_line_discounts.aggregate(total=Sum('discount_amount'))['total'] or 0
    line_transactions = sales_with_line_discounts.count()

    # Total discounts (both types)
    total_discount_amount = payment_discount_total + line_discount_total
    total_transactions = payment_transactions + line_transactions

    context = {
        'payments_with_discounts': payments_with_discounts,
        'sales_with_line_discounts': sales_with_line_discounts,
        'payment_discount_total': payment_discount_total,
        'line_discount_total': line_discount_total,
        'total_discount_amount': total_discount_amount,
        'payment_transactions': payment_transactions,
        'line_transactions': line_transactions,
        'total_transactions': total_transactions,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'reports/discount_report.html', context)


@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def delivery_report(request):
    """
    Delivery Report - Shows all receipts with delivery fees
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Get all receipts with delivery costs
    receipts = Receipt.objects.filter(
        delivery_cost__gt=0
    ).select_related('customer', 'user').prefetch_related('sales', 'sales__product')

    # Apply date filters - default to today if no dates provided
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
            receipts = receipts.filter(date__range=[start_date_obj, end_date_obj])
        except ValueError:
            pass
    else:
        # Default to today's date if no filters applied
        today = timezone.now().date()
        start_date_obj = datetime.combine(today, datetime.min.time())
        end_date_obj = datetime.combine(today, datetime.max.time())
        receipts = receipts.filter(date__range=[start_date_obj, end_date_obj])
        start_date = today.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    # Calculate totals
    total_delivery_fees = receipts.aggregate(total=Sum('delivery_cost'))['total'] or 0
    total_deliveries = receipts.count()
    avg_delivery_fee = total_delivery_fees / total_deliveries if total_deliveries > 0 else 0

    context = {
        'receipts': receipts,
        'total_delivery_fees': total_delivery_fees,
        'total_deliveries': total_deliveries,
        'avg_delivery_fee': avg_delivery_fee,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'reports/delivery_report.html', context)


@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def financial_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    report_type = request.GET.get('report_type', 'revenue')

    sales = Sale.objects.select_related('payment', 'product', 'delivery').prefetch_related('payment__payment_methods')

    # Apply date filters - default to today if no dates provided
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            sales = sales.filter(sale_date__range=[start_date_obj, end_date_obj])
        except (ValueError, TypeError):
            pass
    else:
        # Default to today's date if no filters applied
        today = timezone.now().date()
        start_date_obj = datetime.combine(today, datetime.min.time())
        end_date_obj = datetime.combine(today, datetime.max.time())
        sales = sales.filter(sale_date__range=[start_date_obj, end_date_obj])
        start_date = today.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')

    # Get unique payments to avoid double counting
    unique_payments = Payment.objects.filter(sale__in=sales).distinct()

    # ENHANCED FINANCIAL METRICS

    # 1. Revenue Analysis
    gross_revenue = sales.aggregate(
        total=Sum(ExpressionWrapper(F('product__selling_price') * F('quantity'), output_field=DecimalField()))
    )['total'] or Decimal('0')
    total_revenue = unique_payments.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')

    # 2. Discount Analysis
    total_item_discounts = sales.aggregate(total=Sum('discount_amount'))['total'] or Decimal('0')
    total_payment_discounts = unique_payments.aggregate(total=Sum('discount_amount'))['total'] or Decimal('0')
    total_discounts = total_item_discounts + total_payment_discounts

    # 3. Delivery Fee Analysis — deduplicate via distinct Delivery objects to avoid
    #    double-counting when multiple sales share the same delivery record.
    total_delivery_fees = (
        Delivery.objects.filter(sale__in=sales)
        .distinct()
        .aggregate(total=Sum('delivery_cost'))['total'] or Decimal('0')
    )

    # 4. Cost and Profit Analysis
    total_cost = sales.aggregate(
        total=Sum(ExpressionWrapper(F('product__price') * F('quantity'), output_field=DecimalField()))
    )['total'] or Decimal('0')
    net_revenue = total_revenue - total_delivery_fees  # Revenue without delivery fees
    total_profit = net_revenue - total_cost
    profit_margin = (total_profit / net_revenue * 100) if net_revenue > 0 else 0

    # 5. DETAILED PAYMENT METHOD BREAKDOWN
    payment_method_breakdown = {}
    payment_method_stats = {}

    for payment in unique_payments:
        for pm in payment.payment_methods.filter(status='completed'):
            method_name = pm.get_payment_method_display()

            if method_name not in payment_method_breakdown:
                payment_method_breakdown[method_name] = {
                    'total_amount': 0,
                    'transaction_count': 0,
                    'avg_transaction': 0
                }

            payment_method_breakdown[method_name]['total_amount'] += float(pm.amount)
            payment_method_breakdown[method_name]['transaction_count'] += 1

    # Calculate averages
    for method, data in payment_method_breakdown.items():
        if data['transaction_count'] > 0:
            data['avg_transaction'] = data['total_amount'] / data['transaction_count']

    # Sort by total amount
    payment_method_breakdown = dict(sorted(
        payment_method_breakdown.items(),
        key=lambda x: x[1]['total_amount'],
        reverse=True
    ))

    # 6. Payment Status Analysis
    payment_status_breakdown = {
        'completed': unique_payments.filter(payment_status='completed').count(),
        'partial': unique_payments.filter(payment_status='partial').count(),
        'pending': unique_payments.filter(payment_status='pending').count(),
        'failed': unique_payments.filter(payment_status='failed').count(),
    }

    completed_amount = unique_payments.filter(payment_status='completed').aggregate(
        total=Sum('total_amount'))['total'] or Decimal('0')
    partial_amount = unique_payments.filter(payment_status='partial').aggregate(
        total=Sum('total_paid'))['total'] or Decimal('0')
    pending_amount = unique_payments.filter(payment_status='pending').aggregate(
        total=Sum('total_amount'))['total'] or Decimal('0')

    # 7. Monthly Revenue Trend with detailed breakdown
    monthly_data = []
    monthly_raw = (
        Payment.objects
        .filter(sale__in=sales)
        .annotate(month=TruncMonth('sale__sale_date'))
        .values('month')
        .annotate(
            revenue=Sum('total_amount'),
            discount_amount=Sum('discount_amount'),
            transaction_count=Count('id', distinct=True)
        )
        .order_by('month')
    )

    for month in monthly_raw:
        month_sales = sales.filter(sale_date__month=month['month'].month, sale_date__year=month['month'].year)
        delivery_fees = (
            Delivery.objects.filter(sale__in=month_sales)
            .distinct()
            .aggregate(total=Sum('delivery_cost'))['total'] or Decimal('0')
        )

        monthly_data.append({
            'month': month['month'],
            'revenue': month['revenue'],
            'discount_amount': month['discount_amount'] or 0,
            'delivery_fees': delivery_fees,
            'transaction_count': month['transaction_count'],
            'net_revenue': (month['revenue'] or 0) - delivery_fees
        })

    # 8. Discount Analysis by Type
    discount_breakdown = {
        'item_level_discounts': total_item_discounts,
        'payment_level_discounts': total_payment_discounts,
        'total_discounts': total_discounts,
        'discount_rate': (total_discounts / gross_revenue * 100) if gross_revenue > 0 else 0
    }

    context = {
        # Revenue Metrics
        'gross_revenue': gross_revenue,
        'total_revenue': total_revenue,
        'net_revenue': net_revenue,
        'total_cost': total_cost,
        'total_profit': total_profit,
        'profit_margin': profit_margin,

        # Discount & Fees
        'discount_breakdown': discount_breakdown,
        'total_delivery_fees': total_delivery_fees,

        # Payment Analysis
        'payment_method_breakdown': payment_method_breakdown,
        'payment_status_breakdown': payment_status_breakdown,
        'completed_amount': completed_amount,
        'partial_amount': partial_amount,
        'pending_amount': pending_amount,

        # Trends
        'monthly_data': monthly_data,

        # Filters
        'start_date': start_date,
        'end_date': end_date,
        'report_type': report_type,

        # Summary Stats
        'total_transactions': unique_payments.count(),
        'avg_transaction_value': total_revenue / unique_payments.count() if unique_payments.count() > 0 else 0,
    }

    # Check for export format
    export_format = request.GET.get('export')
    if export_format == 'excel':
        return export_financial_to_excel(context, start_date, end_date)
    elif export_format == 'pdf':
        return export_financial_to_pdf(context, start_date, end_date)

    return render(request, 'reports/financial_report.html', context)


def export_financial_to_excel(context, start_date=None, end_date=None):
    """Export financial report data to Excel (CSV format)"""
    response = HttpResponse(content_type='text/csv')
    filename = f"financial_report_{start_date or 'all'}_to_{end_date or 'now'}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # Report Title
    writer.writerow(['FINANCIAL REPORT'])
    if start_date and end_date:
        writer.writerow([f'Date Range: {start_date} to {end_date}'])
    writer.writerow([])

    # Summary Section
    writer.writerow(['SUMMARY METRICS'])
    writer.writerow(['Metric', 'Value'])
    writer.writerow(['Gross Revenue', f"₦{context['gross_revenue']:,.2f}"])
    writer.writerow(['Net Revenue', f"₦{context['net_revenue']:,.2f}"])
    writer.writerow(['Total Profit', f"₦{context['total_profit']:,.2f}"])
    writer.writerow(['Profit Margin', f"{context['profit_margin']:.2f}%"])
    writer.writerow(['Total Transactions', context['total_transactions']])
    writer.writerow(['Avg Transaction Value', f"₦{context['avg_transaction_value']:,.2f}"])
    writer.writerow(['Total Delivery Fees', f"₦{context['total_delivery_fees']:,.2f}"])
    writer.writerow([])

    # Payment Method Breakdown
    writer.writerow(['PAYMENT METHOD BREAKDOWN'])
    writer.writerow(['Method', 'Total Amount', 'Transactions', 'Average Amount'])
    for method, data in context['payment_method_breakdown'].items():
        writer.writerow([
            method,
            f"₦{data['total_amount']:,.2f}",
            data['transaction_count'],
            f"₦{data['avg_transaction']:,.2f}"
        ])
    writer.writerow([])

    # Discount Analysis
    writer.writerow(['DISCOUNT ANALYSIS'])
    writer.writerow(['Type', 'Amount'])
    db = context['discount_breakdown']
    writer.writerow(['Item-Level Discounts', f"₦{db['item_level_discounts']:,.2f}"])
    writer.writerow(['Payment-Level Discounts', f"₦{db['payment_level_discounts']:,.2f}"])
    writer.writerow(['Total Discounts', f"₦{db['total_discounts']:,.2f}"])
    writer.writerow(['Discount Rate', f"{db['discount_rate']:.2f}%"])
    writer.writerow([])

    # Payment Status
    writer.writerow(['PAYMENT STATUS BREAKDOWN'])
    writer.writerow(['Status', 'Count', 'Amount'])
    psb = context['payment_status_breakdown']
    writer.writerow(['Completed', psb['completed'], f"₦{context['completed_amount']:,.2f}"])
    writer.writerow(['Partial', psb['partial'], f"₦{context['partial_amount']:,.2f}"])
    writer.writerow(['Pending', psb['pending'], f"₦{context['pending_amount']:,.2f}"])
    writer.writerow(['Failed', psb['failed'], "₦0.00"])
    writer.writerow([])

    # Monthly Trends
    writer.writerow(['MONTHLY FINANCIAL TRENDS'])
    writer.writerow(['Month', 'Revenue', 'Discounts', 'Delivery Fees', 'Net Revenue', 'Transactions'])
    for month in context['monthly_data']:
        writer.writerow([
            month['month'].strftime('%b %Y'),
            f"₦{month['revenue']:,.2f}",
            f"₦{month['discount_amount']:,.2f}",
            f"₦{month['delivery_fees']:,.2f}",
            f"₦{month['net_revenue']:,.2f}",
            month['transaction_count']
        ])

    return response


def export_financial_to_pdf(context, start_date=None, end_date=None):
    """Generate PDF for financial report"""
    # Create a file-like buffer to receive PDF data
    buffer = io.BytesIO()

    # Create the PDF object
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=30
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=12,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    normal_style = styles['Normal']

    # Title
    title = "Financial Report"
    if start_date and end_date:
        title += f" ({start_date} to {end_date})"
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))

    # Summary Stats
    summary_data = [
        ['Metric', 'Value'],
        ['Gross Revenue', f"₦{context['gross_revenue']:,.2f}"],
        ['Net Revenue', f"₦{context['net_revenue']:,.2f}"],
        ['Total Profit', f"₦{context['total_profit']:,.2f}"],
        ['Profit Margin', f"{context['profit_margin']:.1f}%"],
        ['Total Transactions', str(context['total_transactions'])],
        ['Avg Transaction Value', f"₦{context['avg_transaction_value']:,.2f}"],
        ['Total Delivery Fees', f"₦{context['total_delivery_fees']:,.2f}"],
    ]

    summary_table = Table(summary_data, colWidths=[200, 150])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
    ]))
    elements.append(Paragraph("Summary", subtitle_style))
    elements.append(summary_table)
    elements.append(Spacer(1, 24))

    # Payment Method Breakdown
    if context['payment_method_breakdown']:
        pm_data = [['Payment Method', 'Total Amount', 'Transactions', 'Avg Amount']]
        for method, data in context['payment_method_breakdown'].items():
            pm_data.append([
                method,
                f"₦{data['total_amount']:,.2f}",
                str(data['transaction_count']),
                f"₦{data['avg_transaction']:,.2f}"
            ])

        pm_table = Table(pm_data, colWidths=[150, 120, 100, 120])
        pm_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (1, 1), (3, -1), 'RIGHT'),
        ]))
        elements.append(Paragraph("Payment Method Breakdown", subtitle_style))
        elements.append(pm_table)
        elements.append(Spacer(1, 24))

    # Discount Analysis
    discount_data = [
        ['Discount Type', 'Amount'],
        ['Item-Level Discounts', f"₦{context['discount_breakdown']['item_level_discounts']:,.2f}"],
        ['Payment-Level Discounts', f"₦{context['discount_breakdown']['payment_level_discounts']:,.2f}"],
        ['Total Discounts', f"₦{context['discount_breakdown']['total_discounts']:,.2f}"],
        ['Discount Rate', f"{context['discount_breakdown']['discount_rate']:.1f}%"],
    ]

    discount_table = Table(discount_data, colWidths=[200, 150])
    discount_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
    ]))
    elements.append(Paragraph("Discount Analysis", subtitle_style))
    elements.append(discount_table)
    elements.append(Spacer(1, 24))

    # Payment Status
    status_data = [
        ['Status', 'Count', 'Amount'],
        ['Completed', str(context['payment_status_breakdown']['completed']), f"₦{context['completed_amount']:,.2f}"],
        ['Partial', str(context['payment_status_breakdown']['partial']), f"₦{context['partial_amount']:,.2f}"],
        ['Pending', str(context['payment_status_breakdown']['pending']), f"₦{context['pending_amount']:,.2f}"],
        ['Failed', str(context['payment_status_breakdown']['failed']), "₦0.00"],
    ]

    status_table = Table(status_data, colWidths=[100, 80, 120])
    status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
    ]))
    elements.append(Paragraph("Payment Status Overview", subtitle_style))
    elements.append(status_table)
    elements.append(Spacer(1, 24))

    # Monthly Data
    if context['monthly_data']:
        monthly_data = [['Month', 'Revenue', 'Discounts', 'Delivery Fees', 'Net Revenue', 'Transactions']]
        for month in context['monthly_data']:
            monthly_data.append([
                month['month'].strftime('%b %Y'),
                f"₦{month['revenue']:,.2f}",
                f"₦{month['discount_amount']:,.2f}",
                f"₦{month['delivery_fees']:,.2f}",
                f"₦{month['net_revenue']:,.2f}",
                str(month['transaction_count'])
            ])

        monthly_table = Table(monthly_data, colWidths=[80, 100, 100, 100, 100, 80])
        monthly_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (1, 1), (4, -1), 'RIGHT'),
        ]))
        elements.append(Paragraph("Monthly Financial Trends", subtitle_style))
        elements.append(monthly_table)

    # Build PDF
    doc.build(elements)

    # FileResponse sets the Content-Disposition header
    buffer.seek(0)
    filename = f"financial_report_{start_date or 'all'}_to_{end_date or 'now'}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response  # ✅ Added missing return statement
    return response


@login_required(login_url='login')
@user_passes_test(lambda u: u.is_superuser or (hasattr(u, 'profile') and u.profile.access_level in ['md', 'accountant', 'admin']), login_url='access_denied')
def tax_report(request):
    """Tax Report - Show tax breakdown for all receipts"""
    import json
    from datetime import datetime, timedelta
    from django.db.models import Sum, Q

    # Date filtering
    filter_type = request.GET.get('filter', 'today')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    now = timezone.now()

    # Set date range based on filter
    if filter_type == 'today':
        start_date = now.date()
        end_date = now.date()
    elif filter_type == 'yesterday':
        start_date = (now - timedelta(days=1)).date()
        end_date = (now - timedelta(days=1)).date()
    elif filter_type == 'this_week':
        start_date = (now - timedelta(days=now.weekday())).date()
        end_date = now.date()
    elif filter_type == 'this_month':
        start_date = now.replace(day=1).date()
        end_date = now.date()
    elif filter_type == 'last_month':
        last_month = now.replace(day=1) - timedelta(days=1)
        start_date = last_month.replace(day=1).date()
        end_date = last_month.date()
    elif filter_type == 'custom' and start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        start_date = now.date()
        end_date = now.date()

    # Get receipts within date range
    receipts = Receipt.objects.filter(
        date__date__gte=start_date,
        date__date__lte=end_date
    ).order_by('-date')

    # Calculate tax summary
    tax_summary = {}
    total_sales = Decimal('0')
    total_tax_collected = Decimal('0')
    inclusive_tax_total = Decimal('0')
    exclusive_tax_total = Decimal('0')

    receipt_details = []

    for receipt in receipts:
        if receipt.tax_details:
            try:
                tax_data = json.loads(receipt.tax_details) if isinstance(receipt.tax_details, str) else receipt.tax_details

                receipt_info = {
                    'receipt': receipt,
                    'total': receipt.total_with_delivery,
                    'tax_amount': receipt.tax_amount,
                    'taxes': []
                }

                for tax_code, tax_info in tax_data.items():
                    tax_amount = Decimal(str(tax_info['amount']))
                    tax_method = tax_info['method']

                    # Add to summary by tax code
                    if tax_code not in tax_summary:
                        tax_summary[tax_code] = {
                            'name': tax_info['name'],
                            'rate': tax_info['rate'],
                            'type': tax_info.get('type', 'percentage'),
                            'inclusive_amount': Decimal('0'),
                            'exclusive_amount': Decimal('0'),
                            'total_amount': Decimal('0'),
                            'count': 0
                        }

                    tax_summary[tax_code]['total_amount'] += tax_amount
                    tax_summary[tax_code]['count'] += 1

                    if tax_method == 'inclusive':
                        tax_summary[tax_code]['inclusive_amount'] += tax_amount
                        inclusive_tax_total += tax_amount
                    else:
                        tax_summary[tax_code]['exclusive_amount'] += tax_amount
                        exclusive_tax_total += tax_amount

                    receipt_info['taxes'].append({
                        'name': tax_info['name'],
                        'code': tax_code,
                        'rate': tax_info['rate'],
                        'amount': tax_amount,
                        'method': tax_method
                    })

                total_sales += receipt.total_with_delivery
                total_tax_collected += receipt.tax_amount
                receipt_details.append(receipt_info)
            except (json.JSONDecodeError, KeyError, TypeError) as e:
                logger.error(f"Error parsing tax details for receipt {receipt.receipt_number}: {e}")
                continue

    # Get active tax configurations for reference
    active_taxes = TaxConfiguration.get_active_taxes()

    context = {
        'filter_type': filter_type,
        'start_date': start_date,
        'end_date': end_date,
        'tax_summary': tax_summary,
        'receipt_details': receipt_details,
        'total_sales': total_sales,
        'total_tax_collected': total_tax_collected,
        'inclusive_tax_total': inclusive_tax_total,
        'exclusive_tax_total': exclusive_tax_total,
        'active_taxes': active_taxes,
        'receipts_count': receipts.count(),
    }

    return render(request, 'Reports/tax_report.html', context)


@user_passes_test(lambda u: u.is_superuser or u.is_md, login_url='access_denied')
@login_required(login_url='login')
def reports_dashboard(request):
    # Get date filter from GET parameters
    filter_type = request.GET.get('filter', 'today')
    now = timezone.now()
    start_date = None
    end_date = now

    # Define date ranges
    if filter_type == 'today':
        start_date = now.date()
        sales = Sale.objects.filter(sale_date__date=start_date)
    elif filter_type == 'this_week':
        start_date = now - timedelta(days=now.weekday())
        sales = Sale.objects.filter(sale_date__gte=start_date)
    elif filter_type == 'this_month':
        start_date = now.replace(day=1)
        sales = Sale.objects.filter(sale_date__gte=start_date)
    elif filter_type == 'this_year':
        start_date = now.replace(month=1, day=1)
        sales = Sale.objects.filter(sale_date__gte=start_date)
    elif filter_type == 'custom':
        start_str = request.GET.get('start_date')
        end_str = request.GET.get('end_date')
        if start_str and end_str:
            try:
                start_date = timezone.datetime.strptime(start_str, '%Y-%m-%d').date()
                end_date = timezone.datetime.strptime(end_str, '%Y-%m-%d').date()
                sales = Sale.objects.filter(sale_date__date__range=[start_date, end_date])
            except:
                start_date = now.date()
                sales = Sale.objects.filter(sale_date__date=start_date)
        else:
            start_date = now.date()
            sales = Sale.objects.filter(sale_date__date=start_date)
    else:
        start_date = now.date()
        sales = Sale.objects.filter(sale_date__date=start_date)

    # Prefetch related payment methods
    sales = sales.select_related('payment', 'product', 'customer', 'delivery').prefetch_related(
        'payment__payment_methods'
    )

    # Use Payment.total_amount and avoid double counting
    unique_payments = Payment.objects.filter(sale__in=sales).distinct()
    total_revenue = sum(payment.total_amount for payment in unique_payments if payment.total_amount)

    total_sales_count = sales.count()

    # Average Order Value (AOV) - use unique receipts to avoid inflating AOV
    unique_receipts = sales.values('receipt').distinct().count()
    avg_order_value = total_revenue / unique_receipts if unique_receipts > 0 else 0

    # Counts
    total_products = Product.objects.count()
    total_customers = Customer.objects.count()
    low_stock_products = Product.objects.filter(quantity__lt=10).count()
    low_stock_items = Product.objects.filter(quantity__lt=10).order_by('quantity')[:5]

    # Recent sales (with enhanced payment method info)
    recent_sales = []
    for sale in sales.order_by('-sale_date')[:5]:
        if sale.payment:
            methods = []
            for pm in sale.payment.payment_methods.all():
                methods.append(f"{pm.get_payment_method_display()} ({pm.amount:.2f})")
            payment_summary = ", ".join(methods) or "Unknown"
            payment_status = sale.payment.get_payment_status_display()
            total_paid = sale.payment.total_paid
            balance_due = sale.payment.balance_due
        else:
            payment_summary = "No Payment"
            payment_status = "N/A"
            total_paid = 0
            balance_due = 0

        recent_sales.append({
            'sale': sale,
            'payment_summary': payment_summary,
            'payment_status': payment_status,
            'total_paid': total_paid,
            'balance_due': balance_due,
        })

    # Monthly revenue trend - use Payment.total_amount
    monthly_data = (
        Payment.objects
        .filter(sale__sale_date__gte=now - timedelta(days=365))
        .annotate(month=TruncMonth('sale__sale_date'))
        .values('month')
        .annotate(
            revenue=Sum('total_amount'),
            discount_amount=Sum('discount_amount'),
            transaction_count=Count('id', distinct=True)
        )
        .order_by('month')[:12]
    )

    # Enhanced Category performance with payment data
    category_performance = {}
    for payment in unique_payments:
        for sale in payment.sale_set.all():
            category = sale.product.category
            if category not in category_performance:
                category_performance[category] = {
                    'product__category': category,
                    'total_revenue': 0,
                    'total_units': 0,
                    'total_discount': 0
                }

            # Proportional split of payment across items
            payment_per_sale = payment.total_amount / payment.sale_set.count() if payment.sale_set.count() > 0 else 0
            discount_per_sale = (
                                            payment.discount_amount or 0) / payment.sale_set.count() if payment.sale_set.count() > 0 else 0

            category_performance[category]['total_revenue'] += payment_per_sale
            category_performance[category]['total_units'] += sale.quantity
            category_performance[category]['total_discount'] += discount_per_sale

    category_performance = sorted(category_performance.values(), key=lambda x: x['total_revenue'], reverse=True)

    # Enhanced Top Selling Products with payment data
    product_revenue = {}
    for payment in unique_payments:
        sales_in_payment = payment.sale_set.all()
        payment_per_sale = payment.total_amount / sales_in_payment.count() if sales_in_payment.count() > 0 else 0

        for sale in sales_in_payment:
            key = (sale.product.brand, sale.product.category)
            if key not in product_revenue:
                product_revenue[key] = {'total_qty': 0, 'total_rev': 0}

            product_revenue[key]['total_qty'] += sale.quantity
            product_revenue[key]['total_rev'] += payment_per_sale

    top_products = []
    for (brand, category), data in product_revenue.items():
        top_products.append({
            'product__brand': brand,
            'product__category': category,
            'total_qty': data['total_qty'],
            'total_rev': data['total_rev']
        })

    top_products = sorted(top_products, key=lambda x: x['total_qty'], reverse=True)[:5]

    # New Customers by Month
    new_customers = (
        Customer.objects
        .filter(created_at__gte=timezone.now() - timedelta(days=365))
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    # Enhanced Payment Status Summary
    all_payments = Payment.objects.filter(sale__in=sales).distinct()
    completed_payments = all_payments.filter(payment_status='completed').count()
    partial_payments = all_payments.filter(payment_status='partial').count()
    pending_payments = all_payments.filter(payment_status='pending').count()
    failed_payments = all_payments.filter(payment_status='failed').count()

    # Enhanced Payment Method Breakdown with statistics
    payment_method_breakdown = {}
    for payment in unique_payments:
        for pm in payment.payment_methods.filter(status='completed'):
            method_name = pm.get_payment_method_display()
            if method_name not in payment_method_breakdown:
                payment_method_breakdown[method_name] = {
                    'total_amount': 0,
                    'transaction_count': 0,
                    'avg_amount': 0
                }

            payment_method_breakdown[method_name]['total_amount'] += float(pm.amount)
            payment_method_breakdown[method_name]['transaction_count'] += 1

    # Calculate averages and sort
    for method, data in payment_method_breakdown.items():
        if data['transaction_count'] > 0:
            data['avg_amount'] = data['total_amount'] / data['transaction_count']

    payment_method_breakdown = dict(sorted(
        payment_method_breakdown.items(),
        key=lambda x: x[1]['total_amount'],
        reverse=True
    ))

    # Delivery Analysis
    delivery_stats = {
        'total_delivery_fees': sum(s.delivery.delivery_cost or 0 for s in sales if s.delivery),
        'delivery_orders': sales.filter(delivery__isnull=False).count(),
        'pickup_orders': sales.filter(delivery__delivery_option='pickup').count(),
        'delivery_orders_count': sales.filter(delivery__delivery_option='delivery').count(),
    }

    # Discount Analysis - both payment-level and line-level
    # Payment-level discounts
    payment_discounts_total = sum(p.discount_amount or 0 for p in unique_payments)
    payment_discounts_count = unique_payments.filter(discount_amount__gt=0).count()

    # Line-level discounts
    line_discounts_total = sum(s.discount_amount or 0 for s in sales if s.discount_amount and s.discount_amount > 0)
    line_discounts_count = sales.filter(discount_amount__gt=0).count()

    discount_stats = {
        'total_discounts': payment_discounts_total + line_discounts_total,
        'payment_discounts': payment_discounts_total,
        'line_discounts': line_discounts_total,
        'discount_transactions': payment_discounts_count + line_discounts_count,
    }

    # Chart colors
    colors = ['#4e73df', '#1cc88a', '#36b9cc', '#f6c23e', '#e74a3b', '#858796', '#5a5c69']
    for i, cat in enumerate(category_performance):
        cat['color'] = colors[i % len(colors)]
        cat['hover_color'] = colors[(i + 2) % len(colors)]

    context = {
        'today': timezone.now().date(),
        'filter_type': filter_type,
        'start_date': start_date,
        'end_date': end_date,
        'total_revenue': total_revenue,
        'total_sales_count': total_sales_count,
        'avg_order_value': round(avg_order_value, 2),
        'total_products': total_products,
        'total_customers': total_customers,
        'low_stock_products': low_stock_products,
        'low_stock_items': low_stock_items,
        'recent_sales': recent_sales,
        'monthly_data': monthly_data,
        'category_performance': category_performance,
        'top_products': top_products,
        'new_customers': new_customers,
        'completed_payments': completed_payments,
        'discount_stats': discount_stats,
        'partial_payments': partial_payments,
        'pending_payments': pending_payments,
        'failed_payments': failed_payments,
        'payment_method_breakdown': payment_method_breakdown,
        'delivery_stats': delivery_stats,
    }

    return render(request, 'reports/dashboard.html', context)


@user_passes_test(is_md, login_url='access_denied')
@login_required(login_url='login')
def inventory_report(request):
    category = request.GET.get('category')
    low_stock = request.GET.get('low_stock')

    products = Product.objects.all()

    if category:
        products = products.filter(category=category)

    if low_stock:
        products = products.filter(quantity__lt=10)

    # Calculate enhanced inventory metrics via DB aggregation (avoids full queryset load)
    totals = products.aggregate(
        total_value=Sum(ExpressionWrapper(F('selling_price') * F('quantity'), output_field=DecimalField())),
        total_cost_value=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
    )
    total_value = totals['total_value'] or Decimal('0')
    total_cost_value = totals['total_cost_value'] or Decimal('0')
    potential_profit = total_value - total_cost_value

    # Additional inventory statistics
    low_stock_count = products.filter(quantity__lt=10).count()
    critical_stock_count = products.filter(quantity__lt=5).count()
    avg_markup = products.aggregate(
        avg_markup=models.Avg('markup')
    )['avg_markup'] or 0

    # Annotate products with additional fields for template
    products = products.annotate(
        total_value=models.F('selling_price') * models.F('quantity')
    )

    context = {
        'products': products,
        'categories': ProductChoices.CATEGORY_CHOICES,
        'total_value': total_value,
        'total_cost_value': total_cost_value,
        'potential_profit': potential_profit,
        'selected_category': category,
        'show_low_stock': bool(low_stock),
        'low_stock_count': low_stock_count,
        'critical_stock_count': critical_stock_count,
        'avg_markup': avg_markup,
    }

    return render(request, 'reports/inventory_report.html', context)

'''========================BARCODE PRINTS=============================='''

@csrf_exempt
@require_POST
def print_multiple_barcodes_directly(request):
    """Print multiple barcodes directly to thermal printer with individual quantities"""
    try:
        data = json.loads(request.body)
        products_data = data.get('products', [])  # [{product_id: 1, quantity: 3}, ...]

        if not products_data:
            return JsonResponse({
                'success': False,
                'error': 'No products specified for printing'
            })

        # Resolve barcode printer: task mapping → barcode PrinterConfiguration → session/OS default
        barcode_mapping_printer = PrinterTaskMapping.get_printer_for_task('barcode_label')
        if barcode_mapping_printer:
            printer_name = barcode_mapping_printer.system_printer_name
            printer_source = 'task_mapping'
        else:
            from ..models import PrinterConfiguration as PC
            barcode_config = PC.objects.filter(printer_type='barcode', is_active=True).first()
            if barcode_config:
                printer_name = barcode_config.system_printer_name
                printer_source = 'barcode_config'
            else:
                printer_name = request.session.get('selected_printer') or win32print.GetDefaultPrinter()
                printer_source = 'fallback'

        if not printer_name:
            return JsonResponse({
                'success': False,
                'error': 'No barcode printer configured. Go to Printer Settings and assign a Barcode Printer.',
            })

        results = []
        total_printed = 0

        for item in products_data:
            product_id = item.get('product_id')
            quantity = item.get('quantity', 1)

            try:
                product = get_object_or_404(Product, id=product_id)

                # Ensure barcode is generated
                if not product.barcode_image or not product.barcode_number:
                    product.generate_barcode()
                    product.save()

                # Get the path to the barcode image
                barcode_path = product.barcode_image.path

                # Print the required number of copies
                copies_printed = 0
                _views = sys.modules['store.views']
                for i in range(quantity):
                    success = _views.print_image(printer_name, barcode_path)
                    if success:
                        copies_printed += 1
                        total_printed += 1

                    # Small delay between copies
                    time.sleep(0.3)

                results.append({
                    'product_id': product_id,
                    'product_name': product.brand,
                    'requested_quantity': quantity,
                    'printed_quantity': copies_printed,
                    'success': copies_printed == quantity
                })

            except Exception as e:
                logger.error(f"Error printing barcode for product {product_id}: {str(e)}")
                results.append({
                    'product_id': product_id,
                    'product_name': f'Product {product_id}',
                    'requested_quantity': quantity,
                    'printed_quantity': 0,
                    'success': False,
                    'error': str(e)
                })

        # Check success rate
        successful_products = sum(1 for result in results if result['success'])
        total_products = len(results)

        # Surface first failure reason as top-level error when nothing printed
        top_error = None
        if successful_products == 0:
            failed = [r for r in results if not r['success'] and r.get('error')]
            if failed:
                top_error = failed[0]['error']
            else:
                top_error = f"Print job sent to '{printer_name}' but 0 copies confirmed. Check the printer is online and the name is correct."

        return JsonResponse({
            'success': successful_products > 0,
            'message': f'Printed {total_printed} barcodes for {successful_products}/{total_products} products',
            'total_printed': total_printed,
            'successful_products': successful_products,
            'total_products': total_products,
            'results': results,
            'printer_name': printer_name,
            'printer_source': printer_source,
            **(({'error': top_error}) if top_error else {}),
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        })
    except Exception as e:
        logger.error(f"Multiple direct print failed: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Multiple direct print failed: {str(e)}'
        })


@csrf_exempt
@require_POST
def print_single_barcode_directly(request, product_id):
    """Print a single barcode with specified quantity"""
    try:
        data = json.loads(request.body)
        quantity = data.get('quantity', 1)

        product = get_object_or_404(Product, id=product_id)

        # Ensure barcode is generated
        if not product.barcode_image or not product.barcode_number:
            product.generate_barcode()
            product.save()

        # Resolve barcode printer: task mapping → barcode PrinterConfiguration → session/OS default
        barcode_mapping_printer = PrinterTaskMapping.get_printer_for_task('barcode_label')
        if barcode_mapping_printer:
            printer_name = barcode_mapping_printer.system_printer_name
        else:
            from ..models import PrinterConfiguration as PC
            barcode_config = PC.objects.filter(printer_type='barcode', is_active=True).first()
            if barcode_config:
                printer_name = barcode_config.system_printer_name
            else:
                printer_name = request.session.get('selected_printer') or win32print.GetDefaultPrinter()

        if not printer_name:
            return JsonResponse({
                'success': False,
                'error': 'No barcode printer configured. Go to Printer Settings and assign a Barcode Printer.',
            })

        # Get the path to the barcode image
        barcode_path = product.barcode_image.path

        # Print the required number of copies
        copies_printed = 0
        _views = sys.modules['store.views']
        for i in range(quantity):
            success = _views.print_image(printer_name, barcode_path)
            if success:
                copies_printed += 1

            # Small delay between copies
            time.sleep(0.3)

        return JsonResponse({
            'success': copies_printed > 0,
            'message': f'Printed {copies_printed}/{quantity} copies of barcode for {product.brand}',
            'product_name': product.brand,
            'requested_quantity': quantity,
            'printed_quantity': copies_printed,
            'printer_name': printer_name
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        })
    except Exception as e:
        logger.error(f"Single direct print failed: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Single direct print failed: {str(e)}'
        })


def print_image(printer_name, image_path):
    """Print an image directly to a printer using Windows GDI"""
    try:
        from PIL import Image, ImageWin
        import win32ui

        # Open the image
        image = Image.open(image_path)

        # Convert to monochrome if needed (better for thermal printers)
        if image.mode != "1":
            image = image.convert("1")

        # Get the printer
        hprinter = win32print.OpenPrinter(printer_name)

        try:
            # Get printer properties
            printer_info = win32print.GetPrinter(hprinter, 2)

            # Create a device context for the printer
            hdc = win32ui.CreateDC()
            hdc.CreatePrinterDC(printer_name)

            # Start document
            hdc.StartDoc("Barcode Print")
            hdc.StartPage()

            # Calculate scaling to fit the page
            printable_area = hdc.GetDeviceCaps(110), hdc.GetDeviceCaps(111)  # HORZRES, VERTRES

            # Scale image to fit printable area
            img_width, img_height = image.size
            scaling_x = printable_area[0] / img_width
            scaling_y = printable_area[1] / img_height
            scaling = min(scaling_x, scaling_y)

            # Calculate position to center image
            x = (printable_area[0] - img_width * scaling) / 2
            y = (printable_area[1] - img_height * scaling) / 2

            # Draw the image
            dib = ImageWin.Dib(image)
            dib.draw(hdc.GetHandleOutput(), (
                int(x),
                int(y),
                int(x + img_width * scaling),
                int(y + img_height * scaling)
            ))

            # End document
            hdc.EndPage()
            hdc.EndDoc()

            return True

        finally:
            win32print.ClosePrinter(hprinter)

    except Exception as e:
        logger.error(f"Error printing image: {str(e)}")
        return False


# Activity Log Views


@login_required(login_url='login')
def activity_log_list(request):
    """
    Display a paginated list of activity logs with filtering options
    """
    # Get all logs ordered by most recent
    logs = ActivityLog.objects.all().select_related('user').order_by('-created_at')

    # Get filter parameters
    action_filter = request.GET.get('action', '')
    user_filter = request.GET.get('user', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    success_filter = request.GET.get('success', '')
    search_query = request.GET.get('search', '')

    # Apply filters
    if action_filter:
        logs = logs.filter(action=action_filter)

    if user_filter:
        logs = logs.filter(username=user_filter)

    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            logs = logs.filter(created_at__gte=date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            # Add one day to include the entire end date
            date_to_obj = date_to_obj + timedelta(days=1)
            logs = logs.filter(created_at__lt=date_to_obj)
        except ValueError:
            pass

    if success_filter:
        logs = logs.filter(success=(success_filter == 'true'))

    if search_query:
        logs = logs.filter(
            Q(description__icontains=search_query) |
            Q(username__icontains=search_query) |
            Q(ip_address__icontains=search_query) |
            Q(object_repr__icontains=search_query)
        )

    # Get unique values for filters
    action_choices = ActivityLog.ACTION_CHOICES
    unique_users = ActivityLog.objects.values_list('username', flat=True).distinct().order_by('username')

    # Pagination
    paginator = Paginator(logs, 50)  # Show 50 logs per page
    page = request.GET.get('page')

    try:
        logs = paginator.page(page)
    except PageNotAnInteger:
        logs = paginator.page(1)
    except EmptyPage:
        logs = paginator.page(paginator.num_pages)

    context = {
        'logs': logs,
        'action_choices': action_choices,
        'unique_users': unique_users,
        'action_filter': action_filter,
        'user_filter': user_filter,
        'date_from': date_from,
        'date_to': date_to,
        'success_filter': success_filter,
        'search_query': search_query,
    }

    return render(request, 'activity/activity_log_list.html', context)


@login_required(login_url='login')
def activity_log_detail(request, log_id):
    """
    Display detailed information about a specific activity log entry
    """
    log = get_object_or_404(ActivityLog, id=log_id)

    context = {
        'log': log,
    }

    return render(request, 'activity/activity_log_detail.html', context)


# ==================== LOYALTY PROGRAM ENDPOINTS ====================

@login_required(login_url='login')


@login_required
def gift_report(request):
    """Report on all items given as gifts"""
    from ..models import Sale
    from django.db.models import Sum, Count

    # Get date filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Get all gift sales
    gift_sales = Sale.objects.filter(is_gift=True).select_related(
        'product', 'customer', 'receipt'
    ).order_by('-sale_date')

    if start_date:
        gift_sales = gift_sales.filter(sale_date__gte=start_date)
    if end_date:
        gift_sales = gift_sales.filter(sale_date__lte=end_date)

    # Calculate statistics
    stats = gift_sales.aggregate(
        total_items=Count('id'),
        total_quantity=Sum('quantity'),
        total_value=Sum('original_value'),
    )

    context = {
        'gift_sales': gift_sales,
        'stats': stats,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'reports/gift_report.html', context)


