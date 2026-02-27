# Standard library
import io
import logging
from datetime import timedelta

# Third-party libraries
import openpyxl
from openpyxl.styles import Font
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

# Django imports
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

# Local app imports
from ..forms import (
    InvoiceForm, GoodsReceivedForm, DeliveryForm
)
from ..models import (
    Product, Invoice, InvoiceProduct, GoodsReceived, Delivery
)
from .auth import is_md, is_cashier, is_superuser, user_required_access

logger = logging.getLogger(__name__)


@login_required(login_url='login')
def invoice(request):
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        if form.is_valid():
            form.save()  # Now simply save the form without manually setting the user, since it's handled in add_product view
            return redirect('invoice_list')  # Redirect to the invoice list after saving
    else:
        form = InvoiceForm()

    return render(request, 'store/invoice.html', {'form': form})


@login_required(login_url='login')
def invoice_list(request):
    # Fetch all invoices, ordered by date descending
    invoices = Invoice.objects.select_related('user').all().order_by('-date')  # Using select_related for efficiency

    # Filter by invoice number if provided
    invoice_number = request.GET.get('invoice_number', '')
    if invoice_number:
        invoices = invoices.filter(invoice_number__icontains=invoice_number)

    # Filter by start date if provided
    start_date = request.GET.get('start_date', '')
    if start_date:
        invoices = invoices.filter(date__gte=start_date)

    # Filter by end date if provided
    end_date = request.GET.get('end_date', '')
    if end_date:
        invoices = invoices.filter(date__lte=end_date)

    return render(request, 'invoice/invoice_list.html', {'invoices': invoices})


def _check_duplicate_invoice(invoice):
    """
    Return the first invoice (created within the last 48 hours, excluding
    *invoice* itself) whose item set matches *invoice* item-for-item by
    product name, color, size, and category.  Returns None when no match.
    """
    cutoff = timezone.now() - timedelta(hours=48)
    new_items = frozenset(
        invoice.invoice_products.values_list(
            'product_name', 'product_color', 'product_size', 'product_category'
        )
    )
    if not new_items:
        return None
    for past in Invoice.objects.filter(date__gte=cutoff).exclude(pk=invoice.pk):
        past_items = frozenset(
            past.invoice_products.values_list(
                'product_name', 'product_color', 'product_size', 'product_category'
            )
        )
        if past_items == new_items:
            return past
    return None


@login_required(login_url='login')
def invoice_detail(request, pk):
    user = request.user
    invoice = get_object_or_404(Invoice, pk=pk)

    # Ensure the invoice has the current user set
    invoice_products = invoice.invoice_products.all()
    total_cost = sum(invoice_product.total_price for invoice_product in invoice_products)
    total_quantity = sum(invoice_product.quantity for invoice_product in invoice_products)  # Add this line

    return render(request, 'invoice/invoice_detail.html', {
        'invoice': invoice,
        'invoice_products': invoice_products,
        'total_cost': total_cost,
        'total_quantity': total_quantity,  # Add this line
        'user': user,
    })


@login_required(login_url='login')
def export_invoice_pdf(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    invoice_products = invoice.invoice_products.all()
    total_cost = sum(invoice_product.total_price for invoice_product in invoice_products)
    total_quantity = sum(invoice_product.quantity for invoice_product in invoice_products)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="invoice_{invoice.invoice_number}.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    # Add invoice details
    p.setFont("Helvetica-Bold", 16)
    p.drawString(1 * inch, height - 1 * inch, f"Invoice #{invoice.invoice_number}")
    p.setFont("Helvetica", 12)
    p.drawString(1 * inch, height - 1.25 * inch, f"Date: {invoice.date.strftime('%B %d, %Y')}")
    p.drawString(1 * inch, height - 1.5 * inch, f"Created by: {invoice.user.username}")

    # Add table headers
    y = height - 2 * inch
    p.setFont("Helvetica-Bold", 12)
    p.drawString(1 * inch, y, "Product")
    p.drawString(4 * inch, y, "Unit Price")
    p.drawString(5.5 * inch, y, "Quantity")
    p.drawString(6.5 * inch, y, "Total")

    # Add products
    y -= 0.2 * inch
    for item in invoice_products:
        # Build spec line for this item
        spec_parts = []
        if item.product_color:
            spec_parts.append(f"Color: {item.product_color}")
        if item.product_size:
            spec_parts.append(f"Size: {item.product_size}")
        if item.product_category:
            spec_parts.append(f"Category: {item.product_category}")
        spec_line = "  |  ".join(spec_parts)

        row_height = (0.42 if spec_line else 0.28) * inch
        y -= row_height
        if y < 1.2 * inch:  # Create new page if needed
            p.showPage()
            y = height - 1 * inch
            y -= row_height

        p.setFont("Helvetica", 11)
        p.setFillColorRGB(0, 0, 0)
        p.drawString(1 * inch, y, item.product_name)
        p.drawString(4 * inch, y, f"{item.product_price:.2f}")
        p.drawString(5.5 * inch, y, str(item.quantity))
        p.drawString(6.5 * inch, y, f"{item.total_price:.2f}")

        if spec_line:
            p.setFont("Helvetica-Oblique", 9)
            p.setFillColorRGB(0.42, 0.47, 0.56)
            p.drawString(1.05 * inch, y - 0.17 * inch, spec_line[:90])
            p.setFillColorRGB(0, 0, 0)

    # Add totals
    y -= 0.5 * inch
    p.setFont("Helvetica-Bold", 12)
    p.drawString(1 * inch, y, f"Total Quantity: {total_quantity}")
    p.drawString(4 * inch, y, f"Total Cost: {total_cost:.2f}")

    p.showPage()
    p.save()
    return response


@login_required(login_url='login')
def export_invoice_excel(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    invoice_products = invoice.invoice_products.all()
    total_cost = sum(invoice_product.total_price for invoice_product in invoice_products)
    total_quantity = sum(invoice_product.quantity for invoice_product in invoice_products)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="invoice_{invoice.invoice_number}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active

    # Sanitize sheet title by removing invalid characters
    sheet_title = f"Invoice {invoice.invoice_number}"
    # Replace invalid characters with underscores
    invalid_chars = ['/', '\\', '*', '?', ':', '[', ']']
    for char in invalid_chars:
        sheet_title = sheet_title.replace(char, '_')
    # Truncate if longer than 31 characters (Excel limit)
    sheet_title = sheet_title[:31]
    ws.title = sheet_title

    # Add headers
    headers = ['Product', 'Unit Price', 'Quantity', 'Total']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Add data
    for row_num, item in enumerate(invoice_products, 2):
        ws.cell(row=row_num, column=1).value = item.product_name
        ws.cell(row=row_num, column=2).value = float(item.product_price)
        ws.cell(row=row_num, column=3).value = item.quantity
        ws.cell(row=row_num, column=4).value = float(item.total_price)

    # Add totals
    last_row = len(invoice_products) + 3
    ws.cell(row=last_row, column=2).value = "Total Quantity:"
    ws.cell(row=last_row, column=3).value = total_quantity
    ws.cell(row=last_row + 1, column=2).value = "Total Cost:"
    ws.cell(row=last_row + 1, column=3).value = float(total_cost)

    # Format currency cells
    for row in ws.iter_rows(min_row=2, max_row=len(invoice_products) + 1, min_col=2, max_col=4):
        for cell in row:
            if cell.column in [2, 4]:  # Price columns
                cell.number_format = '""#,##0.00'

    wb.save(response)
    return response


@login_required(login_url='login')
def goods_received(request):
    if request.method == 'POST':
        form = GoodsReceivedForm(request.POST)
        if form.is_valid():
            goods_received = form.save()
            # Optional: Update product stock
            product = goods_received.product
            product.quantity += goods_received.quantity_received
            product.save()

            messages.success(request, f"✅ {goods_received.quantity_received} units of {product.brand} received (Batch: {goods_received.batch_number}).")
            return redirect('goods_received')
    else:
        form = GoodsReceivedForm()

    # Get recent receipts (last 10)
    recent_receipts = GoodsReceived.objects.select_related('product').order_by('-received_date')[:10]

    return render(request, 'store/goods_received.html', {
        'form': form,
        'recent_receipts': recent_receipts
    })


@login_required(login_url='login')
def delivery(request):
    if request.method == 'POST':
        form = DeliveryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('delivery')
    else:
        form = DeliveryForm()
    return render(request, 'store/delivery.html', {'form': form})
